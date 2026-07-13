# -*- coding: utf-8 -*-
"""
TEST 탭 데이터 자동 생성 파이프라인 (7/7 개정판 · 신규 파일 구조 대응)
==================================================================

입력 (uploads 폴더 · 접두 (EDW)/(EHUB)/(BI, SAP) 유무 모두 허용):
  · [BI, SAP] 단품별판매재고, 반응과.xlsx     (단품별판매재고 시트 + 반응과 시트)
  · [EHUB]    주문_내부_공홈.xlsx / 이랜드몰.xlsx / 카카오톡선물하기.xlsx  (3개 분리)
  · [EDW]     주문_외부(내부포함)_MUSINSA_단품명필요*.csv                   (무신사 주판)
  · [EDW]     주문_외부(내부포함)_NAVER_단품명필요*.csv                     (네이버 주판)
  · [EDW]     주문_내부_ZIGZAG_단품명필요.csv + 주문_외부_ZIGZAG_단품명필요.csv (합산 = 지그재그 주판)
  · [EHUB]    재고_내부_공홈.xlsx / 이랜드몰.xlsx / 카카오톡선물하기.xlsx
  · [EHUB]    재고_내부_네이버, 무신사, 지그재그.xlsx (or 지그재그, 네이버, 무신사)
  · [EDW]     재고_외부_MUSINSA_STOCK*.csv
  · [EDW]     재고_외부_NAVER_STOCK*.csv
  · [EDW]     재고_외부_ZIGZAG_STOCK*.csv  ← T+U+V 합산 (AVAIL_SALES_QTY + REQ_IN_QTY + WAIT_IN_QTY)

출력:
  data/test/data_spao_YYMMDD.csv.gz  (gzip 압축, mock_data가 자동 감지)

실행:
  cd rebal_web_v2.0_deploy && python3 build_test_csv.py

주요 규칙 (사용자 7/7 확정):
  ─ 단품코드: SP로 시작하는 15자리 (스타일 10자리 + 컬러 2자리 + 사이즈 3자리)
  ─ 신상 필터: 5번째 글자가 'G'인 코드만 유지 (대시보드 노출은 신상만)
  ─ 무신사 창고 (KR_MUSINSA_SHIP_QTY 컬럼 R열 사용)
      · 필터: USE_YN='true' AND STATUS='판매중'
      · 수량 컬럼: KR_MUSINSA_SHIP_QTY
  ─ 네이버 창고: ITEM_STATUS='양품' 필터 · USE_QTY 컬럼
  ─ 지그재그 창고: 필터 없음 (전체) · T(AVAIL_SALES_QTY)+U(REQ_IN_QTY)+V(WAIT_IN_QTY) 합산  ← 7/7 규칙 변경
  ─ 지그재그 주판: 내부(주문_내부_ZIGZAG) + 외부(주문_외부_ZIGZAG) 두 CSV 합산
  ─ 무신사/네이버 주판: 외부(내부포함) CSV 하나만 사용 (내·외부 자동 통합됨)
  ─ 재고 합산: inv_채널 = 내부재고 + 외부창고 (대시보드 뺄셈 로직 정합)
  ─ 반응과 재고: 음수 → 0 치환, 매장별 수량 합산
  ─ 누판/주판/출고율: 정수 반올림 후 100으로 나눔 (0~1 소수)
  ─ 정상가 fallback 우선순위:
      1) 단품별판매재고 시트 (발주액/발주량)
      2) 반응과 시트 (표준가/결판가)
      3) 채널 재고 파일들
      4) ZIGZAG 주문 (ORIGINAL_PRICE)
      5) MUSINSA 주문 (PRICE / 0.85로 정상가 역산)
      6) 스타일 10자리 그룹 평균가
  ─ 단품명 fallback: 마스터 → 반응과 시트 → 채널 재고 파일들
"""
from __future__ import annotations
import csv
import gzip
import json
import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("build_test_csv")

# ─────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────
_HERE = Path(__file__).parent
UPLOAD_DIR = Path("/sessions/quirky-dazzling-wozniak/mnt/uploads")   # Cowork sandbox 마운트 경로
OUT_DIR = Path("/sessions/quirky-dazzling-wozniak/mnt/outputs/reba_770_upload/data/test")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 오늘 날짜 YYMMDD (파일명용)
from datetime import date
_TODAY = date.today().strftime("%y%m%d")
OUT_PATH = OUT_DIR / f"data_spao_{_TODAY}.csv.gz"

CHANNELS = ["공홈", "이랜드몰", "무신사", "지그재그", "네이버", "카카오선물하기"]
EXT_CH = ("무신사", "지그재그", "네이버")


def _int(v, d=0):
    try:
        return 0 if v in (None, "") else int(float(v))
    except Exception:
        return d


def _num(v, d=0):
    try:
        return 0 if v in (None, "") else float(v)
    except Exception:
        return d


def _find_by_glob(pattern: str) -> Path:
    """와일드카드로 파일 찾기.
    (EDW) / (EHUB) / (BI, SAP) 등 접두 유무 모두 허용.
    사용자 7/9 fix: 접미 `-XXXX` (자동 rename 파일)도 매칭 대상에 포함 → 최신 mtime 우선.
    """
    candidates: list[Path] = []
    # 원본 패턴 (정확 매칭)
    candidates.extend(UPLOAD_DIR.glob(pattern))
    # 접두 있는 경우
    for prefix in ("(EDW) ", "(EHUB) ", "(BI, SAP) "):
        candidates.extend(UPLOAD_DIR.glob(f"{prefix}{pattern}"))
    # 접두 아무거나
    candidates.extend(UPLOAD_DIR.glob(f"* {pattern}"))
    # 접미 `-XXXX` 처리 (파일 재업로드 자동 rename 대응)
    # 예: "단품별판매재고, 반응과.xlsx" → "단품별판매재고, 반응과*.xlsx"도 매칭
    if pattern.endswith(".xlsx") or pattern.endswith(".csv"):
        stem, ext = pattern.rsplit(".", 1)
        if "*" not in stem:  # 이미 와일드카드 없을 때만
            wide_pat = f"{stem}*.{ext}"
            candidates.extend(UPLOAD_DIR.glob(wide_pat))
            for prefix in ("(EDW) ", "(EHUB) ", "(BI, SAP) "):
                candidates.extend(UPLOAD_DIR.glob(f"{prefix}{wide_pat}"))
    # 중복 제거 + 최신 mtime 우선
    uniq: dict[str, Path] = {p.name: p for p in candidates}
    matches = sorted(uniq.values(), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"파일 없음: {pattern}")
    return matches[0]


def _find_by_any(patterns: list[str]) -> Path:
    """여러 패턴 중 먼저 매칭되는 파일 반환 (파일명 변경 대응)."""
    for pat in patterns:
        try:
            return _find_by_glob(pat)
        except FileNotFoundError:
            continue
    raise FileNotFoundError(f"어떤 패턴에도 매칭 안 됨: {patterns}")


def _read_cp949_csv(path: Path):
    """EUC-KR/CP949 인코딩 CSV → dict rows (외부창고 3파일용)."""
    for enc in ("cp949", "utf-8-sig"):
        try:
            return list(csv.DictReader(path.read_text(encoding=enc).splitlines()))
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"인코딩 감지 실패: {path.name}")


def _read_utf8_csv(path: Path):
    """UTF-8 BOM CSV → dict rows (내부 재고 이랜드몰/카카오톡용)."""
    for enc in ("utf-8-sig", "cp949"):
        try:
            return list(csv.DictReader(path.read_text(encoding=enc).splitlines()))
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"인코딩 감지 실패: {path.name}")


def _iter_xlsx(path: Path, sheet: Optional[str] = None, skip_rows: int = 0):
    """xlsx → dict rows (헤더는 skip_rows 이후 첫 행)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    it = ws.iter_rows(values_only=True)
    for _ in range(skip_rows):
        next(it)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    for r in it:
        yield {hdr[i]: r[i] for i in range(min(len(hdr), len(r)))}
    wb.close()


# ─────────────────────────────────────────────
# Stage 1: 단품별판매재고 시트
# ─────────────────────────────────────────────
def stage1_master(skus: dict) -> None:
    log.info("Stage 1: 단품별판매재고 시트 로딩...")
    path = _find_by_any(["단품별판매재고, 반응과.xlsx"])
    # 캐싱: 동일 파일 재파싱 방지 (mtime + name 기반 pickle)
    import pickle
    cache_key = f"{path.name}_{int(path.stat().st_mtime)}.pkl"
    cache_path = _HERE / ".cache" / cache_key
    cache_path.parent.mkdir(exist_ok=True)
    if cache_path.exists():
        with open(cache_path, "rb") as f:
            cached = pickle.load(f)
        skus.update(cached)
        log.info(f"  → {len(cached):,} 단품 (cached)")
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["단품별판매재고"]
    count = 0
    # 헤더 구조: row 0=라벨('2시즌씩'), row 1=컬럼명, row 2=단위, row 3=전체결과, row 4~ 데이터
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        code = str(row[0] or "").strip().upper()
        if not code or len(code) < 10:
            continue
        # ============================================================
        # 사용자 7/9 확정 스펙 (내일부터 이 기준 그대로 적용):
        # · 누판율 = S열 (col 18) — 파일 %값 그대로
        # · 주판율 = T열 (col 19) — 파일 %값 그대로
        # · 출고율 = K열(출고량) / G열(누적입고량) * 100 — 파일에서 역산
        # · 주간외형매출 = P열 (col 15) — 기간 판매액[외형매출]
        # · 정상가 = D열(발주액[정상가]) / C열(발주량) 역산
        # ============================================================
        buy_qty = _num(row[2])   # col 2: 발주량 (C열)
        buy_amt = _num(row[3])   # col 3: 발주액[정상가] (D열)
        price = int(buy_amt / buy_qty) if buy_qty > 0 else 0
        in_qty = _int(row[6])    # col 6: 누적입고량 (G열)
        ship_qty = _int(row[10]) # col 10: 출고량 (K열)
        cum_qty = _int(row[12])  # col 12: 누판량 (M열)
        wk_qty = _int(row[13])   # col 13: 주판량 (N열)
        wk_sales = _num(row[15]) # col 15: 기간 판매액[외형매출] (P열)
        cum_rate = _num(row[18]) # col 18: 누판율 % (S열)
        wk_rate = _num(row[19])  # col 19: 주판율 % (T열)
        ship_rate = (ship_qty / in_qty * 100) if in_qty > 0 else 0  # K/G × 100
        skus[code] = {
            "단품명": str(row[1] or "").strip(),
            "정상가": price,
            "출고율": ship_rate,
            "누판율": cum_rate,
            "주판율": wk_rate,
            "주간외형매출": int(wk_sales),
            "in_qty": in_qty,
            "cum_qty": cum_qty,
            "wk_qty": wk_qty,
        }
        count += 1
    wb.close()
    with open(cache_path, "wb") as f:
        pickle.dump(dict(skus), f)
    log.info(f"  → {count:,} 단품")


# ─────────────────────────────────────────────
# Stage 2: 반응과 시트
# ─────────────────────────────────────────────
def stage2_bw(bw_qty: dict, bw_name: dict, price_fallback: dict) -> None:
    """반응과 시트 로딩.

    7/7 신규 스키마 (8 cols): [연도, 시즌, 스타일, 컬러, 사이즈, 상품(15자리), 내역, 수량]
      - 매장별 개별 컬럼이 사라짐 → 단일 수량 컬럼 사용
      - 정상가 컬럼이 없어짐 → price_fallback 은 다른 소스에서 채움
    옛 스키마(매장별 col 10~ + 표준가/결판가) 자동 감지 후 후방 호환.
    """
    log.info("Stage 2: 반응과 시트 로딩...")
    path = _find_by_any(["단품별판매재고, 반응과.xlsx"])
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["반응과"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        return

    def _find(*keys: str) -> Optional[int]:
        for i, h in enumerate(header):
            hs = str(h or "")
            for k in keys:
                if k in hs:
                    return i
        return None

    idx_code = _find("상품") if _find("상품") is not None else (
        5 if len(header) <= 12 else 6
    )
    idx_name = _find("내역", "상품명") if _find("내역", "상품명") is not None else (
        6 if len(header) <= 12 else 7
    )
    # 사용자 7/9 fix: 신규 파일 컬럼명이 "수량" → "합계"로 변경됨
    idx_qty_single = _find("수량", "합계")
    idx_price_std = _find("표준가")
    idx_price_kp = _find("결판가")
    is_new_schema = len(header) <= 12 or idx_qty_single is not None

    # 사용자 7/9 fix: 순차 max(0, ...) → 라인 순수 합산 후 최종 max(0, sum)
    # 이전 로직 결함: 음수 라인이 먼저 오면 max(0,...)로 리셋 후 양수만 누적 → 실제보다 큼
    # (1,033개 SKU 오차, 총 10,498장 초과 발생) SPYCG37C0119090 실제 -376장 → 옛 결과 0
    # 수정: 라인들 그대로 sum → 최종에 한 번만 max(0, sum) 적용
    for row in rows_iter:
        if not row:
            continue
        code = str(row[idx_code] or "").strip().upper() if idx_code < len(row) else ""
        if not code or len(code) < 10:
            continue

        if is_new_schema and idx_qty_single is not None and idx_qty_single < len(row):
            qsum = _int(row[idx_qty_single])
        else:
            qsum = sum(_int(v) for v in row[10:])
        # 순수 누적 (max 적용 안 함) — 각 라인 그대로 더함
        bw_qty[code] = bw_qty.get(code, 0) + qsum

        if idx_name is not None and idx_name < len(row) and row[idx_name] and code not in bw_name:
            bw_name[code] = str(row[idx_name]).strip()

        p = 0
        if idx_price_std is not None and idx_price_std < len(row):
            p = _int(row[idx_price_std])
        if p == 0 and idx_price_kp is not None and idx_price_kp < len(row):
            p = _int(row[idx_price_kp])
        if p > 0 and code not in price_fallback:
            price_fallback[code] = p

    # 최종 음수 → 0 clamp (모든 라인 합산 후 딱 한 번만 적용)
    for code in list(bw_qty.keys()):
        if bw_qty[code] < 0:
            bw_qty[code] = 0

    wb.close()
    log.info(f"  → 반응과 재고 {sum(bw_qty.values()):,}장 ({len(bw_qty):,} 단품)")


# ─────────────────────────────────────────────
# Stage 3: 내부 재고 6채널
# ─────────────────────────────────────────────
# (fname_patterns, site_map)
# - EHUB 파일은 모두 xlsx 로 통일 (예전 csv → xlsx 변경 대응)
# - 다중 채널 파일은 사이트명 컬럼값으로 분기
INTERNAL_SOURCES = [
    (["재고_내부_공홈.xlsx"], {"스파오 뉴 공홈": "공홈", "공홈": "공홈"}),
    (["재고_내부_이랜드몰.xlsx", "재고_내부_이랜드몰.csv"], {"이랜드몰": "이랜드몰"}),
    # 사용자 7/8 결정: 카카오톡선물하기 파일의 B열(사이트)에서 '카카오톡선물하기'만 추출.
    # 이전에는 '사방넷'도 카카오 재고에 포함시켰으나, 실제 카카오 채널 재고와 무관한 사이트라
    # 재고 과다 인식 → 재배치 로직에서 실재고보다 많은 이동 계산 유발. '사방넷' 매칭 제거.
    (["재고_내부_카카오톡선물하기.xlsx", "재고_내부_카카오톡선물하기.csv"],
     {"카카오톡선물하기": "카카오선물하기", "카카오선물하기": "카카오선물하기"}),
    ([
        "재고_내부_네이버, 무신사, 지그재그.xlsx",
        "재고_내부_지그재그, 네이버, 무신사.xlsx",
    ], {"무신사": "무신사", "네이버": "네이버", "지그재그": "지그재그"}),
]


def stage3_internal_inv(inv_int: dict, name_fallback: dict) -> None:
    log.info("Stage 3: 내부 재고 로딩...")
    for patterns, site_map in INTERNAL_SOURCES:
        path = _find_by_any(patterns)
        if path.suffix.lower() == ".xlsx":
            src = _iter_xlsx(path)
        else:
            src = _read_utf8_csv(path)
        cnt = 0
        for r in src:
            code = str(r.get("상품코드") or r.get("단품코드") or "").strip().upper()
            if not code or len(code) < 10:
                continue
            site = str(r.get("사이트") or "").strip()
            ch = None
            for k, v in site_map.items():
                if k in site or site in k:
                    ch = v
                    break
            if ch is None and len(site_map) == 1:
                ch = list(site_map.values())[0]
            if not ch:
                continue
            # 사용자 7/8 결정: EHUB 파일 재고 기준을 'M열 = 가용재고'로 고정.
            # 이전: 가용재고 → 총재고 fallback (총재고는 예약·이송 포함이라 실제 이동 가능
            # 재고보다 큼 → 재배치 로직에서 과다 이동 결과 유발). fallback 제거.
            qty = _int(r.get("가용재고") or 0)
            inv_int.setdefault(ch, {})
            inv_int[ch][code] = inv_int[ch].get(code, 0) + qty
            nm = str(r.get("상품명") or "").strip()
            if nm and code not in name_fallback:
                name_fallback[code] = nm
            cnt += 1
        log.info(f"  {path.name}: {cnt:,}행 처리")


# ─────────────────────────────────────────────
# Stage 4: 외부창고 3채널
# ─────────────────────────────────────────────
_MUS_OPT_RX = re.compile(r"\[(\w+)\][^\^]*\^\s*\w*\[(\w+)\]")


def stage4_ext_wh(inv_ext: dict, name_fallback: dict, price_fallback: dict) -> None:
    log.info("Stage 4: 외부창고 로딩...")

    # ── MUSINSA (7/9 신규: EDW NAVER_STOCK_★S열 파일 사용 · S열=LOCAL_AVAIL_QTY)
    # 사용자 7/9 확정: 무신사 재고 파일이 지그재그/네이버와 같은 EDW 형식으로 변경.
    # 파일명은 (EDW) 재고_내부_NAVER_STOCK_★S열.csv 이지만 실질 = 무신사 위탁창고 재고.
    # 컬럼: F열(col 5)=SUPPLIER_BARCODE 단품코드(15자리), S열(col 18)=LOCAL_AVAIL_QTY
    path = _find_by_any([
        "재고_내부_NAVER_STOCK_*S열*.csv",  # 7/9 신규 파일명
        "재고_내부_NAVER_STOCK_★S열*.csv",
        "재고_외부_MUSINSA_STOCK_*.csv",   # 이전 파일명 fallback
        "재고_외부_MUSINSA_STOCK*.csv",
    ])
    rows = _read_cp949_csv(path)
    per = defaultdict(int)
    used_new_schema = False
    for r in rows:
        # 신규 스키마: SUPPLIER_BARCODE (F열) + LOCAL_AVAIL_QTY (S열)
        code = str(r.get("SUPPLIER_BARCODE") or "").strip().upper()
        if code and len(code) >= 12:
            used_new_schema = True
            per[code] += _int(r.get("LOCAL_AVAIL_QTY", 0))
            nm = str(r.get("PRODUCT_NAME") or "").strip()
            if nm and code not in name_fallback:
                name_fallback[code] = nm
            continue
        # 옛 스키마 fallback: STYLE_NO + OPTION_TXT + KR_MUSINSA_SHIP_QTY
        if str(r.get("USE_YN", "")).strip().lower() != "true":
            continue
        if r.get("STATUS") != "판매중":
            continue
        style = str(r.get("STYLE_NO") or "").strip().upper()
        opt = str(r.get("OPTION_TXT") or "").strip()
        m = _MUS_OPT_RX.search(opt)
        if not (m and style):
            continue
        code = f"{style}{m.group(1).zfill(2)}{m.group(2).zfill(3)}"
        if len(code) < 12:
            continue
        per[code] += _int(r.get("KR_MUSINSA_SHIP_QTY", 0))
        nm = str(r.get("PROD_NM") or "").strip()
        if nm and code not in name_fallback:
            name_fallback[code] = nm
    inv_ext["무신사"] = dict(per)
    log.info(f"  MUSINSA ({'LOCAL_AVAIL_QTY(신규)' if used_new_schema else 'KR_MUSINSA_SHIP_QTY(옛)'}): {len(per):,}단품, 재고 {sum(per.values()):,}장 [파일: {path.name}]")

    # ── NAVER (ITEM_STATUS='양품' · USE_QTY)
    path = _find_by_any(["재고_외부_NAVER_STOCK_*.csv", "재고_외부_NAVER_STOCK*.csv"])
    rows = _read_cp949_csv(path)
    per = defaultdict(int)
    for r in rows:
        if r.get("ITEM_STATUS") != "양품":
            continue
        code = str(r.get("ITEM_CD") or "").strip().upper()
        if not code or code == "SPAO_UNKNOWN" or len(code) < 12:
            continue
        per[code] += _int(r.get("USE_QTY", 0))
        nm = str(r.get("ITEM_NM") or "").strip()
        if nm and code not in name_fallback:
            name_fallback[code] = nm
    inv_ext["네이버"] = dict(per)
    log.info(f"  NAVER: {len(per):,}단품, 재고 {sum(per.values()):,}장 [파일: {path.name}]")

    # ── ZIGZAG (7/7 규칙 변경: 필터 없음 · T+U+V 합산)
    #     T = AVAIL_SALES_QTY  (판매 가능 = 가용재고)
    #     U = REQ_IN_QTY       (입고 신청)
    #     V = WAIT_IN_QTY      (입고 대기)
    path = _find_by_any(["재고_외부_ZIGZAG_STOCK_*.csv", "재고_외부_ZIGZAG_STOCK*.csv"])
    rows = _read_cp949_csv(path)
    per: dict = defaultdict(int)
    for r in rows:
        code = str(r.get("PRODUCT_BARCODE_NO") or "").strip().upper()
        if not code or len(code) < 12:
            continue
        tuv = (
            _int(r.get("AVAIL_SALES_QTY", 0))
            + _int(r.get("REQ_IN_QTY", 0))
            + _int(r.get("WAIT_IN_QTY", 0))
        )
        per[code] += tuv
        nm = str(r.get("PRODUCT_NM") or "").strip()
        if nm and code not in name_fallback:
            name_fallback[code] = nm
        p = _int(r.get("SALES_PRICE", 0))
        if p > 0 and code not in price_fallback:
            price_fallback[code] = p
    inv_ext["지그재그"] = dict(per)
    log.info(
        f"  ZIGZAG (T+U+V, no filter): {len(per):,}단품, 재고 "
        f"{sum(per.values()):,}장 [파일: {path.name}]"
    )


# ─────────────────────────────────────────────
# Stage 5: 내부 주문 (공홈+이랜드몰+카카오)
# ─────────────────────────────────────────────
_CH_SITE_MAP = {"차세대공홈": "공홈", "이랜드몰": "이랜드몰",
                "카카오": "카카오선물하기", "사방넷": "카카오선물하기"}


def _parse_internal_order_xlsx(path: Path, per_ch: dict) -> int:
    """단일 EHUB 내부주문 xlsx 파싱. 첫 행이 곧 헤더 (7/7 신규 구조)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet0"] if "Sheet0" in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    hi = {h: i for i, h in enumerate(header)}
    # 7/7 파일은 첫 행이 헤더. 예전 파일은 첫 행이 비어있었음 → 헤더 미검출 시 한 줄 더 읽어 보정
    if "SAP단품코드" not in hi:
        header = [str(c).strip() if c is not None else "" for c in next(it)]
        hi = {h: i for i, h in enumerate(header)}
    cnt = 0
    for row in it:
        code = str(row[hi.get("SAP단품코드", 12)] or "").strip().upper()
        if not code or len(code) < 12:
            continue
        site = str(row[hi.get("거래채널", 10)] or "").strip()
        ch = None
        for k, v in _CH_SITE_MAP.items():
            if k in site:
                ch = v
                break
        if not ch:
            continue
        per_ch[ch][code] += _int(row[hi.get("실수량", 23)])
        cnt += 1
    wb.close()
    return cnt


def stage5_orders_internal(ord_int: dict) -> None:
    """내부 주문 로딩 · 각 xlsx 파일마다 별도 pickle 캐시.
    타임아웃 발생 시 다음 실행에서 캐시된 파일은 건너뜀."""
    log.info("Stage 5: 내부 주문 로딩...")
    import pickle
    cache_dir = _HERE / ".cache"
    cache_dir.mkdir(exist_ok=True)

    per_ch: dict = defaultdict(lambda: defaultdict(int))
    for existing_ch, mp in ord_int.items():
        for code, q in mp.items():
            per_ch[existing_ch][code] = q

    total = 0
    for patterns in (
        ["주문_내부_공홈.xlsx"],
        ["주문_내부_이랜드몰.xlsx"],
        ["주문_내부_카카오톡선물하기.xlsx"],
        ["주문_내부_공홈,이랜드몰,카카오톡.xlsx"],
    ):
        try:
            path = _find_by_any(patterns)
        except FileNotFoundError:
            continue
        cache_p = cache_dir / f"ord_int_{path.name}_{int(path.stat().st_mtime)}.pkl"
        if cache_p.exists():
            with open(cache_p, "rb") as f:
                cached = pickle.load(f)
            for ch, mp in cached.items():
                for code, q in mp.items():
                    per_ch[ch][code] += q
            log.info(f"  {path.name}: (cached)")
            continue
        this_file: dict = defaultdict(lambda: defaultdict(int))
        cnt = _parse_internal_order_xlsx(path, this_file)
        for ch, mp in this_file.items():
            for code, q in mp.items():
                per_ch[ch][code] += q
        with open(cache_p, "wb") as f:
            pickle.dump({ch: dict(mp) for ch, mp in this_file.items()}, f)
        log.info(f"  {path.name}: {cnt:,}행 처리 (cached)")
        total += cnt

    for ch, mp in per_ch.items():
        ord_int[ch] = dict(mp)
    log.info(f"  → 총 {total:,}행 처리")


# ─────────────────────────────────────────────
# Stage 6: 외부 주문 (MUSINSA/NAVER/ZIGZAG 시트, 단품코드 조합)
# ─────────────────────────────────────────────
_STYLE_RX = re.compile(r"(SP[A-Z0-9]{8})", re.IGNORECASE)
_OPT_TXT_RX = re.compile(r"\[(\d{1,3})\][^\^]*\^\s*\w*\[(\d{1,4})\]")
_OPT_JSN_RX = re.compile(r"\[(\d{1,3})\]")
_NAVER_COLOR_RX = re.compile(r"\((\d{1,3})\)")
_NAVER_SIZE_A_RX = re.compile(r"사이즈[^:]*:\s*[^/]*?(\d{2,4})")
_NAVER_SIZE_B_RX = re.compile(r"/\s*[^(/]*?\((\d{2,4})\)")


def _extract_style(text: str) -> Optional[str]:
    if not text:
        return None
    m = _STYLE_RX.findall(text.upper())
    return m[-1] if m else None


def _parse_musinsa_order_csv(path: Path, per: dict) -> None:
    rows = _read_cp949_csv(path)
    for r in rows:
        name = str(r.get("GOODS_NM") or "")
        style = str(r.get("STYLE_NO") or "").strip().upper() or _extract_style(name)
        if not style:
            continue
        opt_txt = str(r.get("GOODS_OPT") or r.get("OPTION_TXT") or "")
        cs = None
        if opt_txt:
            m = _OPT_TXT_RX.search(opt_txt)
            if m:
                cs = (m.group(1).zfill(2), m.group(2).zfill(3))
        if not cs:
            opt_json = str(r.get("GOODS_OPTION_NAME") or "")
            if opt_json.startswith("["):
                try:
                    arr = json.loads(opt_json)
                    if len(arr) >= 2:
                        m1 = _OPT_JSN_RX.search(arr[0])
                        m2 = _OPT_JSN_RX.search(arr[1])
                        if m1 and m2:
                            cs = (m1.group(1).zfill(2), m2.group(1).zfill(3))
                except Exception:
                    pass
        if not cs:
            continue
        state = str(r.get("ORD_STATE") or "")
        if any(x in state for x in ("취소", "반품", "교환")):
            continue
        code = f"{style}{cs[0]}{cs[1]}"
        per[code] += _int(r.get("QTY", 0)) or 1


def _parse_naver_order_csv(path: Path, per: dict) -> None:
    rows = _read_cp949_csv(path)
    for r in rows:
        name = str(r.get("PRODUCT_NAME") or "")
        style = _extract_style(name)
        if not style:
            continue
        opt = str(r.get("PRODUCT_OPTION_CONTENTS") or "")
        cm = _NAVER_COLOR_RX.search(opt)
        sm = _NAVER_SIZE_A_RX.search(opt) or _NAVER_SIZE_B_RX.search(opt)
        if not (cm and sm):
            continue
        claim = str(r.get("CLAIM_STATUS") or "")
        if "CANCEL" in claim.upper():
            continue
        code = f"{style}{cm.group(1).zfill(2)}{sm.group(1).zfill(3)}"
        per[code] += 1  # 사용자 확정: 각 행 = 1건


def _parse_zigzag_order_csv(path: Path, per: dict) -> None:
    rows = _read_cp949_csv(path)
    for r in rows:
        code = str(r.get("CUSTOM_PRODUCT_ITEM_CODE") or "").strip().upper()
        if not code or len(code) < 12:
            continue
        status = str(r.get("ORDER_STATUS") or "")
        if any(x in status.upper() for x in ("CANCEL", "REFUND")):
            continue
        per[code] += _int(r.get("QUANTITY", 0)) or 1


def stage6_orders_ext(ord_ext: dict) -> None:
    log.info("Stage 6: 외부 주문 로딩 (EDW CSV, 7/7 신규 구조)...")

    # ── MUSINSA (7/9 신규: 내부/외부 두 파일 합산 · 지그재그와 동일 처리)
    # 이전: 외부(내부포함) 파일 하나. 7/9부터 지그재그처럼 내부+외부 두 파일 분리 → 합산.
    per: dict = defaultdict(int)
    for patterns in (
        ["주문_내부_MUSINSA_단품명필요*.csv"],
        ["주문_외부_MUSINSA_단품명필요*.csv"],
        # 옛 통합 파일명 fallback
        ["주문_외부(내부포함)_MUSINSA_단품명필요*.csv"],
        ["주문_내부+외부_무신사*.csv"],
    ):
        try:
            path = _find_by_any(patterns)
        except FileNotFoundError:
            continue
        before = sum(per.values())
        _parse_musinsa_order_csv(path, per)
        added = sum(per.values()) - before
        if added > 0:
            log.info(f"  MUSINSA[{path.name}]: +{added:,}주문")
    ord_ext["무신사"] = dict(per)
    log.info(f"  MUSINSA 합계: {sum(per.values()):,}주문 ({len(per):,}단품)")

    # ── NAVER (외부(내부포함) 파일 하나) ──
    per = defaultdict(int)
    path = _find_by_any([
        "주문_외부(내부포함)_NAVER_단품명필요*.csv",
        "주문_내부+외부_네이버*.csv",
    ])
    _parse_naver_order_csv(path, per)
    ord_ext["네이버"] = dict(per)
    log.info(f"  NAVER: {sum(per.values()):,}주문 ({len(per):,}단품) [{path.name}]")

    # ── ZIGZAG (내부 + 외부 두 파일 합산) ──
    per = defaultdict(int)
    for patterns in (
        ["주문_내부_ZIGZAG_단품명필요*.csv"],
        ["주문_외부_ZIGZAG_단품명필요*.csv"],
    ):
        try:
            path = _find_by_any(patterns)
        except FileNotFoundError:
            log.warning(f"  ZIGZAG 파일 없음: {patterns}")
            continue
        before = sum(per.values())
        _parse_zigzag_order_csv(path, per)
        log.info(f"  ZIGZAG[{path.name}]: +{sum(per.values()) - before:,}주문")
    ord_ext["지그재그"] = dict(per)
    log.info(f"  ZIGZAG 합계: {sum(per.values()):,}주문 ({len(per):,}단품)")


def stage6b_price_fallback(price_fallback: dict) -> None:
    """주문 파일에서 정상가 fallback 추출 (EDW CSV 기준)."""
    log.info("Stage 6b: 주문 파일에서 정상가 fallback...")

    # ZIGZAG — ORIGINAL_PRICE (내부/외부 파일 모두 순회)
    add = 0
    for patterns in (
        ["주문_내부_ZIGZAG_단품명필요*.csv"],
        ["주문_외부_ZIGZAG_단품명필요*.csv"],
    ):
        try:
            path = _find_by_any(patterns)
        except FileNotFoundError:
            continue
        for r in _read_cp949_csv(path):
            code = str(r.get("CUSTOM_PRODUCT_ITEM_CODE") or "").strip().upper()
            p = _int(r.get("ORIGINAL_PRICE", 0))
            if code and p > 0 and code not in price_fallback:
                price_fallback[code] = p
                add += 1
    log.info(f"  ZIGZAG 정상가 fallback: +{add}")

    # MUSINSA — PRICE / 0.85 = 정상가 역산 (NORMAL_PRICE 있으면 우선 사용)
    # 7/9 신규: 내부/외부 두 파일 지원 (지그재그와 동일 구조)
    add = 0
    musinsa_paths = []
    for patterns in (
        ["주문_내부_MUSINSA_단품명필요*.csv"],
        ["주문_외부_MUSINSA_단품명필요*.csv"],
        ["주문_외부(내부포함)_MUSINSA_단품명필요*.csv"],
        ["주문_내부+외부_무신사*.csv"],
    ):
        try:
            musinsa_paths.append(_find_by_any(patterns))
        except FileNotFoundError:
            continue
    for path in musinsa_paths:
        for r in _read_cp949_csv(path):
            style = str(r.get("STYLE_NO") or "").strip().upper() or _extract_style(
                str(r.get("GOODS_NM") or "")
            )
            if not style:
                continue
            opt_txt = str(r.get("GOODS_OPT") or r.get("OPTION_TXT") or "")
            cs = None
            if opt_txt:
                m = _OPT_TXT_RX.search(opt_txt)
                if m:
                    cs = (m.group(1).zfill(2), m.group(2).zfill(3))
            if not cs:
                opt_json = str(r.get("GOODS_OPTION_NAME") or "")
                if opt_json.startswith("["):
                    try:
                        arr = json.loads(opt_json)
                        if len(arr) >= 2:
                            m1 = _OPT_JSN_RX.search(arr[0])
                            m2 = _OPT_JSN_RX.search(arr[1])
                            if m1 and m2:
                                cs = (m1.group(1).zfill(2), m2.group(1).zfill(3))
                    except Exception:
                        pass
            if not cs:
                continue
            code = f"{style}{cs[0]}{cs[1]}"
            # NORMAL_PRICE 컬럼이 있으면 그대로, 없으면 PRICE/0.85 (15% 할인 가정)
            np_val = _int(r.get("NORMAL_PRICE", 0))
            p = _int(r.get("PRICE", 0))
            price = np_val if np_val > 0 else (int(p / 0.85) if p > 0 else 0)
            if code and price > 0 and code not in price_fallback:
                price_fallback[code] = price
                add += 1
    log.info(f"  MUSINSA 정상가 fallback: +{add}")


# ─────────────────────────────────────────────
# Stage 7: 통합 + 신상 필터 + 랭킹 + 금액
# ─────────────────────────────────────────────
def stage7_finalize(skus_master: dict, bw_qty: dict, bw_name: dict,
                    inv_int: dict, inv_ext: dict,
                    ord_int: dict, ord_ext: dict,
                    price_fallback: dict, name_fallback: dict) -> dict:
    log.info("Stage 7: 통합 + 신상 필터 + 랭킹 + 금액...")

    # 스타일 그룹 평균가 (10자리 스타일 → 평균 정상가)
    style_prices = defaultdict(list)
    for code, m in skus_master.items():
        if m.get("정상가", 0) > 0 and len(code) >= 10:
            style_prices[code[:10]].append(m["정상가"])
    for code, p in price_fallback.items():
        if p > 0 and len(code) >= 10:
            style_prices[code[:10]].append(p)
    style_avg = {s: int(sum(ps) / len(ps)) for s, ps in style_prices.items()}
    log.info(f"  스타일 평균가: {len(style_avg):,} 스타일")

    # 등장하는 모든 단품 집합
    all_codes = set(skus_master) | set(bw_qty)
    for m in inv_int.values():
        all_codes |= set(m)
    for m in inv_ext.values():
        all_codes |= set(m)
    for m in ord_int.values():
        all_codes |= set(m)
    for m in ord_ext.values():
        all_codes |= set(m)

    # SP로 시작하는 15자리 코드만
    all_codes = {c for c in all_codes if c and len(c) >= 12
                 and c != "SPAO_UNKNOWN" and c.startswith("SP")}

    def mkrow():
        r = {"단품코드": "", "단품명": "", "정상가": 0,
             "매출랭킹": 9999, "온라인랭킹": 9999,
             "출고율": 0, "온라인비중": 0, "누판율": 0, "주판율": 0,
             "주간외형매출": 0, "in_qty": 0, "cum_qty": 0, "wk_qty": 0,
             "_last_date": _TODAY, "inv_반응과": 0}
        for ch in CHANNELS:
            r[f"inv_{ch}"] = 0
            r[f"ord_{ch}"] = 0
            r[f"daily_{ch}"] = 0
            r[f"daily_amt_{ch}"] = 0
            r[f"inv_amt_{ch}"] = 0
        for ch in EXT_CH:
            r[f"wh_{ch}"] = 0
            r[f"wh_amt_{ch}"] = 0
        return r

    skus = {}
    for code in all_codes:
        r = mkrow()
        r["단품코드"] = code
        m = skus_master.get(code, {})
        for k in ("단품명", "정상가", "출고율", "누판율", "주판율",
                  "주간외형매출", "in_qty", "cum_qty", "wk_qty"):
            if k in m:
                r[k] = m[k]
        # 반응과
        r["inv_반응과"] = max(0, int(bw_qty.get(code, 0)))
        # 단품명 fallback
        if not r["단품명"] and bw_name.get(code):
            r["단품명"] = bw_name[code]
        if not r["단품명"] and code in name_fallback:
            r["단품명"] = name_fallback[code]
        # 정상가 fallback: 우선순위 - 마스터 → 개별 → 스타일 평균
        if r["정상가"] == 0:
            if code in price_fallback:
                r["정상가"] = price_fallback[code]
            elif code[:10] in style_avg:
                r["정상가"] = style_avg[code[:10]]
        # 재고: 내부 + 외부창고 (대시보드 뺄셈 로직 정합)
        for ch in CHANNELS:
            internal = max(0, int(inv_int.get(ch, {}).get(code, 0)))
            ext = max(0, int(inv_ext.get(ch, {}).get(code, 0))) if ch in EXT_CH else 0
            r[f"inv_{ch}"] = internal + ext
            if ch in EXT_CH:
                r[f"wh_{ch}"] = ext
        # 주문 (내부)
        for ch, mp in ord_int.items():
            q = max(0, int(mp.get(code, 0)))
            r[f"ord_{ch}"] = q
            r[f"daily_{ch}"] = q
        # 주문 (외부)
        for ch, mp in ord_ext.items():
            q = max(0, int(mp.get(code, 0)))
            r[f"ord_{ch}"] = q
            r[f"daily_{ch}"] = q
        skus[code] = r

    # 7/13: 신상 필터 해제 — 이월상품 포함 (사용자 정책 변경)
    log.info(f"  전체(신상+이월): {len(skus):,} 단품")

    # 매출 랭킹 (주간외형매출 순)
    sorted_codes = sorted(skus.keys(),
                          key=lambda c: skus[c].get("주간외형매출", 0),
                          reverse=True)
    for i, code in enumerate(sorted_codes):
        skus[code]["매출랭킹"] = i + 1
        skus[code]["온라인랭킹"] = i + 1

    # 금액 (재고액 = 정상가 × 재고)
    for r in skus.values():
        p = r.get("정상가", 0)
        for ch in CHANNELS:
            r[f"inv_amt_{ch}"] = int(p * r[f"inv_{ch}"])
            r[f"daily_amt_{ch}"] = int(p * r[f"daily_{ch}"])
        for ch in EXT_CH:
            r[f"wh_amt_{ch}"] = int(p * r[f"wh_{ch}"])

    # 누판/주판/출고율: 정수 반올림 후 100으로 나누기 (0~1 소수)
    for r in skus.values():
        for k in ("누판율", "주판율", "출고율"):
            v = _num(r.get(k, 0))
            r[k] = round(v) / 100

    return skus


# ─────────────────────────────────────────────
# Stage 8: CSV gzip 저장
# ─────────────────────────────────────────────
def stage8_save(skus: dict) -> None:
    log.info(f"Stage 8: gzip 저장 → {OUT_PATH.name}")
    cols = ["단품코드", "단품명", "매출랭킹", "온라인랭킹", "정상가",
            "출고율", "온라인비중", "누판율", "주판율", "주간외형매출",
            "in_qty", "cum_qty", "wk_qty", "_last_date", "inv_반응과"]
    for ch in CHANNELS:
        cols += [f"inv_{ch}", f"inv_amt_{ch}", f"ord_{ch}",
                 f"daily_{ch}", f"daily_amt_{ch}"]
    for ch in EXT_CH:
        cols += [f"wh_{ch}", f"wh_amt_{ch}"]

    with gzip.open(OUT_PATH, "wt", encoding="utf-8-sig", newline="",
                   compresslevel=9) as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for code in sorted(skus.keys()):
            w.writerow(skus[code])
    log.info(f"  → {OUT_PATH.stat().st_size:,} bytes")


# ─────────────────────────────────────────────
# 실행
# ─────────────────────────────────────────────
import pickle as _pkl

_STATE_KEYS = (
    "skus_master", "bw_qty", "bw_name",
    "inv_int", "inv_ext", "ord_int", "ord_ext",
    "price_fallback", "name_fallback",
)


def _state_path(stage: str) -> Path:
    d = _HERE / ".cache"
    d.mkdir(exist_ok=True)
    return d / f"state_{stage}.pkl"


def _save_state(stage: str, state: dict) -> None:
    with open(_state_path(stage), "wb") as f:
        _pkl.dump({k: state[k] for k in _STATE_KEYS}, f)
    log.info(f"  💾 state saved → .cache/state_{stage}.pkl")


def _load_state(stage: str) -> Optional[dict]:
    p = _state_path(stage)
    if not p.exists():
        return None
    with open(p, "rb") as f:
        return _pkl.load(f)


def main() -> None:
    print('[MAIN ENTRY]', flush=True); import os
    resume = os.environ.get("REBA_RESUME_FROM", "")
    state: dict = {k: {} for k in _STATE_KEYS}
    if resume:
        loaded = _load_state(resume)
        if loaded:
            state.update(loaded)
            log.info(f"▶ resumed from checkpoint '{resume}'")

    if not resume:
        stage1_master(state["skus_master"])
        _save_state("s1", state)
    if resume in ("", "s1"):
        stage2_bw(state["bw_qty"], state["bw_name"], state["price_fallback"])
        _save_state("s2", state)
        resume = ""
    if resume in ("", "s2"):
        stage3_internal_inv(state["inv_int"], state["name_fallback"])
        _save_state("s3", state)
        resume = ""
    if resume in ("", "s3"):
        stage4_ext_wh(state["inv_ext"], state["name_fallback"], state["price_fallback"])
        _save_state("s4", state)
        resume = ""
    if resume in ("", "s4"):
        stage5_orders_internal(state["ord_int"])
        _save_state("s5", state)
        resume = ""
    if resume in ("", "s5"):
        stage6_orders_ext(state["ord_ext"])
        stage6b_price_fallback(state["price_fallback"])
        _save_state("s6", state)

    skus = stage7_finalize(
        state["skus_master"], state["bw_qty"], state["bw_name"],
        state["inv_int"], state["inv_ext"],
        state["ord_int"], state["ord_ext"],
        state["price_fallback"], state["name_fallback"],
    )
    stage8_save(skus)

    # ── 검증 요약
    log.info("─" * 60)
    log.info(f"검증 요약 ({OUT_PATH.name})")
    log.info(f"총 단품수: {len(skus):,}")
    price_ok = sum(1 for r in skus.values() if r["정상가"] > 0)
    if len(skus):
        log.info(f"정상가 커버: {price_ok:,} ({price_ok/len(skus)*100:.1f}%)")

    tot_inv = sum(r[f"inv_amt_{ch}"] for r in skus.values() for ch in CHANNELS)
    tot_ext = sum(r[f"wh_amt_{ch}"] for r in skus.values() for ch in EXT_CH)
    log.info(f"총재고액: {tot_inv/1e8:.1f}억")
    log.info(f"내부창고: {(tot_inv-tot_ext)/1e8:.1f}억")
    log.info(f"외부창고: {tot_ext/1e8:.1f}억")
    for ch in EXT_CH:
        wq = sum(r[f"wh_{ch}"] for r in skus.values())
        wq = sum(r[f"wh_{ch}"] for r in skus.values())
        wa = sum(r[f"wh_amt_{ch}"] for r in skus.values())
        log.info(f"    {ch}: {wq:>7,}장 - {wa/1e8:.1f}억")

    log.info("완료. GitHub push 필요:")
    log.info(f"  -> data/test/{OUT_PATH.name}")


if __name__ == "__main__":
    main()
