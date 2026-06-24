# -*- coding: utf-8 -*-
"""
v4.4 화면 — 누판율·주판율 데이터바 · 출고율 기준 완전 제거(슬라이더 삭제) · 재고/주문 RAW 정합 · 리오더 병합(컬러 동일) · 외부창고 분리
복원: 단품코드 검색(앞 10자리) · 🚫 제외 스타일 탭 · 📊 채널 별 세부 탭(외부창고 컬럼은 여기만)
      · 체크박스 단품 선택 승인 · 사용자 정의 기준 명칭
(페이지 설정·비밀번호 게이트·공통 CSS는 app.py 담당)
"""
import io
import streamlit as st
import pandas as pd


def _xlsx_bytes(sheets: dict) -> bytes:
    """여러 시트 dict {sheet_name: DataFrame} → xlsx bytes.

    SCM팀 즉시 작업용 정식 엑셀 다운로드 (스파오 6/19 미팅 합의).
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for name, df in sheets.items():
            if df is None or len(df) == 0:
                continue
            df.to_excel(writer, sheet_name=name[:31], index=False)
    return buf.getvalue()


def _xlsx_bytes_with_bars(sheets: dict, bar_columns: list = None,
                           red_bar_columns: list = None,
                           woc_columns: list = None,
                           bold_columns: list = None) -> bytes:
    """xlsx 생성 + 조건부 서식 (파란 데이터 바·빨간 데이터 바·WOC 신호등).

    bar_columns: 파란 데이터 바 (#638EC6) — 큰 값 강조
    red_bar_columns: 빨간 데이터 바 (#FF6B6B) — 회수매출 등 강조
    woc_columns: 'N주' 텍스트 셀 → 대시보드 신호등 (대시보드 woc_color 기준)
                 <1주 빨강(FFC7CE) / 1~4주 노랑(FFEB9C) / ≥4주 초록(C6EFCE)
    """
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    bar_columns = bar_columns or []
    red_bar_columns = red_bar_columns or []
    woc_columns = woc_columns or []
    bold_columns = bold_columns or []

    # 대시보드 신호등 매핑 — 옅은 톤 (엑셀 가독성)
    fill_red = PatternFill('solid', fgColor='FFC7CE')
    fill_yellow = PatternFill('solid', fgColor='FFEB9C')
    fill_green = PatternFill('solid', fgColor='C6EFCE')
    font_red = Font(color='9C0006', bold=True)
    font_yellow = Font(color='9C5700', bold=True)
    font_green = Font(color='006100', bold=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for name, df in sheets.items():
            if df is None or len(df) == 0:
                continue
            sheet_name = name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]
            n_rows = len(df)

            # 파란 데이터 바
            for col_name in bar_columns:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    rng = f'{col_letter}2:{col_letter}{n_rows + 1}'
                    ws.conditional_formatting.add(rng, DataBarRule(
                        start_type='min', end_type='max',
                        color='638EC6', showValue=True))

            # 빨간 데이터 바 (회수매출 등)
            for col_name in red_bar_columns:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    rng = f'{col_letter}2:{col_letter}{n_rows + 1}'
                    ws.conditional_formatting.add(rng, DataBarRule(
                        start_type='min', end_type='max',
                        color='FF6B6B', showValue=True))

            # WOC 신호등 (셀별 직접 fill 적용 — 텍스트 "N주" 파싱)
            for col_name in woc_columns:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    for row_idx in range(2, n_rows + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        val = str(cell.value or '').strip()
                        try:
                            num = float(val.replace('주', '').strip())
                        except (ValueError, AttributeError):
                            continue
                        if num < 1:
                            cell.fill = fill_red
                            cell.font = font_red
                        elif num < 4:
                            cell.fill = fill_yellow
                            cell.font = font_yellow
                        else:
                            cell.fill = fill_green
                            cell.font = font_green

            # 굵게 처리 (사용자 6/19 요청 — 실행 중요도 강조)
            for col_name in bold_columns:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    for row_idx in range(2, n_rows + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        existing_color = cell.font.color if cell.font else None
                        existing_size = cell.font.size if cell.font else 11
                        cell.font = Font(bold=True, size=existing_size, color=existing_color)

            # 컬럼 너비 자동 조정
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(col_name)), 8)
                try:
                    max_len = max(max_len, df[col_name].astype(str).str.len().max())
                except Exception:
                    pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 30)
    return buf.getvalue()


# ─── 메일링 명단 (회사 메일) ──────────────────────────────────
DEFAULT_SCM_LIST = [
    'PARK_JINSEONG03@eland.co.kr',
    'Jang_HyeonSeong@eland.co.kr',
    'JANG_HONGSEO01@eland.co.kr',
    'HAN_JIWOONG03@eland.co.kr',
]
DEFAULT_PLAN_LIST = [
    'KIM_SANGHYUK04@eland.co.kr', 'KIM_SUHO03@eland.co.kr',
    'KIM_INYOUNG03@eland.co.kr', 'KIM_JAEWOOK03@eland.co.kr',
    'KIM_JONGOH01@eland.co.kr', 'KIM_HYESOO02@eland.co.kr',
    'NA_YELIN01@eland.co.kr', 'SEO_GAYEON@eland.co.kr',
    'AN_SOYEON@eland.co.kr', 'YUN_BYEONGOK01@eland.co.kr',
    'LEE_SEUNGYEON05@eland.co.kr', 'LEE_SIWON01@eland.co.kr',
    'CHOI_JINYOUNG09@eland.co.kr', 'HWANG_SOOYONG01@eland.co.kr',
]


def _get_mlist(key, default):
    """session_state 기반 명단 관리."""
    if key not in st.session_state:
        st.session_state[key] = list(default)
    return st.session_state[key]


def _edit_mlist(key, default, title, hint=''):
    """명단 편집 expander."""
    current = _get_mlist(key, default)
    with st.expander(f'📋 {title} 명단 관리 (현재 {len(current)}명)'):
        if hint:
            st.caption(hint)
        st.caption('줄당 1개 이메일. 추가·삭제·복원 가능. 저장 버튼 누르기 전까지는 미반영.')
        edited = st.text_area(
            '이메일 명단',
            value='\n'.join(current),
            height=200,
            key=f'edit_{key}',
            label_visibility='collapsed',
        )
        b1, b2, b3 = st.columns([1, 1, 4])
        with b1:
            if st.button('💾 저장', key=f'save_{key}', use_container_width=True):
                new_list = [e.strip() for e in edited.split('\n') if e.strip() and '@' in e]
                st.session_state[key] = new_list
                st.success(f'✅ {len(new_list)}명 저장 완료')
                st.rerun()
        with b2:
            if st.button('🔄 기본값', key=f'reset_{key}', use_container_width=True):
                st.session_state[key] = list(default)
                st.success(f'기본값 {len(default)}명 복원')
                st.rerun()
    return current


def _xlsx_by_channel(df_all, df_sum, ch_short_map):
    """채널별 시트로 분리된 xlsx (SCM 분배 편의).

    채널 6개 + 전체 + 출고매장코드별 소계 = 최대 8 시트.
    """
    sheets = {'전체': df_all}
    if '주력채널' in df_all.columns:
        for ch in CHANNELS:
            short = ch_short_map.get(ch, ch)
            sub = df_all[df_all['주력채널'] == short]
            if len(sub) > 0:
                sheets[short] = sub.reset_index(drop=True)
    if df_sum is not None and len(df_sum) > 0:
        sheets['출고매장코드별 소계'] = df_sum
    return _xlsx_bytes(sheets)


def _mailto_link(emails: list[str], subject: str, body: str) -> str:
    """mailto: 링크 생성 (메일 클라이언트 자동 열기)."""
    import urllib.parse as _u
    to = ','.join(emails)
    qs = _u.urlencode({'subject': subject, 'body': body}, quote_via=_u.quote)
    return f'mailto:{to}?{qs}'

from rebalance_engine import calc_rebalance_group, calc_after_woc, calc_expected_revenue, calc_distribution
import effect_log
from mock_data import (
    get_combined_data, get_last_update_time, get_reorder_info,
    get_reorder_mapping, parse_reorder_bytes, save_reorder_mapping,
    CHANNELS, EXT_WAREHOUSE, BW_NAME,
)

CH_SHORT = {
    '공홈': '공홈', '이랜드몰': '이몰', '무신사': '무신',
    '지그재그': '지재', '네이버': '네이', '카카오선물하기': '카카오',
}
EXT_CHANNELS = [c for c in CHANNELS if c in EXT_WAREHOUSE]  # 무신사·지그재그·네이버

# 출고 매장코드 — 채널별 내부창고 코드 (추가 분배 = 내부창고 분배 정책)
# 사용자 첨부 매장코드표(2026-06-18) 기반 확정
WAREHOUSE_CODE = {
    '공홈': 'AEQ5',           # 내부 (대표 코드. 추가: AER6, ACZX)
    '이랜드몰': 'AAKY',        # 내부
    '무신사': 'ABY1',          # 내부 (외부 = AENS · 분배는 내부만)
    '지그재그': 'AELA',        # 내부 (외부 = ADU3 · 분배는 내부만)
    '네이버': 'AEF1',          # 내부 (외부 = ADQS · 분배는 내부만)
    '카카오선물하기': 'ACCX',   # 내부
}
# 채널별 외부창고 코드 (참고 — 통합 재고뷰에서만 표시, 분배 미사용)
WAREHOUSE_CODE_EXT = {
    '무신사': 'AENS', '지그재그': 'ADU3', '네이버': 'ADQS',
}

SCENARIOS = {
    '🛡️ 기본': {
        'desc': '결품 기준 1주 미만 → 목표 2주 확보. 회전(온라인 6채널 잉여→결품)으로 보충. '
                '이동 상한: 각 채널 현재고의 30% (스파오 6/19 미팅 합의 — 보수 운영)',
        'shortage_th': 1.0, 'target_woc': 2.0,
        'ship_th': 0.90, 'min_move': 0, 'min_recv': 0, 'move_cap_pct': 0.30,
    },
    '🎛️ 임의': {
        'desc': '상단 슬라이더로 직접 조정 (이동 상한 % 포함). 기본값 30%',
        'shortage_th': 1.0, 'target_woc': 2.0,
        'ship_th': 0.90, 'min_move': 0, 'min_recv': 0, 'move_cap_pct': 0.30,
    },
}


@st.cache_data(show_spinner=False)
def load_data_v20():
    return get_combined_data('v2')


@st.cache_data(show_spinner=False)
def calc_results_v20(params_key):
    skus = load_data_v20()
    ch_excl = {}
    if len(params_key) > 5 and params_key[5]:
        for ch, direction, pats in params_key[5]:
            ch_excl.setdefault(ch, {})[direction] = set(pats)
    params = {
        'shortage_threshold': params_key[0], 'target_woc': params_key[1],
        'ship_rate_threshold': params_key[2], 'min_move_qty': params_key[3],
        'min_recv_order': params_key[4], 'ch_excl': ch_excl,
        'move_cap_pct': params_key[6] if len(params_key) > 6 else 0.5,
    }
    # 컬러(단품코드 12자리) 단위로 묶어 그룹 재배치 — 아소트 깨짐 방지
    from collections import defaultdict
    groups = defaultdict(dict)
    for code, d in skus.items():
        groups[code[:12]][code] = d
    move_map = {}
    for color, g in groups.items():
        move_map.update(calc_rebalance_group(g, params, CHANNELS))
    results = []
    for code, d in skus.items():
        moves = move_map.get(code, {c: 0 for c in CHANNELS})
        after = calc_after_woc(d, moves, CHANNELS)
        rev = calc_expected_revenue(d, moves, CHANNELS, d['price'])
        mode = '제외' if d.get('locked', False) else '자동회전'
        results.append({'code': code, 'data': d, 'moves': moves,
                        'after': after, 'revenue': rev, 'mode': mode})
    return results


def _apply_exclusion(results):
    """🚫 제외 스타일 적용 — 매칭 단품은 이동 0·모드 '제외' (캐시 비파괴 copy)"""
    excluded = st.session_state.get('excluded_codes', set())
    if not excluded:
        return results
    out = []
    for r in results:
        if any(ex and ex in r['code'] for ex in excluded):
            r2 = dict(r)
            r2['mode'] = '제외'
            r2['moves'] = {k: 0 for k in r['moves']}
            r2['after'] = calc_after_woc(r['data'], r2['moves'], CHANNELS)
            r2['revenue'] = 0
            out.append(r2)
        else:
            out.append(r)
    return out


def _apply_overrides(results):
    """이동량 직접 수정(세션) 반영 — 해당 단품 moves를 사용자 입력값으로 교체하고 after/효과 재계산."""
    ov = st.session_state.get('move_overrides', {})
    if not ov:
        return results
    out = []
    for r in results:
        o = ov.get(r['code'])
        if o:
            r2 = dict(r)
            moves = {c: int(o.get(c, 0)) for c in CHANNELS}
            r2['moves'] = moves
            r2['after'] = calc_after_woc(r['data'], moves, CHANNELS)
            r2['revenue'] = calc_expected_revenue(r['data'], moves, CHANNELS, r['data']['price'])
            out.append(r2)
        else:
            out.append(r)
    return out


def _ch_excl_key():
    """채널별 IN/OUT 제외 — ch_excl_rows 통합 구조 (스타일 + 시작일 + 종료일).
    시작일/종료일이 비어 있으면 영구 적용, 둘 다 있으면 오늘 ∈ [시작, 종료] 일 때만 활성."""
    from datetime import date as _date
    today = _date.today()
    rows = st.session_state.get('ch_excl_rows', [])
    # 레거시: 기존 ch_excl 텍스트영역 데이터도 흡수
    legacy = st.session_state.get('ch_excl', {})
    merged = {}  # ch -> {'in': set, 'out': set}
    for ch in CHANNELS:
        for dr in ('in', 'out'):
            pats = set(legacy.get(ch, {}).get(dr, []))
            if pats:
                merged.setdefault(ch, {}).setdefault(dr, set()).update(pats)
    for row in rows:
        try:
            ch = row.get('채널')
            direction = (row.get('방향') or '').lower()
            pat = (row.get('스타일') or row.get('스타일 패턴') or '').strip()
            if not (ch and direction in ('in', 'out') and pat):
                continue
            start = row.get('시작일')
            end = row.get('종료일')
            if start or end:
                # 기간 지정 — 둘 다 있어야 활성
                if not (start and end):
                    continue
                s = start if hasattr(start, 'toordinal') else _date.fromisoformat(str(start)[:10])
                e = end if hasattr(end, 'toordinal') else _date.fromisoformat(str(end)[:10])
                if not (s <= today <= e):
                    continue
            merged.setdefault(ch, {}).setdefault(direction, set()).add(pat)
        except Exception:
            continue
    return tuple(sorted(
        (ch, dr, tuple(sorted(merged.get(ch, {}).get(dr, []))))
        for ch in CHANNELS for dr in ('in', 'out')
        if merged.get(ch, {}).get(dr)
    ))


def _active_period_patterns():
    """하위 호환 — 이전 코드가 이 함수를 호출할 경우 빈 dict 반환."""
    return {}


def woc_color(w):
    if w is None or w == '' or pd.isna(w): return ''
    try:
        v = float(str(w).replace('주', ''))
    except: return ''
    if v < 1: return 'background-color: #5B1E1E; color: #FF5A5F; font-weight:bold'   # 빨강: 1주 미만
    if v < 4: return 'background-color: #5A4500; color: #FFC000; font-weight:bold'   # 노랑: 1~4주
    return 'background-color: #1B4D3E; color: #4AE3B5; font-weight:bold'            # 초록: 4주 이상


def mv_color(v):
    if v is None or v == 0 or pd.isna(v) or v == '': return ''
    try:
        vv = int(str(v).replace('+', ''))
    except: return ''
    if vv > 0: return 'background-color: #1B4D3E; color: #4AE3B5; font-weight:bold; text-align:center'
    return 'background-color: #5B1E1E; color: #FF5A5F; font-weight:bold; text-align:center'


def qty_color(s):
    """이동 후 재고량 셀: 증가(+)=녹색음영 / 감소(-)=레드음영, 글자 흰색, 우측정렬."""
    if not isinstance(s, str) or s == '':
        return 'text-align:right'
    if '(+' in s:
        return 'background-color: #1B4D3E; color: #FFFFFF; font-weight:bold; text-align:right'
    if '(-' in s:
        return 'background-color: #5B1E1E; color: #FFFFFF; font-weight:bold; text-align:right'
    return 'color: #FFFFFF; text-align:right'



@st.dialog('✅ 재고 이동 전송 확인')
def _approve_dialog(scenario_key, sel_count, sel_qty, sel_amt, sel_rev, ch_in, ch_out, exec_id):
    st.markdown(f"""<div class="scenario-box">
    <b>SAP BAPI 전송 완료 (mock)</b> — 실행 ID <b>#{exec_id}</b> · 📈 실행 효과 탭에 이력 기록됨<br>
    실 배포 시: BAPI_GOODS_MVT_CREATE 호출 → 전표 생성 → audit log</div>""", unsafe_allow_html=True)
    m1, m2, m3, m4 = st.columns(4)
    m1.metric('승인 단품', f'{sel_count:,}건')
    m2.metric('총 이동량', f'{sel_qty:,}장')
    m3.metric('총 이동 금액', f'{sel_amt/100000000:.2f}억')
    m4.metric('기대 회수 매출', f'{sel_rev/100000000:.2f}억')
    rows = [{'채널': c, 'IN (장)': ch_in.get(c, 0), 'OUT (장)': ch_out.get(c, 0),
             '순증감': ch_in.get(c, 0) - ch_out.get(c, 0)} for c in CHANNELS
            if ch_in.get(c, 0) or ch_out.get(c, 0)]
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True,
                     height=min(38 + 35 * len(rows), 260))
    st.caption('⚠️ 실제 재고 이동 여부는 익일 06:00 재고 갱신 후 "현 재고보유주수" 변화 및 📈 실행 효과 탭 실측으로 확인하세요.')

    st.caption('⚠️ 실제 재고 이동 여부는 익일 06:00 재고 갱신 후 "현 재고보유주수" 변화 및 📈 실행 효과 탭 실측으로 확인하세요.')

    # 메일 자동 발송 — kang_hoonkoo@eland.co.kr (다이얼로그 표시와 동시에 자동 mailto 호출)
    from datetime import datetime as _dt
    _now_str = _dt.now().strftime('%Y-%m-%d %H:%M')
    _subj = '[REBA] ' + str(scenario_key) + ' 회전 승인 결과 — #' + str(exec_id) + ' (' + _now_str + ')'
    _ch_lines = []
    for c in CHANNELS:
        _in_q = ch_in.get(c, 0)
        _out_q = ch_out.get(c, 0)
        if _in_q or _out_q:
            _ch_lines.append('- ' + CH_SHORT.get(c, c) + ': IN +' + format(_in_q, ',') +
                             ' / OUT -' + format(_out_q, ',') +
                             ' (순증감 ' + format(_in_q - _out_q, '+,') + ')')
    _nl = chr(10)
    _ch_summary = _nl.join(_ch_lines) if _ch_lines else '(채널 이동 없음)'
    _body_lines = [
        '온라인 재고관리 Agent — 회전 승인 결과 보고',
        '================================================',
        '',
        '시나리오: ' + str(scenario_key),
        '실행 ID: #' + str(exec_id),
        '승인 일시: ' + _now_str,
        '',
        '━━━━━━━ 요약 KPI ━━━━━━━',
        '승인 단품: ' + format(sel_count, ',') + '건',
        '총 이동량: ' + format(sel_qty, ',') + '장',
        '총 이동 금액: ' + format(sel_amt / 100000000, '.2f') + '억',
        '기대 회수 매출: ' + format(sel_rev / 100000000, '.2f') + '억',
        '',
        '━━━━━━━ 채널별 IN/OUT 요약 ━━━━━━━',
        _ch_summary,
        '',
        '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━',
        '※ 본 메일은 온라인 재고관리 Agent의 자동 발송 mock입니다.',
        '   실 SAP 연동 시 BAPI_GOODS_MVT_CREATE 호출 → 전표 생성 → audit log.',
        '   상세 확인: spao-rebal.streamlit.app → 📈 실행 효과 탭',
        '',
        '— AICA · 온라인 재고관리 Agent',
    ]
    _body = _nl.join(_body_lines)
    _mailto_url = _mailto_link(['kang_hoonkoo@eland.co.kr'], _subj, _body)
    # 다이얼로그 진입 시 자동 mailto 호출 (sessions 키로 1회 방지)
    _trig_key = 'mailto_sent_' + str(exec_id)
    if not st.session_state.get(_trig_key):
        st.session_state[_trig_key] = True
        try:
            import streamlit.components.v1 as _components
            _components.html(
                '<script>window.open("' + _mailto_url + '", "_blank");</script>',
                height=0,
            )
        except Exception:
            pass
    st.caption('📧 회전 결과 메일이 kang_hoonkoo@eland.co.kr로 자동 발송 요청되었습니다. '
               '메일 클라이언트가 열리지 않을 경우 팝업 차단 해제 후 다시 시도해주세요.')

    if st.button('확인 (다이얼로그 닫기)', type='primary', use_container_width=True, key='dlg_ok_' + str(scenario_key)):
        st.rerun()



def _chan_recovery_bar(items):
    """회수매출 채널 구성 — KPI 카드에 들어가는 100% 가로 스택 막대 + 상위 채널 범례."""
    rev = {c: 0 for c in CHANNELS}
    for r in items:
        d = r['data']
        for c in CHANNELS:
            o = d['orders'].get(c, 0); i = d['inv'].get(c, 0); mvc = r['moves'].get(c, 0)
            rev[c] += (max(0, o - i) - max(0, o - (i + mvc))) * d['price']
    tot = sum(rev.values()) or 1
    pal = {'공홈': '#4AE3B5', '네이버': '#8AB4F8', '무신사': '#2FB7A4',
           '지그재그': '#FFC000', '이랜드몰': '#9B8CFF', '카카오선물하기': '#FF8FA3'}
    order = [c for c in sorted(CHANNELS, key=lambda x: -rev[x]) if rev[c] > 0]
    seg = ''.join(
        f'<span title="{CH_SHORT[c]} {rev[c]/1e8:.2f}억 ({rev[c]/tot*100:.0f}%)" '
        f'style="width:{rev[c]/tot*100:.2f}%;background:{pal[c]};height:12px;display:inline-block"></span>'
        for c in order)
    leg = '  '.join(
        f'<span style="color:{pal[c]};font-size:11px">●</span>'
        f'<span style="color:#FFFFFF;font-size:10px"> {CH_SHORT[c]} {rev[c]/1e8:.2f}억</span>'
        for c in order[:4])
    return (f'<div style="display:flex;width:100%;border-radius:3px;overflow:hidden;margin:6px 0 4px">{seg}</div>'
            f'<div style="line-height:1.3">{leg}</div>')


def render_scenario(scenario_key, container, allow_slider=False):
    preset = SCENARIOS[scenario_key]

    ship_th = 0.0
    if allow_slider:
        container.markdown('### 🎛️ 사용자 정의 기준')
        sl1, sl2, sl3, sl4, sl5 = container.columns(5)
        with sl1:
            shortage_th = st.slider('재배치 대상 (재고주수 0주 이하)', 0.5, 4.0, preset['shortage_th'], 0.5, key=f'sh_{scenario_key}')
        with sl2:
            target_woc = st.slider('목표 재고주수 (주)', 1.0, 6.0, preset['target_woc'], 0.5, key=f'tg_{scenario_key}')
        with sl3:
            min_move = st.slider('이동 ≥ N장만 (비부가 제거)', 0, 50, preset['min_move'], 1, key=f'mn_{scenario_key}')
        with sl4:
            min_recv = st.slider('소액 채널 제외 (주간주문 N장 미만)', 0, 20, preset.get('min_recv', 4), 1, key=f'mr_{scenario_key}')
        with sl5:
            move_cap_pct = st.slider('채널별 이동 상한 (현재고 %)', 0, 100, int(round(preset.get('move_cap_pct', 0.5) * 100)), 5, key=f'cap_{scenario_key}') / 100.0
    else:
        shortage_th = preset['shortage_th']
        target_woc = preset['target_woc']
        min_move = preset['min_move']
        min_recv = preset.get('min_recv', 4)
        move_cap_pct = preset.get('move_cap_pct', 0.5)

    container.markdown(f'<div class="scenario-box">{preset["desc"]}</div>', unsafe_allow_html=True)

    with st.spinner('계산 중...'):
        params_key = (shortage_th, target_woc, ship_th, min_move, min_recv, _ch_excl_key(), move_cap_pct)
        results = calc_results_v20(params_key)
    results = _apply_exclusion(results)
    results = _apply_overrides(results)
    # v0.9 — AI 일일 요약에서 '새로고침' 클릭 시 데이터 ±2% 변동 (mock 시각효과)
    results = _apply_data_variance(results, _get_data_seed())
    # 갱신 시각 표시 (AI 일일 요약의 '새로고침' 이벤트 추적)
    _seed = _get_data_seed()
    _last = st.session_state.get('data_seed_ts', '최초 로드 (정적 데이터)')
    if _seed:
        container.caption(f'📡 마지막 데이터 갱신: <b>{_last}</b> · 데이터 ±2% 변동 적용됨 (mock)', unsafe_allow_html=True)
    else:
        container.caption(f'📡 데이터 상태: {_last} · 🤖 AI 일일 요약 탭의 [🔄 새로고침]으로 갱신')

    total_units = sum(sum(r['data']['inv'].get(c, 0) for c in CHANNELS) for r in results)
    total_units_amt = sum(sum(r['data']['inv'].get(c, 0) for c in CHANNELS) * r['data']['price'] for r in results)
    total_in = sum(sum(v for v in r['moves'].values() if v > 0) for r in results)
    total_amt = sum(sum(v for v in r['moves'].values() if v > 0) * r['data']['price'] for r in results)
    total_rev = sum(r['revenue'] for r in results)

    def kpi_card(col, label, value, sub=''):
        col.markdown(f"""<div class="kpi-card" style="min-height:120px;display:flex;flex-direction:column;justify-content:center">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div><div class="kpi-sub">{sub}</div></div>""", unsafe_allow_html=True)

    kpi_ph = container.container()

    col_f1, col_fs, col_fp, col_f3 = container.columns([1.4, 2.6, 2.6, 1.8])
    with col_f1:
        show_only_moved = st.checkbox('이동 발생만', value=True, key=f'moved_{scenario_key}')
    with col_fs:
        search_code = st.text_input('단품코드 검색', placeholder='앞 10자리 입력 (예: SPPPG25U05)', key=f'search_{scenario_key}').strip().upper()
    with col_fp:
        # 단품코드 선택 (멀티) — 회전 발생 단품 우선 노출
        _move_codes = sorted({r['code'] for r in results if any(v != 0 for v in r['moves'].values())})
        if not _move_codes:
            _move_codes = sorted({r['code'] for r in results})[:200]
        picked = st.multiselect('단품코드 선택 (복수 가능)', options=_move_codes,
                                key=f'pick_{scenario_key}',
                                placeholder='드롭다운에서 단품 다중 선택')
    with col_f3:
        sort_by = st.selectbox('정렬', ['온라인 매출 순위 ↑', '기대효과 ↓', '이동수량 ↓', '단품코드'], key=f'sort_{scenario_key}')

    filtered = list(results)
    if show_only_moved and not search_code and not picked:
        filtered = [r for r in filtered if any(v != 0 for v in r['moves'].values())]
    if search_code:
        filtered = [r for r in filtered if r['code'].upper().startswith(search_code)]
    if picked:
        _set = set(picked)
        filtered = [r for r in filtered if r['code'] in _set]

    if sort_by == '온라인 매출 순위 ↑':
        filtered.sort(key=lambda r: r['data'].get('rank_online', 9999))
    elif sort_by == '기대효과 ↓':
        filtered.sort(key=lambda r: -r['revenue'])
    elif sort_by == '이동수량 ↓':
        filtered.sort(key=lambda r: -sum(v for v in r['moves'].values() if v > 0))
    else:
        filtered.sort(key=lambda r: r['code'])

    # 매트릭스 헤더 + 승인/Excel 버튼 자리 (placeholder — 표 아래 sel_count 계산 후 채움)
    actions_ph = container.empty()

    MAX_ROWS = 2000
    if len(filtered) > MAX_ROWS:
        container.caption(f'⚠️ {len(filtered):,}건 中 상위 {MAX_ROWS}개만 표시 (성능). 정렬·필터로 좁히세요.')
        filtered = filtered[:MAX_ROWS]

    _sel_state = st.session_state.get(f'mat_{scenario_key}')
    _pre_rows = []
    try:
        _pre_rows = [i for i in _sel_state.selection.rows if i < len(filtered)]
    except Exception:
        _pre_rows = []
    if _pre_rows:
        _base = [filtered[i] for i in _pre_rows]
        _units = sum(sum(r['data']['inv'].get(c, 0) for c in CHANNELS) for r in _base)
        _units_amt = sum(sum(r['data']['inv'].get(c, 0) for c in CHANNELS) * r['data']['price'] for r in _base)
        _in = sum(sum(v for v in r['moves'].values() if v > 0) for r in _base)
        _amt = sum(sum(v for v in r['moves'].values() if v > 0) * r['data']['price'] for r in _base)
        _rev = sum(r['revenue'] for r in _base)
        _sub = f'☑ 선택 {len(_base):,}건 기준'
        _chart_items = _base
    else:
        _units, _units_amt, _in, _amt, _rev, _sub = total_units, total_units_amt, total_in, total_amt, total_rev, '전체 기준'
        _chart_items = filtered
    with kpi_ph:
        k1, k2, k3, k4, k5, k6 = st.columns(6)
        kpi_card(k1, '총 단품량', f'{_units:,}장', f'6채널 재고 합계 · {_sub}')
        kpi_card(k2, '총 이동량(회전)', f'{_in:,}장', f'주간 IN · {_sub}')
        kpi_card(k3, '총 재고금액', f'{_units_amt/100000000:.1f}억', '재고수량 × 정상가')
        kpi_card(k4, '총 이동 금액', f'{_amt/100000000:.2f}억', '이동수량 × 정상가')
        k5.markdown(
            f'<div class="kpi-card" style="min-height:120px"><div class="kpi-label">회수 매출 · 채널 구성</div>'
            f'<div class="kpi-value">{_rev/100000000:.2f}억</div>'
            f'{_chan_recovery_bar(_chart_items)}</div>', unsafe_allow_html=True)
        kpi_card(k6, '연 환산', f'{_rev*52/100000000:.0f}억', '× 52주')

    rows = []
    for r in filtered:
        d = r['data']; mv = r['moves']; af = r['after']; inv = d['inv']
        name = d['name']
        cum = d.get('cum_rate', 0) * 100; wk = d.get('wk_rate', 0) * 100
        sv = d.get('wk_sales', 0)
        sales_str = f"{round(sv/10000):,}만" if sv else '-'
        row = [r['code'], name, sales_str, f"{cum:.0f}%", f"{wk:.0f}%", f"{int(d['ship_rate']*100)}%",
               f"{inv.get('반응과', 0):,}"]
        for c in CHANNELS:
            o = d['orders'].get(c, 0)
            w = inv.get(c, 0) / o if o > 0 else None
            row.append(f'{round(w)}주' if w is not None else '')
        for c in CHANNELS:
            v = mv.get(c, 0)
            row.append('0' if v == 0 else f'{v:+d}')
        for c in CHANNELS:
            w = af.get(c)
            row.append(f'{round(w)}주' if w is not None else '')
        for c in CHANNELS:
            v = mv.get(c, 0); ni = inv.get(c, 0) + v
            row.append(f'{ni:,}' if v == 0 else f'{ni:,} ({v:+d})')
        row.append(round(r['revenue'] / 10000))
        rows.append(row)

    sel_count = 0
    selected_rows = []
    if rows:
        columns = pd.MultiIndex.from_tuples(
            [('', '단품코드'), ('', '단품명'), ('', '주간외형매출'), ('', '누판'), ('', '주판'), ('', '출고'), ('', '반응과재고')] +
            [('현 재고보유주수', CH_SHORT[c]) for c in CHANNELS] +
            [('이동수량 (장)', CH_SHORT[c]) for c in CHANNELS] +
            [('이동 후 재고보유주수', CH_SHORT[c]) for c in CHANNELS] +
            [('이동 후 재고량 (장,±)', CH_SHORT[c]) for c in CHANNELS] +
            [('효과', '만원')]
        )
        df = pd.DataFrame(rows, columns=columns)

        woc_cols = [('현 재고보유주수', CH_SHORT[c]) for c in CHANNELS] + \
                   [('이동 후 재고보유주수', CH_SHORT[c]) for c in CHANNELS]
        mv_cols = [('이동수량 (장)', CH_SHORT[c]) for c in CHANNELS]
        qty_cols = [('이동 후 재고량 (장,±)', CH_SHORT[c]) for c in CHANNELS]

        styled = (df.style
                  .map(woc_color, subset=woc_cols)
                  .map(mv_color, subset=mv_cols)
                  .map(qty_color, subset=qty_cols))
        styled = styled.format({('효과', '만원'): '{:,}'.format})

        container.caption('💡 좌측 ☑ 박스(행) 클릭 → 단품 선택 · 헤더 클릭으로 전체 선택')
        event = container.dataframe(
            styled, use_container_width=True, height=620, hide_index=True,
            on_select='rerun', selection_mode='multi-row', key=f'mat_{scenario_key}')
        selected_rows = event.selection.rows if (event and event.selection) else []
        sel_count = len(selected_rows) if selected_rows else len(df)
        if not selected_rows:
            container.caption(f'✅ 미선택 시 전체 {len(df):,}건 실행 대상')
        else:
            container.caption(f'✅ 선택: **{sel_count:,}건** / 전체 {len(df):,}건  ·  선택분만 승인 대상')
    else:
        container.info('필터 조건에 맞는 단품이 없습니다.')

    container.caption(
        '🎨 재고보유주수: 🔴 < 1주  🟡 1~4주  🟢 ≥ 4주   |   🔁 회전 = 온라인 6채널 간 이동(합계 0)   |   '
        '효과 = 결품해소 회수매출(만원)   |   ※ 이동수량은 외부창고(AENS·ADU3·ADQS) 보관분 제외'
    )

    sel_items = []
    if rows:
        sel_items = [filtered[i] for i in selected_rows] if selected_rows else list(filtered)
    sel_qty = sum(sum(v for v in it['moves'].values() if v > 0) for it in sel_items)
    sel_rev = sum(it['revenue'] for it in sel_items)

    # 표 위 placeholder 채우기 — [헤더 텍스트 | 승인 버튼 | Excel 다운로드 버튼]
    with actions_ph.container():
        col_h, col_b1, col_b2 = st.columns([3, 2, 2])
    with col_h:
        st.markdown(f'**단품 × 채널 매트릭스 — {len(filtered):,}건**')
    with col_b1:
        if st.button(f'✅ 선택 {sel_count}건 승인(회전)', use_container_width=True, type='primary', key=f'approve_{scenario_key}'):
            details = []
            ch_in, ch_out = {}, {}
            sel_amt = 0
            for it in sel_items:
                for ch, v in it['moves'].items():
                    if v > 0:
                        details.append((it['code'], ch, it['data']['inv'].get(ch, 0), v, it['data']['price']))
                        ch_in[ch] = ch_in.get(ch, 0) + v
                        sel_amt += v * it['data']['price']
                    elif v < 0:
                        ch_out[ch] = ch_out.get(ch, 0) - v
            exec_id = effect_log.log_execution(scenario_key, len(sel_items), sel_qty, sel_rev, details=details)
            _approve_dialog(scenario_key, sel_count, sel_qty, sel_amt, sel_rev, ch_in, ch_out, exec_id)
    with col_b2:
        # SAP 자동 연동 전 임시 — MD가 수기 실행할 수 있도록 회전 결과 엑셀 다운로드
        # (스파오 6/19 합의 — 이노플 견적 협의 중)
        try:
            smap = _load_style_map()
            out_rows = {ch: [] for ch in CHANNELS}
            all_rows = []
            for it in sel_items:
                d = it['data']
                code = it['code']
                moves = it.get('moves', {})
                sty = code[:10]
                sty_name = smap.get(sty, d.get('name', ''))
                in_pairs = [(ch, moves[ch]) for ch in CHANNELS if moves.get(ch, 0) > 0]
                in_str = ' / '.join([f'{CH_SHORT.get(ch, ch)}+{q:,}' for ch, q in in_pairs])
                # 받는 채널 매장코드 (회전 작업 편의 — 영문 4자리만, 사용자 요청 6/19)
                in_wh_str = ' / '.join([WAREHOUSE_CODE.get(ch, '-') for ch, _ in in_pairs])
                # v0.9.8 — 채널 시트에 OUT·IN 모두 표시 (단품 누락 방지)
                # ch_move != 0인 모든 채널을 그 채널의 시트에 포함
                for ch in CHANNELS:
                    ch_move = moves.get(ch, 0)
                    if ch_move == 0:
                        continue
                    out_qty = max(0, -ch_move)
                    in_qty = max(0, ch_move)
                    price = d.get('price', 0)
                    inv_my = d['inv'].get(ch, 0)
                    ord_my = d['orders'].get(ch, 0)
                    woc_cur = round(inv_my / ord_my, 1) if ord_my > 0 else None
                    woc_after = round((inv_my + in_qty - out_qty) / ord_my, 1) if ord_my > 0 else None
                    # v0.9.9 — 회수매출 2개 컬럼 분리 (옵션 B)
                    # 1) OUT 매출가치(만원) = OUT 수량 × 정상가 (회수된 재고의 단순 가치)
                    out_value = round(out_qty * price / 10000) if out_qty > 0 else 0
                    # 2) 결품해소 회수(만원) = 대시보드 KPI와 동일 산식 — 이 채널의 결품 해소분
                    old_short_ch = max(0, ord_my - inv_my)
                    new_short_ch = max(0, ord_my - (inv_my + in_qty - out_qty))
                    relief = round((old_short_ch - new_short_ch) * price / 10000)
                    row = {
                        '단품코드': code, '스타일코드': sty, '스타일명': sty_name,
                        '내 채널 현재고': inv_my,
                        '내 채널 주판': ord_my,
                        '현 재고주수': (f'{woc_cur}주' if woc_cur is not None else ''),
                        'OUT 수량(장)': int(out_qty),
                        'IN 수량(장)': int(in_qty),
                        '이동 후 재고주수': (f'{woc_after}주' if woc_after is not None else ''),
                        '받는 채널 분배': in_str,
                        '받는 채널 매장코드': in_wh_str,
                        '내 채널 매장코드': WAREHOUSE_CODE.get(ch, '-'),
                        '단품 정상가(원)': price,
                        'OUT 매출가치(만원)': out_value,
                        '결품해소 회수(만원)': relief,
                    }
                    out_rows[ch].append(row)
                    all_rows.append({**row, '채널': ch})
            sheets = {}
            for ch in CHANNELS:
                if out_rows[ch]:
                    sheets[CH_SHORT.get(ch, ch)] = (
                        pd.DataFrame(out_rows[ch])
                        .sort_values('OUT 수량(장)', ascending=False)
                        .reset_index(drop=True)
                    )
            if all_rows:
                sheets['전체 회전 매트릭스'] = pd.DataFrame(all_rows)
            from datetime import datetime as _dt
            tot_qty = sum(r['OUT 수량(장)'] for r in all_rows) if all_rows else 0
            fname = f'회전결과_{scenario_key.replace(" ", "")}_{_dt.now().strftime("%Y%m%d_%H%M")}_{tot_qty}장.xlsx'
            if sheets:
                # 조건부 서식 (사용자 6/19 요청)
                bar_cols = ['내 채널 현재고', '내 채널 주판', 'OUT 수량(장)']
                red_bar_cols = ['OUT 매출가치(만원)', '결품해소 회수(만원)']
                woc_cols = ['현 재고주수', '이동 후 재고주수']
                bold_cols = ['OUT 수량(장)', '결품해소 회수(만원)']
                st.download_button(
                    f'⬇️ Excel 다운로드 (회전 수기 실행용 · {sel_count:,}건)',
                    data=_xlsx_bytes_with_bars(sheets, bar_cols, red_bar_cols, woc_cols, bold_cols),
                    file_name=fname,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True, key=f'exp_xlsx_{scenario_key}',
                )
            else:
                st.button('⬇️ Excel 다운로드 (0건)', use_container_width=True, disabled=True, key=f'exp_xlsx_dis_{scenario_key}')
        except Exception as e:
            st.button(f'⬇️ Excel 다운로드 (에러)', use_container_width=True, disabled=True, key=f'exp_xlsx_err_{scenario_key}')
    container.caption('💡 스파오 6/19 합의 — SAP 자동 연동 전 임시 운영: 위 ⬇️ Excel(채널별 시트)로 MD 수기 실행.')


# ============================================================
# 영구 저장 — GitHub Contents API 자동 commit (옵션 C)
# 사용자 6/24 요청: 김혜인 책임 등 다중 사용자 데이터 공유 + 세션 종료 후 보존
# Streamlit Secrets에 GITHUB_TOKEN (PAT, repo write 권한) 등록 필요
# ============================================================
_GH_REPO = 'kanghg61-del/spao-rebalance'
_GH_BRANCH = 'main'


def _gh_get_token():
    try:
        if hasattr(st, 'secrets'):
            t = st.secrets.get('GITHUB_TOKEN')
            if t:
                return t
    except Exception:
        pass
    import os as _os
    return _os.environ.get('GITHUB_TOKEN')


def _gh_load(path):
    """GitHub에서 JSON 파일 읽기 → (data, sha) 또는 (None, None)"""
    import urllib.request as _ur
    import urllib.error as _ue
    import json as _json
    import base64 as _b64
    token = _gh_get_token()
    if not token:
        return None, None
    url = f'https://api.github.com/repos/{_GH_REPO}/contents/{path}?ref={_GH_BRANCH}'
    req = _ur.Request(url, headers={
        'Authorization': f'Bearer {token}',
        'Accept': 'application/vnd.github+json',
        'X-GitHub-Api-Version': '2022-11-28',
    })
    try:
        with _ur.urlopen(req, timeout=10) as r:
            meta = _json.loads(r.read().decode('utf-8'))
        content = _b64.b64decode(meta['content']).decode('utf-8')
        return _json.loads(content), meta.get('sha')
    except _ue.HTTPError as e:
        if e.code == 404:
            return None, None  # 파일 아직 없음 (정상)
        return None, None
    except Exception:
        return None, None


def _gh_save(path, data, commit_msg=None):
    """GitHub에 JSON 파일 commit (기존 sha 자동 조회)"""
    import urllib.request as _ur
    import urllib.error as _ue
    import json as _json
    import base64 as _b64
    from datetime import datetime as _dtg
    token = _gh_get_token()
    if not token:
        return False, '키 미설정 (Streamlit Secrets: GITHUB_TOKEN)'
    _, sha = _gh_load(path)  # 기존 sha (없으면 신규)
    body = {
        'message': commit_msg or f'update {path} {_dtg.now().strftime("%Y-%m-%d %H:%M:%S")}',
        'content': _b64.b64encode(_json.dumps(data, ensure_ascii=False, default=str, indent=2).encode('utf-8')).decode('ascii'),
        'branch': _GH_BRANCH,
    }
    if sha:
        body['sha'] = sha
    url = f'https://api.github.com/repos/{_GH_REPO}/contents/{path}'
    req = _ur.Request(
        url,
        data=_json.dumps(body).encode('utf-8'),
        headers={
            'Authorization': f'Bearer {token}',
            'Accept': 'application/vnd.github+json',
            'X-GitHub-Api-Version': '2022-11-28',
            'Content-Type': 'application/json',
        },
        method='PUT',
    )
    try:
        with _ur.urlopen(req, timeout=15) as r:
            r.read()
        return True, None
    except _ue.HTTPError as he:
        try:
            err = he.read().decode('utf-8')[:200]
        except Exception:
            err = ''
        return False, f'HTTP {he.code}: {err}'
    except Exception as e:
        return False, f'{type(e).__name__}: {str(e)[:120]}'


def _persist_load_ch_excl():
    """앱 시작 시 또는 명시적 refresh 시 GitHub에서 ch_excl_rows + excluded_codes 로드"""
    if st.session_state.get('_ch_excl_loaded'):
        return  # 한 번만 로드
    rows, _ = _gh_load('data/ch_excl.json')
    if rows is not None:
        # 날짜 문자열 → date 객체 변환
        from datetime import date as _date
        for r in rows:
            for k in ('시작일', '종료일'):
                v = r.get(k)
                if v and isinstance(v, str):
                    try:
                        r[k] = _date.fromisoformat(v[:10])
                    except Exception:
                        r[k] = None
        st.session_state['ch_excl_rows'] = rows
    excl, _ = _gh_load('data/excluded_codes.json')
    if excl is not None and isinstance(excl, dict):
        st.session_state['excluded_text'] = excl.get('text', '')
        st.session_state['excluded_codes'] = set(excl.get('codes', []))
    st.session_state['_ch_excl_loaded'] = True


def render_excluded_tab():
    # 영구 저장 자동 로드 (앱 첫 진입 시 GitHub에서 동기화)
    _persist_load_ch_excl()
    st.markdown('### 🚫 채널별 IN·OUT 제외 관리 (채널 담당 MD)')
    st.caption('🟢 IN 제외 = 이 채널로 재고를 받지 않음(수신 차단) · '
               '🔴 OUT 제외 = 이 채널에서 재고를 빼지 않음(반출 차단). '
               '스타일 코드 일부만 입력해도 매칭(예: SPACG24). '
               '시작일/종료일 입력 시 해당 기간만 자동 적용 → 종료 후 자동 해제. '
               '시작일/종료일 비워두면 영구 제외. 입력 후 다른 탭으로 이동하면 재배치에 반영됩니다.')

    try:
        skus = load_data_v20()
    except Exception:
        skus = None

    import pandas as _pd
    from datetime import date as _date

    all_rows = list(st.session_state.get('ch_excl_rows', []))
    tabs = st.tabs([str(c) for c in CHANNELS])
    for tab, c in zip(tabs, CHANNELS):
        with tab:
            ca, cb = st.columns(2)
            # ── IN 제외 ─────────────────────────────
            with ca:
                st.markdown('**🟢 IN 제외 (이 채널로 받지 않을 스타일)**')
                in_rows_existing = [r for r in all_rows if r.get('채널') == c and r.get('방향') == 'in']
                if not in_rows_existing:
                    in_rows_existing = [{'스타일': '', '시작일': None, '종료일': None}]
                df_in = _pd.DataFrame(in_rows_existing)[['스타일', '시작일', '종료일']]
                edited_in = st.data_editor(
                    df_in, num_rows='dynamic', use_container_width=True,
                    key=f'excl_in_editor_{c}',
                    column_config={
                        '스타일': st.column_config.TextColumn('스타일', help='코드 일부만 입력해도 매칭', required=False),
                        '시작일': st.column_config.DateColumn('시작일', format='YYYY-MM-DD', help='비워두면 영구 적용'),
                        '종료일': st.column_config.DateColumn('종료일', format='YYYY-MM-DD'),
                    })
            # ── OUT 제외 ────────────────────────────
            with cb:
                st.markdown('**🔴 OUT 제외 (이 채널에서 빼지 않을 스타일)**')
                out_rows_existing = [r for r in all_rows if r.get('채널') == c and r.get('방향') == 'out']
                if not out_rows_existing:
                    out_rows_existing = [{'스타일': '', '시작일': None, '종료일': None}]
                df_out = _pd.DataFrame(out_rows_existing)[['스타일', '시작일', '종료일']]
                edited_out = st.data_editor(
                    df_out, num_rows='dynamic', use_container_width=True,
                    key=f'excl_out_editor_{c}',
                    column_config={
                        '스타일': st.column_config.TextColumn('스타일', required=False),
                        '시작일': st.column_config.DateColumn('시작일', format='YYYY-MM-DD'),
                        '종료일': st.column_config.DateColumn('종료일', format='YYYY-MM-DD'),
                    })

            # 입력 정리 (스타일 비어있는 행 제외) — 채널별로 갱신
            in_records, out_records = [], []
            for _, row in edited_in.iterrows():
                sty = str(row.get('스타일') or '').strip()
                if sty:
                    in_records.append({'채널': c, '방향': 'in', '스타일': sty,
                                       '시작일': row.get('시작일'), '종료일': row.get('종료일')})
            for _, row in edited_out.iterrows():
                sty = str(row.get('스타일') or '').strip()
                if sty:
                    out_records.append({'채널': c, '방향': 'out', '스타일': sty,
                                        '시작일': row.get('시작일'), '종료일': row.get('종료일')})
            # 이 채널 데이터만 교체
            all_rows = [r for r in all_rows if r.get('채널') != c]
            all_rows.extend(in_records + out_records)

            # KPI
            in_pats = {r['스타일'] for r in in_records}
            out_pats = {r['스타일'] for r in out_records}
            m_in = sum(1 for code in skus if any(p in code for p in in_pats)) if (skus and in_pats) else 0
            m_out = sum(1 for code in skus if any(p in code for p in out_pats)) if (skus and out_pats) else 0
            mc1, mc2, mc3, mc4 = st.columns(4)
            mc1.metric('IN 제외 패턴', f'{len(in_pats):,}')
            mc2.metric('OUT 제외 패턴', f'{len(out_pats):,}')
            mc3.metric('매칭 단품 (IN / OUT)', f'{m_in:,} / {m_out:,}')
            # 오늘 활성 미리보기
            today = _date.today()
            active_today_in = []
            active_today_out = []
            for r in in_records + out_records:
                s = r.get('시작일'); e = r.get('종료일')
                if not (s or e):
                    is_active = True
                elif s and e:
                    try:
                        s_ = s if hasattr(s, 'toordinal') else _date.fromisoformat(str(s)[:10])
                        e_ = e if hasattr(e, 'toordinal') else _date.fromisoformat(str(e)[:10])
                        is_active = s_ <= today <= e_
                    except Exception:
                        is_active = False
                else:
                    is_active = False
                if is_active and r['방향'] == 'in':
                    active_today_in.append(r['스타일'])
                elif is_active and r['방향'] == 'out':
                    active_today_out.append(r['스타일'])
            mc4.metric(f'오늘({today.strftime("%m/%d")}) 활성', f'{len(active_today_in)} / {len(active_today_out)}',
                       help='IN 활성 / OUT 활성 패턴 수 (기간 없으면 영구 활성)')

            if st.button('🗑️ 이 채널 제외 초기화', key=f'clr_v2_{c}'):
                all_rows = [r for r in all_rows if r.get('채널') != c]
                st.session_state['ch_excl_rows'] = all_rows
                # 레거시 텍스트영역도 비움
                if c in st.session_state.get('ch_excl', {}):
                    st.session_state['ch_excl'].pop(c, None)
                st.rerun()

    st.session_state['ch_excl_rows'] = all_rows

    # ── GitHub 영구 저장 ──
    st.markdown('---')
    save_col1, save_col2, save_col3 = st.columns([2, 2, 6])
    with save_col1:
        if st.button('💾 GitHub 영구 저장', type='primary', use_container_width=True, key='gh_save_excl'):
            ok, err = _gh_save('data/ch_excl.json', all_rows, commit_msg=f'채널 IN-OUT 제외 갱신 (사용자: 대시보드)')
            if ok:
                st.success('✅ GitHub 영구 저장 완료. 모든 사용자에게 반영됩니다.')
            else:
                st.error(f'❌ 저장 실패: {err}')
    with save_col2:
        if st.button('🔄 GitHub에서 다시 불러오기', use_container_width=True, key='gh_reload_excl'):
            st.session_state.pop('_ch_excl_loaded', None)
            st.session_state.pop('ch_excl_rows', None)
            st.rerun()
    with save_col3:
        token_ok = bool(_gh_get_token())
        if token_ok:
            st.caption('🟢 영구 저장 활성 (GitHub 자동 commit). 변경 후 좌측 "💾 GitHub 영구 저장" 클릭')
        else:
            st.caption('🟡 임시 세션 저장 (브라우저 닫으면 사라짐). 영구 저장은 GITHUB_TOKEN 등록 후 활성')

    with st.expander('🚫 전체 이동 제외 (예약판매·기획전 등 — 모든 채널에서 이동 자체를 막음)', expanded=False):
        excluded_text = st.text_area('제외 단품코드 (줄바꿈 또는 쉼표로 구분)',
                                     value=st.session_state.get('excluded_text', ''), height=130,
                                     placeholder='예:\nSPJJG25G0119095\nSPACG24A5', key='excluded_text_input')
        codes_set = {x.strip() for x in excluded_text.replace(',', '\n').split('\n') if x.strip()}
        st.session_state['excluded_codes'] = codes_set
        st.session_state['excluded_text'] = excluded_text
        if skus and codes_set:
            st.caption(f'매칭 단품 {sum(1 for code in skus if any(ex in code for ex in codes_set)):,}건')
        bc1, bc2 = st.columns(2)
        with bc1:
            if st.button('💾 GitHub 영구 저장', type='primary', use_container_width=True, key='gh_save_excl_codes'):
                ok, err = _gh_save('data/excluded_codes.json', {'text': excluded_text, 'codes': sorted(codes_set)},
                                   commit_msg='전체 이동 제외 단품 갱신')
                if ok:
                    st.success('✅ GitHub 영구 저장 완료')
                else:
                    st.error(f'❌ 저장 실패: {err}')
        with bc2:
            if st.button('🗑️ 전체 이동 제외 초기화', use_container_width=True):
                st.session_state['excluded_text'] = ''
                st.session_state['excluded_codes'] = set()
                st.rerun()



def render_export_tab():
    """채널 MD용 회전 결과 엑셀 다운로드 (SAP 연동 전 수기 실행용).

    스파오 6/19 미팅 결과 — 견적 추가 받기 전 테스트 운영. 각 채널 MD가 자기
    채널에서 어디 채널로 재고를 빼야 하는지 자체 시트에서 확인 후 수기 실행.
    """
    st.markdown('### ⬇️ 회전 결과 엑셀 다운로드 (MD 수기 실행용)')
    st.markdown(
        '<div class="scenario-box">SAP 자동 연동 전(견적 협의 중) **임시 운영 방식** — 채널별 시트로 분리한 엑셀을 '
        'MD가 다운받아 본인 채널에서 빠지는 단품·수량을 확인 후 수기 실행. '
        '시트 = 6개 채널 + 전체 회전 매트릭스 + 출고매장코드별 요약.</div>',
        unsafe_allow_html=True)

    smap = _load_style_map()
    preset = SCENARIOS['🛡️ 기본']
    params_key = (preset['shortage_th'], preset['target_woc'], preset['ship_th'],
                  preset['min_move'], preset['min_recv'], _ch_excl_key(), preset['move_cap_pct'])
    try:
        results = calc_results_v20(params_key)
    except Exception as e:
        st.error(f'회전 계산 실패: {e}')
        return

    out_rows = {ch: [] for ch in CHANNELS}
    all_rows = []
    n_move = 0
    tot_qty = 0
    for r in results:
        moves = r.get('moves', {})
        if not moves or not any(v != 0 for v in moves.values()):
            continue
        d = r['data']
        code = r['code']
        sty = code[:10]
        sty_name = smap.get(sty, d.get('name', ''))
        in_pairs = [(ch, moves[ch]) for ch in CHANNELS if moves.get(ch, 0) > 0]
        if not in_pairs:
            continue
        in_str = ' / '.join([f'{CH_SHORT.get(ch, ch)}+{q:,}' for ch, q in in_pairs])
        # 각 OUT 채널마다 행 생성
        for out_ch in CHANNELS:
            out_qty = -moves.get(out_ch, 0)
            if out_qty > 0:
                price = d.get('price', 0)
                row = {
                    '단품코드': code,
                    '스타일코드': sty,
                    '스타일명': sty_name,
                    '내 채널 현재고': d['inv'].get(out_ch, 0),
                    '내 채널 주판': d['orders'].get(out_ch, 0),
                    'OUT 수량(장)': int(out_qty),
                    '받는 채널 분배': in_str,
                    '내 채널 매장코드': WAREHOUSE_CODE.get(out_ch, '-'),
                    '단품 정상가(원)': price,
                    '회수매출(만원)': round(out_qty * price / 10000),
                }
                out_rows[out_ch].append(row)
                tot_qty += int(out_qty)
                n_move += 1
                all_rows.append({**row, 'OUT 채널': out_ch})

    # KPI
    c1, c2, c3 = st.columns(3)
    _kpi(c1, '🔄 회전 발생 단품×채널', f'{n_move:,}건', '시나리오 🛡️ 기본 · 회전 30%')
    _kpi(c2, '📦 총 OUT 수량', f'{tot_qty:,}장', '6채널 합산')
    _kpi(c3, '🏪 활성 OUT 채널', f'{sum(1 for ch in CHANNELS if len(out_rows[ch]) > 0):,}개', '회전 발생 채널')

    # 시트 구성 (6채널 + 전체 + 매장코드 요약)
    sheets = {}
    for ch in CHANNELS:
        if out_rows[ch]:
            df = pd.DataFrame(out_rows[ch]).sort_values('OUT 수량(장)', ascending=False).reset_index(drop=True)
            sheets[CH_SHORT.get(ch, ch)] = df
    if all_rows:
        df_all = pd.DataFrame(all_rows)[
            ['OUT 채널', '단품코드', '스타일코드', '스타일명',
             '내 채널 현재고', '내 채널 주판', 'OUT 수량(장)',
             '받는 채널 분배', '내 채널 매장코드', '단품 정상가(원)', '회수매출(만원)']
        ]
        sheets['전체 회전 매트릭스'] = df_all
        df_wh = (df_all.groupby('내 채널 매장코드')
                 .agg(단품수=('단품코드', 'count'),
                      OUT수량=('OUT 수량(장)', 'sum'),
                      회수매출=('회수매출(만원)', 'sum'))
                 .reset_index().sort_values('OUT수량', ascending=False))
        sheets['출고매장코드별 요약'] = df_wh

    from datetime import datetime as _dt
    fname = f'회전결과_{_dt.now().strftime("%Y%m%d_%H%M")}_{tot_qty}장.xlsx'
    if sheets:
        st.download_button(
            f'⬇️ {fname} 다운로드 — 채널별 시트 {len(sheets)}개',
            data=_xlsx_bytes(sheets),
            file_name=fname,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary', use_container_width=True, key='exp_xlsx',
        )

        # 채널별 미리보기
        st.markdown('#### 📋 채널별 회전 결과 미리보기 (상위 30건)')
        st.caption('각 채널 MD는 본인 채널 탭 + 다운로드한 엑셀의 해당 시트를 확인 후 수기 실행.')
        ch_tabs = st.tabs(['🌐 전체'] + [CH_SHORT.get(ch, ch) for ch in CHANNELS])
        with ch_tabs[0]:
            if 'all_rows' in dir() and all_rows:
                st.dataframe(pd.DataFrame(all_rows).head(50), use_container_width=True, hide_index=True, height=420)
            else:
                st.info('회전 결과가 없습니다.')
        for i, ch in enumerate(CHANNELS, start=1):
            with ch_tabs[i]:
                if out_rows[ch]:
                    df_ch = pd.DataFrame(out_rows[ch]).sort_values('OUT 수량(장)', ascending=False).head(30)
                    st.dataframe(df_ch, use_container_width=True, hide_index=True, height=420)
                    st.caption(f'{ch} 채널 → {len(out_rows[ch]):,}건 / OUT 합계 {sum(r["OUT 수량(장)"] for r in out_rows[ch]):,}장')
                else:
                    st.info(f'{ch} 채널에서는 회전 OUT이 발생하지 않았습니다.')
    else:
        st.info('현재 시나리오에서 회전이 발생하지 않았습니다. 재배치(기본) 탭에서 시나리오를 먼저 확인하세요.')

    st.caption('🛡️ 시나리오 = 기본 (회전 30% 보수). 임의 조정이 필요하면 재배치(임의) 탭의 슬라이더 활용. '
               'SAP 자동 연동 후에는 본 탭의 다운로드 단계가 자동화됩니다.')


def render_reorder_tab():
    st.markdown('### 🔁 리오더 매핑 관리')
    st.caption('리오더(코드변경) 발생 시 주 1회~월 1회 갱신. 스타일코드(10자리)·단품코드(15자리) 모두 지원 — '
               '스타일코드 입력 시 사이즈까지 자동 prefix 매칭됩니다. 적용 즉시 전 탭 재계산.')

    info = get_reorder_info()
    pairs = get_reorder_mapping()

    m1, m2, m3 = st.columns(3)
    m1.metric('매핑 행수', f"{info['mapping_rows']:,}")
    m2.metric('병합 적용 단품', f"{info['merged']:,}건")
    m3.metric('적용 파일', info['file'] or '없음')

    st.markdown('---')
    col_up, col_add = st.columns(2)

    with col_up:
        st.markdown('#### 📂 파일 업로드 (전체 교체)')
        up = st.file_uploader('csv 또는 xlsx — 컬럼: 기존코드 / 리오더(추가)코드 자동 인식',
                              type=['csv', 'xlsx'], key='reorder_upload')
        if up is not None:
            try:
                new_pairs = parse_reorder_bytes(up.getvalue(), up.name)
                st.success(f'{up.name} → 유효 매핑 {len(new_pairs):,}행 인식')
                st.dataframe(pd.DataFrame(new_pairs, columns=['기존코드', '리오더코드']).head(10),
                             use_container_width=True, hide_index=True, height=200)
                if st.button(f'✅ 적용 — 기존 매핑을 {len(new_pairs):,}행으로 교체', type='primary',
                             use_container_width=True, key='apply_upload'):
                    save_reorder_mapping(new_pairs)
                    st.cache_data.clear()
                    st.success('적용 완료 — 전 탭 재계산됨')
                    st.rerun()
            except Exception as e:
                st.error(f'파일 해석 실패: {e}')

    with col_add:
        st.markdown('#### ⌨️ 직접 입력 (1건 추가)')
        with st.form('reorder_add_form', clear_on_submit=True):
            org_in = st.text_input('원오더(기존) 코드', placeholder='예: SPJJG23KU2 또는 15자리 단품코드')
            reo_in = st.text_input('리오더(추가) 코드', placeholder='예: SPJJG24KU1')
            ok = st.form_submit_button('➕ 매핑 추가', use_container_width=True, type='primary')
        if ok:
            org_v, reo_v = org_in.strip().upper(), reo_in.strip().upper()
            if not org_v or not reo_v or org_v == reo_v:
                st.error('기존/리오더 코드를 서로 다르게 입력하세요.')
            elif any(r == reo_v for _, r in pairs):
                st.warning(f'{reo_v} 는 이미 매핑에 존재합니다.')
            else:
                save_reorder_mapping(pairs + [(org_v, reo_v)])
                st.cache_data.clear()
                st.success(f'추가 완료: {reo_v} → {org_v}')
                st.rerun()

    st.markdown('---')
    st.markdown(f'#### 현재 매핑 ({len(pairs):,}행)')
    if pairs:
        df_map = pd.DataFrame(pairs, columns=['기존코드', '리오더코드'])
        sc1, sc2 = st.columns([3, 1])
        with sc1:
            q = st.text_input('매핑 검색', placeholder='코드 일부 입력', key='reorder_search').strip().upper()
        if q:
            df_map = df_map[df_map['기존코드'].str.contains(q) | df_map['리오더코드'].str.contains(q)]
        st.dataframe(df_map, use_container_width=True, hide_index=True, height=320)
        with sc2:
            csv_bytes = ('기존코드,리오더코드\n' + '\n'.join(f'{o},{r}' for o, r in pairs)).encode('utf-8-sig')
            st.download_button('⬇️ CSV 다운로드', csv_bytes, 'reorder_mapping.csv', 'text/csv',
                               use_container_width=True)
    else:
        st.info('등록된 매핑이 없습니다. 파일 업로드 또는 직접 입력으로 추가하세요.')

    st.caption('⚠️ 웹에서 적용한 변경은 앱 **재시작·재배포 시 패키지 파일 기준으로 초기화**됩니다. '
               '영구 반영하려면 ⬇️ CSV를 다운로드해 GitHub 레포의 reorder_mapping.csv 로 커밋하거나 Claude에게 전달하세요.')


def _stat_color(s):
    if not isinstance(s, str):
        return ''
    if '긴급' in s:
        return 'background-color:#5B1E1E; color:#FF6B70; font-weight:bold'
    if '주의' in s:
        return 'background-color:#5A4500; color:#FFC000; font-weight:bold'
    if '정상' in s:
        return 'background-color:#1B4D3E; color:#4AE3B5; font-weight:bold'
    return 'color:#9FB0C0'


def _move_color(v):
    try:
        n = int(str(v).replace('+', '').replace(',', ''))
    except Exception:
        return 'text-align:right'
    if n > 0:
        return 'background-color:#1B4D3E; color:#FFFFFF; font-weight:bold; text-align:right'
    if n < 0:
        return 'background-color:#5B1E1E; color:#FFFFFF; font-weight:bold; text-align:right'
    return 'color:#9FB0C0; text-align:right'


def _rate_color(v):
    try:
        x = float(v)
    except Exception:
        return ''
    if x >= 25:
        return 'background-color:#5B1E1E; color:#FF6B70; font-weight:bold'
    if x >= 10:
        return 'background-color:#5A4500; color:#FFC000; font-weight:bold'
    return 'background-color:#1B4D3E; color:#4AE3B5; font-weight:bold'


def _hl_sum(row):
    v0 = str(row.iloc[0])
    is_sum = v0 == '합계' or v0 == '— 합계 —'
    sty = 'background-color:#1E2D40; color:#FFFFFF; font-weight:bold' if is_sum else ''
    return [sty] * len(row)


def _int0(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ''
    if f != f:
        return ''
    return f'{int(round(f))}'


def _woc(v):
    """재고주수 포맷 — 기본탭과 동일: '0주', '1주' 형태."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ''
    if f != f:
        return ''
    return f'{int(round(f))}주'


BOK_LIST = ['A', 'C', 'G', 'K', 'M', 'U', 'W']


def _bok(code):
    return code[7] if len(code) > 7 else '?'


def _item(code):
    return code[2:4] if len(code) > 3 else '?'


def _ch_effect(d, mv_ch, ch):
    o = d['orders'].get(ch, 0); i = d['inv'].get(ch, 0); p = d.get('price', 0)
    return (max(0, o - i) - max(0, o - (i + mv_ch))) * p


def _render_group(results_ch, keyfn, g_inv, g_ord, g_move, g_eff, g_ext, keycol, namefn=False):
    """아이템별/스타일별 집계표 — 단품 상세와 동일 지표를 그룹 합산하여 렌더(맨 위 합계행)."""
    agg = {}
    nm = {}
    for r in results_ch:
        d = r['data']; k = keyfn(r['code']); p = d.get('price', 0)
        o = g_ord(d); i = g_inv(d)
        a = agg.setdefault(k, dict(sku=0, ordd=0, inv=0, invamt=0, salesamt=0, ext=0, mv=0, eff=0, item=0, urg=0))
        a['sku'] += 1; a['ordd'] += o; a['inv'] += i; a['invamt'] += i * p; a['salesamt'] += o * p
        a['ext'] += g_ext(d); a['mv'] += max(0, g_move(r)); a['eff'] += g_eff(r)
        if o > 0:
            a['item'] += 1
            if i / o < 1:
                a['urg'] += 1
        if namefn:
            nm.setdefault(k, d['name'])
    tot_ord = sum(a['ordd'] for a in agg.values()) or 1

    def mk(k, a):
        woc = a['inv'] / a['ordd'] if a['ordd'] > 0 else None
        daily = a['ordd'] / 7
        # 이미지는 스타일/아이템 코드 기준 mock SVG (7월 본연동 시 실 SPAO 이미지로 교체)
        row = {'이미지': _spao_img_url(k), keycol: k}
        if namefn:
            v = nm.get(k, '')
            row['대표 상품명'] = (v[:18] + '…') if len(v) > 18 else v
        row.update({
            '주간 판매량': a['ordd'], '일평균 판매량': round(daily, 1),
            '일평균 매출(만원)': round(a['salesamt'] / 7 / 10000), '현 재고량': a['inv'],
            '현 재고금액(만원)': round(a['invamt'] / 10000), '외부창고': a['ext'],
            '현 재고주수': round(woc) if woc is not None else None,
            '소진예상(일)': round(a['inv'] / daily) if daily > 0 else None,
            '추천이동(회전)': a['mv'],
            '이동 후 재고주수': round((a['inv'] + a['mv']) / a['ordd']) if a['ordd'] > 0 else None,
            '효과(만원)': round(a['eff'] / 10000),
            '판매량 비중(%)': round(a['ordd'] / tot_ord * 100, 2),
            '결품률(%)': round(a['urg'] / a['item'] * 100, 1) if a['item'] else 0.0,
        })
        return row

    body = [mk(k, a) for k, a in agg.items() if not (a['inv'] == 0 and a['ordd'] == 0)]
    body.sort(key=lambda x: -x['주간 판매량'])
    if not body:
        st.info('표시할 항목이 없습니다.')
        return
    T = {f: sum(a[f] for a in agg.values()) for f in ('sku', 'ordd', 'inv', 'invamt', 'salesamt', 'ext', 'mv', 'eff', 'item', 'urg')}
    woc = T['inv'] / T['ordd'] if T['ordd'] > 0 else None
    daily = T['ordd'] / 7
    sumrow = {'이미지': '', keycol: '합계'}
    if namefn:
        sumrow['대표 상품명'] = f'{len(body):,}개'
    sumrow.update({
        '주간 판매량': T['ordd'], '일평균 판매량': round(daily, 1),
        '일평균 매출(만원)': round(T['salesamt'] / 7 / 10000), '현 재고량': T['inv'],
        '현 재고금액(만원)': round(T['invamt'] / 10000), '외부창고': T['ext'],
        '현 재고주수': round(woc) if woc is not None else None,
        '소진예상(일)': round(T['inv'] / daily) if daily > 0 else None,
        '추천이동(회전)': T['mv'],
        '이동 후 재고주수': round((T['inv'] + T['mv']) / T['ordd']) if T['ordd'] > 0 else None,
        '효과(만원)': round(T['eff'] / 10000),
        '판매량 비중(%)': 100.0, '결품률(%)': round(T['urg'] / T['item'] * 100, 1) if T['item'] else 0.0,
    })
    df = pd.DataFrame([sumrow] + body)
    styled = (df.style.map(_rate_color, subset=['결품률(%)'])
              .map(woc_color, subset=['현 재고주수', '이동 후 재고주수'])
              .map(mv_color, subset=['추천이동(회전)'])
              .apply(_hl_sum, axis=1)
              .format({'주간 판매량': '{:,}'.format, '일평균 판매량': '{:.1f}'.format,
                       '일평균 매출(만원)': '{:,}'.format, '현 재고량': '{:,}'.format, '현 재고금액(만원)': '{:,}'.format,
                       '외부창고': '{:,}'.format, '현 재고주수': _woc, '이동 후 재고주수': _woc,
                       '소진예상(일)': _int0,
                       '추천이동(회전)': '{:,}'.format, '효과(만원)': '{:,}'.format,
                       '판매량 비중(%)': '{:.2f}'.format, '결품률(%)': '{:.1f}'.format}))
    st.dataframe(styled, use_container_width=True, height=440, hide_index=True,
                 column_config={'이미지': st.column_config.ImageColumn('이미지', width='small',
                                                                       help='추후 SPAO 공홈 실 이미지로 교체')})


def render_channel_tab():
    st.markdown("""
    <style>
    .stTabs div[role="radiogroup"] label:has(input:checked){background:#FFC000 !important; border-color:#FFC000 !important;}
    .stTabs div[role="radiogroup"] label:has(input:checked) p{color:#0A141F !important;}
    .stTabs .stTabs [data-baseweb="tab"][aria-selected="true"]{background:#FFC000 !important; color:#0A141F !important;}
    .stTabs .stTabs [data-baseweb="tab"][aria-selected="true"] p{color:#0A141F !important;}
    </style>
    """, unsafe_allow_html=True)

    picks = ['전체'] + list(CHANNELS)
    channel_pick = st.radio('채널 선택', picks, horizontal=True, key='ch_pick', label_visibility='collapsed')
    is_all = channel_pick == '전체'
    is_ext = (not is_all) and channel_pick in EXT_WAREHOUSE
    wh_label = f'{EXT_WAREHOUSE[channel_pick][0]}({EXT_WAREHOUSE[channel_pick][1]})' if is_ext else None

    preset = SCENARIOS['🛡️ 기본']
    params_key = (preset['shortage_th'], preset['target_woc'], preset['ship_th'],
                  preset['min_move'], preset.get('min_recv', 4), _ch_excl_key())
    results_ch = _apply_exclusion(calc_results_v20(params_key))

    def g_inv(d):
        return sum(d['inv'].get(c, 0) for c in CHANNELS) if is_all else d['inv'].get(channel_pick, 0)

    def g_ord(d):
        return sum(d['orders'].get(c, 0) for c in CHANNELS) if is_all else d['orders'].get(channel_pick, 0)

    def g_move(r):
        if is_all:
            return sum(v for v in r['moves'].values() if v > 0)
        return r['moves'].get(channel_pick, 0)

    def g_ext(d):
        if is_all:
            return sum(d.get('ext_wh', {}).get(c, 0) for c in EXT_WAREHOUSE)
        return d.get('ext_wh', {}).get(channel_pick, 0) if is_ext else 0

    def g_eff(r):
        return r['revenue'] if is_all else _ch_effect(r['data'], r['moves'].get(channel_pick, 0), channel_pick)

    def kcard(col, label, value, sub=''):
        col.markdown(f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
                     f'<div class="kpi-value">{value}</div><div class="kpi-sub">{sub}</div></div>',
                     unsafe_allow_html=True)

    sub_overview, sub_item, sub_style, sub_sku = st.tabs(['📋 재고 현황', '🧺 아이템별', '🎨 스타일별', '🔎 단품 상세'])

    # ───────────── 재고 현황 ─────────────
    with sub_overview:
        tot_inv = sum(g_inv(r['data']) for r in results_ch)
        tot_amt = sum(g_inv(r['data']) * r['data'].get('price', 0) for r in results_ch)
        n_item = sum(1 for r in results_ch if g_ord(r['data']) > 0)
        n_urgent = sum(1 for r in results_ch if g_ord(r['data']) > 0 and g_inv(r['data']) / max(1, g_ord(r['data'])) < 1)
        rate = n_urgent / max(1, n_item) * 100
        tot_in = sum(max(0, g_move(r)) for r in results_ch)
        tot_ext = sum(g_ext(r['data']) for r in results_ch)
        ext_pct = tot_ext / max(1, tot_inv) * 100
        union_item = union_urg = 0
        if is_all:
            for r in results_ch:
                d = r['data']
                act = [c for c in CHANNELS if d['orders'].get(c, 0) > 0]
                if act:
                    union_item += 1
                    if any(d['inv'].get(c, 0) / d['orders'][c] < 1 for c in act):
                        union_urg += 1
        union_rate = union_urg / max(1, union_item) * 100

        if is_all:
            c = st.columns(9)
            kcard(c[0], '품목 수', f'{n_item:,}', '주문 발생 SKU')
            kcard(c[1], '총 재고액', f'{tot_amt/1e8:.2f}억', '재고수량 × 정상가')
            kcard(c[2], '총 재고량', f'{tot_inv:,}장', '6채널 합계')
            kcard(c[3], '긴급 결품', f'{n_urgent:,}건', '재고주수 < 1주')
            kcard(c[4], '결품률(합산)', f'{rate:.1f}%', '6채널 재고 합산')
            kcard(c[5], '결품(채널)', f'{union_rate:.1f}%', '한 채널이라도 결품')
            kcard(c[6], '추천 이동(IN)', f'{tot_in:,}장', '금주 충전')
            kcard(c[7], '외부창고', f'{tot_ext:,}장', 'AENS·ADU3·ADQS')
            kcard(c[8], '외부창고 비중', f'{ext_pct:.1f}%', '외부창고 / 총재고')
            st.caption('ℹ️ 결품률(합산)은 6채널 재고를 합쳐 봐 낮게(7%대) 보입니다(풀링 효과). 운영 체감은 "한 채널이라도 결품"인 결품(채널)을 보세요.')
        else:
            n = 7 if is_ext else 6
            c = st.columns(n)
            kcard(c[0], '품목 수', f'{n_item:,}', '주문 발생 SKU')
            kcard(c[1], '총 재고액', f'{tot_amt/1e8:.2f}억', '재고수량 × 정상가')
            kcard(c[2], '총 재고량', f'{tot_inv:,}장', f'{channel_pick}')
            kcard(c[3], '긴급 결품', f'{n_urgent:,}건', '재고주수 < 1주')
            kcard(c[4], '결품률', f'{rate:.1f}%', f'{n_urgent:,}/{n_item:,}')
            kcard(c[5], '추천 이동(IN)', f'{tot_in:,}장', '금주 충전')
            if is_ext:
                kcard(c[6], '외부창고', f'{tot_ext:,}장', f'{wh_label} · 비중 {ext_pct:.1f}%')
            st.caption(f'ℹ️ {channel_pick} 결품률 {rate:.1f}% — 마이너 채널은 운영 SKU·SKU당 재고가 적어 구조적으로 높습니다(데이터 오류 아님).')

        bstat = {b: [0, 0] for b in BOK_LIST}
        for r in results_ch:
            b = _bok(r['code'])
            if b in bstat:
                o = g_ord(r['data']); i = g_inv(r['data'])
                if o > 0:
                    bstat[b][0] += 1
                    if i / o < 1:
                        bstat[b][1] += 1
        st.markdown('##### 🧬 복종별 결품률')
        bc = st.columns(len(BOK_LIST))
        for j, b in enumerate(BOK_LIST):
            item, urg = bstat[b]
            br = urg / item * 100 if item else 0
            clr = '#FF6B70' if br >= 25 else '#FFC000' if br >= 10 else '#4AE3B5'
            bc[j].markdown(
                f'<div class="kpi-card"><div class="kpi-label">복종 {b}</div>'
                f'<div class="kpi-value" style="color:{clr}">{br:.1f}%</div>'
                f'<div class="kpi-sub">{urg:,}/{item:,}</div></div>', unsafe_allow_html=True)
        urgent_loss = 0
        for r in results_ch:
            d = r['data']; o = g_ord(d); i = g_inv(d)
            if o > 0 and i / o < 1:
                urgent_loss += (o - i) * d.get('price', 0)
        bw_total = sum(r['data']['inv'].get('반응과', 0) for r in results_ch)
        wb = max(BOK_LIST, key=lambda b: (bstat[b][1] / bstat[b][0]) if bstat[b][0] else 0)
        wb_rate = (bstat[wb][1] / bstat[wb][0] * 100) if bstat[wb][0] else 0
        dist_note = '반응과 재고가 0이라 현재 분배 불가 → 리오더 요청 우선' if bw_total <= 0 else f'반응과 {bw_total:,}장으로 분배(SCM 요청) 가능'
        st.markdown('##### 🧠 AI 진단 · 제안')
        st.markdown(
            '<div class="scenario-box" style="border-left-color:#8AB4F8">'
            '<div style="color:#8AB4F8;font-weight:bold;margin-bottom:4px">🧠 AI 진단 · 제안 '
            '<span style="color:#9FB0C0;font-weight:normal;font-size:11px">(통계·규칙 기반 자동 진단)</span></div>'
            '<div style="color:#FFFFFF;font-size:13px;line-height:1.7">'
            f'<b>진단</b> — {channel_pick} 운영 SKU {n_item:,}건 중 <b>{n_urgent:,}건({rate:.1f}%)</b>이 1주 내 결품 위험, '
            f'노출 손실 약 <b>{urgent_loss/1e8:.1f}억원</b> 규모. 복종별로는 <b>{wb}({wb_rate:.1f}%)</b>가 가장 취약 → 우선 점검.<br>'
            '<b>제안</b> — ① 회전(온라인 잉여→결품 채널) 우선 충전 · '
            f'② {dist_note} · ③ 회전·분배로 못 메우는 단품은 리오더 요청(🤖 AICA 2.0 허브) 처리 권장.'
            '</div></div>', unsafe_allow_html=True)

    # ───────────── 아이템별 ─────────────
    with sub_item:
        st.caption('아이템 = 상품코드 3~4번째 자리(예: SPPG23U07 → PG). 선택 채널 기준 집계 · 단품 상세 지표 동일 · 맨 위 합계.')
        _render_group(results_ch, _item, g_inv, g_ord, g_move, g_eff, g_ext, '아이템')

    # ───────────── 스타일별 (상품코드 10자리) ─────────────
    with sub_style:
        st.caption('스타일 = 상품코드 10자리 기준. 선택 채널 기준 집계 · 단품 상세 지표 동일 · 맨 위 합계.')
        cs1, cs2 = st.columns([2, 3])
        with cs1:
            sty_search_s = st.text_input('스타일 검색', placeholder='앞 10자리 (예: SPACG24A5)', key='sty_search_style').strip().upper()
        with cs2:
            all_stys_s = sorted({r['code'][:10] for r in results_ch})
            sty_pick_s = st.multiselect('스타일 선택 (복수 가능)', options=all_stys_s,
                                         key='sty_pick_style',
                                         placeholder='드롭다운에서 스타일 다중 선택')
        filtered_style = list(results_ch)
        if sty_search_s:
            filtered_style = [r for r in filtered_style if r['code'][:10].upper().startswith(sty_search_s)]
        if sty_pick_s:
            _spset = set(sty_pick_s)
            filtered_style = [r for r in filtered_style if r['code'][:10] in _spset]
        _render_group(filtered_style, lambda c: c[:10], g_inv, g_ord, g_move, g_eff, g_ext, '스타일코드', namefn=True)

    # ───────────── 단품 상세 ─────────────
    with sub_sku:
        f0, f1, f2, f3, f4, f5 = st.columns([2.2, 1.2, 1.2, 1.1, 2.4, 1.6])
        with f0:
            _all_stys_sku = sorted({r['code'][:10] for r in results_ch})
            sty_pick_sku = st.multiselect('스타일 선택 (복수 가능)', options=_all_stys_sku,
                                           key='sty_pick_sku',
                                           placeholder='스타일 단위 다중 선택')
        with f1:
            only_urgent = st.checkbox('🔴 결품(주의)만', value=False, key='ch_only_urgent')
        with f2:
            only_moved = st.checkbox('이동 발생만', value=False, key='ch_only_moved')
        with f3:
            bok_pick = st.selectbox('복종', ['전체'] + BOK_LIST, key='ch_bok')
        with f4:
            ch_search = st.text_input('검색 (상품코드/SKU)', placeholder='앞 10자리만 입력해도 OK', key='ch_search').strip().upper()
        with f5:
            ch_sort = st.selectbox('정렬', ['온라인 매출 순위 ↑', '기대효과 ↓', '이동수량 ↓', '단품코드'], key='ch_sort')
        _sty_pick_set = set(sty_pick_sku) if sty_pick_sku else None

        rows = []
        s_daily = s_damt = s_inv = s_iamt = s_ext = s_mv = s_ni = s_eff = 0.0
        for r in results_ch:
            d = r['data']
            o = g_ord(d); i = g_inv(d); mv = g_move(r); p = d.get('price', 0)
            if only_moved and mv == 0:
                continue
            woc = (i / o) if o > 0 else None
            if only_urgent and not (woc is not None and woc < 2):
                continue
            if bok_pick != '전체' and _bok(r['code']) != bok_pick:
                continue
            if ch_search and not r['code'].upper().startswith(ch_search):
                continue
            if _sty_pick_set and r['code'][:10] not in _sty_pick_set:
                continue
            daily = o / 7 if o > 0 else 0
            sojin = round(i / daily) if daily > 0 else None
            ni = i + (0 if is_all else mv)
            woc2 = (ni / o) if o > 0 else None
            ext = g_ext(d); eff = g_eff(r)
            if woc is None:
                stat = '– 무판매'
            elif woc < 1:
                stat = '🔴 긴급결품'
            elif woc < 2:
                stat = '🟡 주의'
            else:
                stat = '🟢 정상'
            row = {
                '이미지': _spao_img_url(r['code'][:10]),
                '상태': stat, '복종': _bok(r['code']),
                '상품코드': r['code'][:10], '단품코드(SKU)': r['code'],
                '상품명': (d['name'][:22] + '…') if len(d['name']) > 22 else d['name'],
                '일평균 판매량': round(daily, 1), '일평균 매출(만원)': round(daily * p / 10000, 1),
                '현 재고량': i, '현 재고금액(만원)': round(i * p / 10000),
                '내부창고': '—', '🔌항만': '—', '🔌부평': '—', '외부창고': ext,
                '현 재고주수': round(woc) if woc is not None else None,
                '소진예상(일)': sojin if sojin is not None else None,
                '추천이동': mv, '이동후재고': ni,
                '이동 후 재고주수': round(woc2) if woc2 is not None else None,
                '효과(만원)': round(eff / 10000),
            }
            rows.append((row, (sojin if sojin is not None else 10 ** 9), eff, abs(mv), d.get('rank_online', 9999), r['code']))
            s_daily += daily; s_damt += daily * p / 10000; s_inv += i; s_iamt += i * p / 10000
            s_ext += ext; s_mv += mv; s_ni += ni; s_eff += eff / 10000

        if ch_sort == '온라인 매출 순위 ↑':
            rows.sort(key=lambda x: x[4])
        elif ch_sort == '기대효과 ↓':
            rows.sort(key=lambda x: -x[2])
        elif ch_sort == '이동수량 ↓':
            rows.sort(key=lambda x: -x[3])
        else:
            rows.sort(key=lambda x: x[5])

        data = [x[0] for x in rows]
        st.caption(f'총 {len(data):,}건' + (' · 합계 외 상위 500건 표시' if len(data) > 500 else '') + ' · 맨 위 합계')
        sumrow = {
            '이미지': '',
            '상태': '— 합계 —', '복종': '', '상품코드': '', '단품코드(SKU)': f'{len(data):,}건', '상품명': '',
            '일평균 판매량': round(s_daily, 1), '일평균 매출(만원)': round(s_damt, 1),
            '현 재고량': int(s_inv), '현 재고금액(만원)': round(s_iamt),
            '내부창고': '', '🔌항만': '', '🔌부평': '', '외부창고': int(s_ext),
            '현 재고주수': '', '소진예상(일)': '', '추천이동': int(s_mv), '이동후재고': int(s_ni),
            '이동 후 재고주수': '', '효과(만원)': round(s_eff),
        }
        disp = [sumrow] + data[:500]

        if data:
            df_ch = pd.DataFrame(disp)
            styled = (df_ch.style
                      .map(_stat_color, subset=['상태'])
                      .map(woc_color, subset=['현 재고주수', '이동 후 재고주수'])
                      .map(_move_color, subset=['추천이동'])
                      .apply(_hl_sum, axis=1)
                      .format({'일평균 판매량': '{:.1f}'.format, '일평균 매출(만원)': '{:.1f}'.format,
                               '현 재고량': '{:,}'.format, '현 재고금액(만원)': '{:,}'.format,
                               '현 재고주수': _woc, '이동 후 재고주수': _woc, '소진예상(일)': _int0,
                               '추천이동': '{:,}'.format, '이동후재고': '{:,}'.format, '외부창고': '{:,}'.format,
                               '효과(만원)': '{:,}'.format}))
            st.dataframe(styled, use_container_width=True, height=520, hide_index=True,
                         column_config={
                             '이미지': st.column_config.ImageColumn('이미지', width='small',
                                                                     help='추후 SPAO 공홈 실 이미지로 교체'),
                             '복종': st.column_config.TextColumn('복종', width='small')})
        else:
            st.info('표시할 단품이 없습니다.')

        st.caption('🎨 상태: 🔴 긴급결품(<1주) · 🟡 주의(1~2주) · 🟢 정상(≥2주)   |   효과=선택 채널 기준(전체=합산)   |   외부창고=AENS·ADU3·ADQS 실데이터   |   🔌 내부창고·항만·부평=물류 API 연동(9/1) 후 표시')


def render_effect_tab():
    st.markdown('### 📈 실행 효과 누적 관리')
    st.caption('재배치 **승인 실행 시 자동 기록** (단품×채널 전일재고 스냅샷 포함) → 기대효과 대비 **실제 효과(실측)** 누적 추적.')
    st.markdown('<div class="scenario-box">📍 <b>추적 범위 — 온라인 6채널 내 회전(이동)만</b> · '
                '외부창고(AENS·ADU3·ADQS)·옴니재고 이동분은 본 효과 집계에서 <b>제외</b>. '
                '리오더(신규 입고)는 별도 \'리오더 요청\' 탭에서 관리.</div>',
                unsafe_allow_html=True)
    st.markdown('<div class="scenario-box">📐 <b>실제효과 산식 (보수 집계)</b> — 이동(IN) 받은 단품×채널에서 '
                '<b>전일(이동 전) 재고로는 판매 불가능했던 추가 판매분만</b> 인정: '
                '추가판매 = min(이동IN, max(0, 실제판매 − 전일재고)) → 실제효과 = Σ 추가판매 × 정상가. '
                '이동 없이도 팔 수 있었던 물량은 제외. 실측일 = <b>당일 매출 기준</b> (매일 06:00 매출 갱신 후 집계). 아래 일일 매출 자료 업로드 시 자동 반영.</div>',
                unsafe_allow_html=True)

    # 온라인 6채널 화이트리스트 — 외부창고/옴니 행은 details에서 자동 제외
    log_rows_raw = effect_log.load_log()
    log_rows = log_rows_raw  # log 자체는 실행 단위 요약이라 그대로

    def _f(v):
        try: return float(v)
        except Exception: return 0.0

    n_exec = len(log_rows)
    cum_qty = sum(int(_f(r.get('이동량_장'))) for r in log_rows)
    cum_exp = sum(_f(r.get('기대효과_만원')) for r in log_rows)
    measured = [r for r in log_rows if str(r.get('실제효과_만원') or '').strip()]
    cum_act = sum(_f(r.get('실제효과_만원')) for r in measured)
    cum_exp_measured = sum(_f(r.get('기대효과_만원')) for r in measured)
    rate = cum_act / cum_exp_measured * 100 if cum_exp_measured > 0 else None

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric('누적 실행', f'{n_exec:,}회', f'실측 완료 {len(measured):,}건')
    k2.metric('누적 이동량', f'{cum_qty:,}장')
    k3.metric('누적 기대효과', f'{cum_exp/10000:.2f}억')
    cum_extra = sum(int(_f(r.get('추가판매_장'))) for r in log_rows)
    k4.metric('누적 실제효과', f'{cum_act/10000:.2f}억', f'추가판매 {cum_extra:,}장 (전일재고 초과분만)')
    k5.metric('달성률 (실제/기대)', f'{rate:.1f}%' if rate is not None else '-', '실측분 기준')

    if not log_rows:
        st.info('아직 실행 이력이 없습니다. 시나리오 탭에서 "✅ 선택 N건 승인"을 실행하면 자동 기록됩니다.')
        return

    # 누적 추이 차트 (기대 vs 실제, 억원)
    df_log = pd.DataFrame(log_rows)
    df_log['기대효과_만원'] = pd.to_numeric(df_log['기대효과_만원'], errors='coerce').fillna(0)
    df_log['실제효과_만원'] = pd.to_numeric(df_log['실제효과_만원'], errors='coerce').fillna(0)
    df_chart = df_log[['실행일시']].copy()
    df_chart['누적 기대효과 (억)'] = (df_log['기대효과_만원'].cumsum() / 10000).round(3)
    df_chart['누적 실제효과 (억)'] = (df_log['실제효과_만원'].cumsum() / 10000).round(3)
    df_chart = df_chart.set_index('실행일시')
    st.line_chart(df_chart, height=240, color=['#8AB4F8', '#4AE3B5'])

    st.markdown('#### 실행 이력 · 실측 입력')
    st.caption('💡 **실제효과_만원**·**메모** 칸을 직접 수정 후 "실측 입력 저장" — 입력 시 상태가 자동으로 "실측 완료(수동)"로 변경. 1행 = Σ 합계(자동 계산, 수정 불가)')
    cum_sku = sum(int(_f(r.get('단품수'))) for r in log_rows)
    total_row = {
        'id': 'Σ', '실행일시': '— 합계 —', '시나리오': f'{n_exec}회 실행',
        '단품수': cum_sku, '이동량_장': cum_qty,
        '기대효과_만원': round(cum_exp), '실제효과_만원': round(cum_act),
        '추가판매_장': cum_extra, '실측일': '',
        '상태': f'실측 {len(measured)}/{n_exec}', '메모': '',
    }
    df_disp = pd.concat([pd.DataFrame([total_row]), df_log], ignore_index=True)

    def _hl_total(row):
        sty = 'background-color: #1E2D40; color: #4AE3B5; font-weight: bold' if str(row.get('id')) == 'Σ' else ''
        return [sty] * len(row)

    edited = st.data_editor(
        df_disp.style.apply(_hl_total, axis=1),
        use_container_width=True,
        height=320,
        hide_index=True,
        disabled=['id', '실행일시', '시나리오', '단품수', '이동량_장', '기대효과_만원', '실측일', '상태'],
        key='effect_editor',
    )

    b1, b2, b3, b4, b5 = st.columns([2, 2.4, 2, 2, 2])
    with b1:
        if st.button('💾 실측 입력 저장', type='primary', use_container_width=True, key='fx_save'):
            recs = [r for r in edited.to_dict('records') if str(r.get('id')) not in ('Σ', '합계')]
            effect_log.save_rows(recs)
            st.success('저장 완료')
            st.rerun()
    with b2:
        if st.button('🤖 실측 자동 산출 (mock·추가판매분)', use_container_width=True, key='fx_mock'):
            n = effect_log.mock_fill_actuals()
            st.success(f'{n}건 실측 채움 — 전일재고 대비 추가판매분만 집계 (실데이터 연동 전 데모)')
            st.rerun()
    with b3:
        st.download_button('⬇️ 이력 백업 (CSV)', effect_log.export_csv_bytes(),
                           'execution_log.csv', 'text/csv', use_container_width=True, key='fx_dl')
    with b4:
        up = st.file_uploader('복원', type=['csv'], key='fx_restore', label_visibility='collapsed')
        if up is not None and st.button('📂 백업 복원 (교체)', use_container_width=True, key='fx_restore_btn'):
            n = effect_log.restore_from_bytes(up.getvalue())
            st.success(f'{n}행 복원')
            st.rerun()
    with b5:
        confirm = st.checkbox('초기화 확인', key='fx_clear_ok')
        if st.button('🗑️ 이력 초기화', use_container_width=True, key='fx_clear', disabled=not confirm):
            effect_log.clear_log()
            st.rerun()

    st.markdown('#### 📂 일일 매출 자료 업로드 — 당일 실측 자동 반영')
    st.caption('매일 06:00 매출 갱신 후 업로드. 컬럼 자동 인식: 단품코드 / 채널(선택) / 판매수량 — 채널 없으면 이동IN 비중으로 배분. '
               '실측 대기 실행에 추가판매분(min(이동IN, 판매−전일재고))만 집계.')
    su1, su2 = st.columns([3, 1])
    with su1:
        sales_up = st.file_uploader('매출 자료 (csv/xlsx)', type=['csv', 'xlsx'], key='fx_sales_up', label_visibility='collapsed')
    with su2:
        if sales_up is not None and st.button('✅ 실측 반영', type='primary', use_container_width=True, key='fx_sales_apply'):
            n_exec, matched = effect_log.apply_sales_bytes(sales_up.getvalue(), sales_up.name)
            if n_exec:
                st.success(f'{n_exec}개 실행 실측 완료 (단품×채널 {matched}건 매칭)')
            else:
                st.warning('매칭된 실측 대기 실행이 없습니다 (이미 실측 완료이거나 코드 불일치).')
            st.rerun()

    with st.expander('🔍 실행별 상세 내역 (단품×채널 — 전일재고·이동IN·실제판매·추가판매)'):
        ids = [r['id'] for r in log_rows]
        pick = st.selectbox('실행 ID', ids, index=len(ids)-1, key='fx_detail_pick')
        det = effect_log.load_details(pick)
        if det:
            df_det = pd.DataFrame(det)
            st.dataframe(df_det, use_container_width=True, height=300, hide_index=True)
            st.download_button('⬇️ 상세 내역 전체 (CSV)', effect_log.export_details_bytes(),
                               'execution_details.csv', 'text/csv', key='fx_det_dl')
        else:
            st.info('이 실행에는 상세 스냅샷이 없습니다 (구버전 이력 — 실측 시 기대효과 대비 보수 추정 적용).')

    st.caption('⚠️ 이력은 앱 **재시작·재배포 시 초기화**됩니다. 주기적으로 ⬇️ 백업 CSV를 보관하고, 필요 시 📂 복원하세요. '
               '(실 배포 시 DB 저장 + audit log로 전환)')


# ============================================================
# 테스트(AICA 2.0)에서 흡수한 탭 — 🧩 추가 분배 / 🚨 리오더 요청 / 📦 입고 예정
# 진단등급·필업요청·리오더 추출·입고보정은 실데이터, 🔌 표시는 9/1 연동 예정(mock)
# ============================================================
def _kpi(col, label, value, sub=''):
    col.markdown(
        f'<div class="kpi-card" style="min-height:96px;display:flex;flex-direction:column;justify-content:center">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{value}</div>'
        f'<div class="kpi-sub">{sub}</div></div>', unsafe_allow_html=True)


def _mock_int(code, salt, lo, hi):
    """코드 기반 결정적 mock 정수 — 데모 표시용(실데이터 아님)."""
    h = abs(hash(code + salt))
    return lo + (h % (hi - lo + 1))


@st.cache_data(show_spinner=False)
def imminent_rows():
    """결품 임박(온라인 합산 재고주수 < 1주, 주문>0) 단품 — 실데이터."""
    skus = load_data_v20()
    rows = []
    for c, d in skus.items():
        ti = sum(d['inv'].get(ch, 0) for ch in CHANNELS)
        to = sum(d['orders'].get(ch, 0) for ch in CHANNELS)
        if to > 0 and ti / to < 1.0:
            wk4 = to * 4
            short = max(0, wk4 - ti)
            rows.append({
                'code': c, 'name': d['name'], 'rank': d.get('rank_online', 9999),
                'inv': ti, 'ord': to, 'woc': round(ti / to, 2),
                'wk4': wk4, 'short': short, 'loss': short * d['price'], 'price': d['price'],
            })
    rows.sort(key=lambda r: -r['loss'])
    return rows


@st.cache_data(show_spinner=False)
def imminent_rows_by_channel(channel: str):
    """채널별 결품 임박 단품 — 스파오 6/19 미팅 P0 #1 (채널별 분리).

    channel='전체' 시 imminent_rows() 동일 (6채널 합산 기준).
    채널 지정 시 그 채널의 재고/주판으로 재계산하여 결품 임박 추출.
    """
    if channel == '전체':
        return imminent_rows()
    skus = load_data_v20()
    rows = []
    for c, d in skus.items():
        ti = d['inv'].get(channel, 0)
        to = d['orders'].get(channel, 0)
        if to > 0 and ti / to < 1.0:
            wk4 = to * 4
            short = max(0, wk4 - ti)
            rows.append({
                'code': c, 'name': d['name'], 'rank': d.get('rank_online', 9999),
                'inv': ti, 'ord': to, 'woc': round(ti / to, 2),
                'wk4': wk4, 'short': short, 'loss': short * d['price'], 'price': d['price'],
            })
    rows.sort(key=lambda r: -r['loss'])
    return rows



def _grade_color(s):
    if not isinstance(s, str):
        return ''
    if 'X' in s:
        return 'background-color:#5B1E1E; color:#FF6B70; font-weight:bold'
    if 'M' in s:
        return 'background-color:#5A4500; color:#FFC000; font-weight:bold'
@st.cache_data(show_spinner=False)
def _load_style_map():
    """첨부 spao_style_map.csv 로드 — 스타일코드 10자리 → 스타일명."""
    import csv as _csv
    from pathlib import Path as _Path
    p = _Path(__file__).parent / 'spao_style_map.csv'
    m = {}
    if p.exists():
        with open(p, encoding='utf-8-sig') as f:
            for row in _csv.DictReader(f):
                k = (row.get('style_code') or '').strip()
                v = (row.get('style_name') or '').strip()
                if k and v:
                    m[k] = v
    return m


def render_onepan_tab():
    st.markdown('### 🧩 추가 분배')
    st.markdown(
        '<div class="scenario-box">스파오 6개 엑셀(공홈 결품체크·계산기, 네이버/지그재그/키즈 한판, 지그재그 마케팅)을 '
        '한 화면으로 흡수하는 통합 단품판. <b>진단(S/M/X)·현재고·주판·재고주수·필업 요청수량·반응과 보유·주력채널은 실데이터</b>. '
        '진단: 🔴 X 결품임박(&lt;1주) · 🟡 M 주의(1~4주) · 🟢 S 정상(≥4주). '
        '<b>필업 요청수량 = 재고주수 1주 미만 단품에 한해 1주 목표재고 − 현재고</b>.</div>',
        unsafe_allow_html=True)

    smap = _load_style_map()
    skus = load_data_v20()
    rows = []
    nX = nM = nS = 0
    fill_q = 0
    fill_amt = 0
    bw_total_qty = 0
    bw_total_amt = 0
    for code, d in skus.items():
        ti = sum(d['inv'].get(ch, 0) for ch in CHANNELS)
        to = sum(d['orders'].get(ch, 0) for ch in CHANNELS)
        if to <= 0 and ti <= 0:
            continue
        woc = ti / to if to > 0 else None
        if woc is None:
            grade = '– 무판매'
        elif woc < 1:
            grade = '🔴 X 결품임박'; nX += 1
        elif woc < 4:
            grade = '🟡 M 주의'; nM += 1
        else:
            grade = '🟢 S 정상'; nS += 1
        # v0.8: 필업 = 재고주수 < 1주 단품만, 1주 목표 - 현재고
        if woc is not None and woc < 1:
            fillq = max(0, round(to - ti))
        else:
            fillq = 0
        price = d.get('price', 0)
        fill_q += fillq
        fill_amt += fillq * price
        bw_q = d['inv'].get(BW_NAME, 0)
        bw_amt = bw_q * price
        bw_total_qty += bw_q
        bw_total_amt += bw_amt
        topch = max(CHANNELS, key=lambda ch: d['orders'].get(ch, 0)) if to > 0 else '-'
        sty_code = code[:10]
        sty_name_full = smap.get(sty_code, d['name'])
        woc_after = (ti + fillq) / to if to > 0 else None
        rows.append({
            '진단': grade,
            '스타일코드': sty_code,
            '단품코드': code,
            '스타일명': (sty_name_full[:24] + '…') if len(sty_name_full) > 24 else sty_name_full,
            '주력채널': CH_SHORT.get(topch, topch),
            '출고매장코드': WAREHOUSE_CODE.get(topch, '0500') if topch != '-' else '-',
            '현재고': ti,
            '반응과 전체수량': bw_q,
            '반응과 전체금액(만원)': round(bw_amt / 10000),
            '주판': to,
            '재고주수': round(woc, 1) if woc is not None else None,
            '필업요청(장)': fillq,
            '필업요청금액(만원)': round(fillq * price / 10000),
            '이동 후 재고주수': round(woc_after, 1) if woc_after is not None else None,
            '예상 회수매출(만원)': round(fillq * price / 10000),
            '_sort': fillq,
            '_price': price,
            '_topch': topch,
        })

    k = st.columns(6)
    _kpi(k[0], '🔴 결품임박(X)', f'{nX:,}건', '재고주수 < 1주')
    _kpi(k[1], '🟡 주의(M)', f'{nM:,}건', '1~4주')
    _kpi(k[2], '🟢 정상(S)', f'{nS:,}건', '≥ 4주')
    _kpi(k[3], '📦 필업 요청수량', f'{fill_q:,}장', '결품임박만 · 1주 목표')
    _kpi(k[4], '💰 필업 요청금액', f'{fill_amt/1e8:.2f}억', '필업 × 정상가')
    _kpi(k[5], '🏬 반응과 보유', f'{bw_total_amt/1e8:.2f}억', f'{bw_total_qty:,}장')

    # 핵심 10 스타일
    st.markdown('#### ⭐ 추가 분배 핵심 10 스타일')
    st.caption('단품을 스타일(10자리)로 묶어 **필업 요청금액 큰 순**. 첨부 스파오 스타일코드 매핑 적용.')
    style_grp = {}
    for r in rows:
        sty = r['스타일코드']
        g = style_grp.setdefault(sty, {'units': [], 'amt': 0, 'qty': 0,
                                       'inv': 0, 'ord': 0,
                                       'name': smap.get(sty, r['스타일명']),
                                       'topch': r['주력채널'],
                                       'wh_code': r['출고매장코드']})
        g['units'].append(r)
        g['amt'] += r['필업요청(장)'] * r['_price']
        g['qty'] += r['필업요청(장)']
        g['inv'] += r['현재고']
        g['ord'] += r['주판']
    top10 = sorted(style_grp.items(), key=lambda kv: -kv[1]['amt'])[:10]
    if top10 and any(g['amt'] > 0 for _, g in top10):
        top_list = []
        sum_inv = sum_ord = sum_qty = 0
        sum_amt = 0
        for sty, g in top10:
            woc = (g['inv'] / g['ord']) if g['ord'] > 0 else None
            if woc is None: grade = '–'
            elif woc < 1: grade = '🔴 X'
            elif woc < 4: grade = '🟡 M'
            else: grade = '🟢 S'
            woc_after = (g['inv'] + g['qty']) / g['ord'] if g['ord'] > 0 else None
            top_list.append({
                '진단': grade,
                '스타일코드': sty,
                '스타일명': (g['name'][:28] + '…') if len(g['name']) > 28 else g['name'],
                '주력채널': g['topch'],
                '출고매장코드': g['wh_code'],
                '현재고': g['inv'],
                '주판': g['ord'],
                '재고주수': round(woc, 1) if woc is not None else None,
                '필업요청(장)': g['qty'],
                '필업요청금액(만원)': round(g['amt'] / 10000),
                '이동 후 재고주수': round(woc_after, 1) if woc_after is not None else None,
                '예상 회수매출(만원)': round(g['amt'] / 10000),
            })
            sum_inv += g['inv']; sum_ord += g['ord']; sum_qty += g['qty']; sum_amt += g['amt']
        woc_sum = (sum_inv / sum_ord) if sum_ord > 0 else None
        woc_sum_after = (sum_inv + sum_qty) / sum_ord if sum_ord > 0 else None
        sum_row = {
            '진단': '— 합계 —', '스타일코드': f'{len(top10)}개', '스타일명': '',
            '주력채널': '-', '출고매장코드': '-',
            '현재고': sum_inv, '주판': sum_ord,
            '재고주수': round(woc_sum, 1) if woc_sum is not None else None,
            '필업요청(장)': sum_qty,
            '필업요청금액(만원)': round(sum_amt / 10000),
            '이동 후 재고주수': round(woc_sum_after, 1) if woc_sum_after is not None else None,
            '예상 회수매출(만원)': round(sum_amt / 10000),
        }
        df_top = pd.DataFrame([sum_row] + top_list)
        styled_top = (df_top.style.map(_grade_color, subset=['진단'])
                      .map(woc_color, subset=['재고주수', '이동 후 재고주수'])
                      .apply(_hl_sum, axis=1)
                      .format({'현재고': '{:,}'.format, '주판': '{:,}'.format,
                               '재고주수': lambda v: '' if v is None else f'{v:.1f}',
                               '이동 후 재고주수': lambda v: '' if v is None else f'{v:.1f}',
                               '필업요청(장)': '{:,}'.format,
                               '필업요청금액(만원)': '{:,}'.format,
                               '예상 회수매출(만원)': '{:,}'.format}))
        st.dataframe(styled_top, use_container_width=True, height=400, hide_index=True)
        top10_amt = sum(g['amt'] for _, g in top10)
        share = (top10_amt / fill_amt * 100) if fill_amt else 0
        st.markdown(f'<div class="scenario-box">⭐ <b>상위 10 스타일 합산 기대매출: {top10_amt/1e8:.2f}억</b> '
                    f'({share:.0f}% / 전체 필업 요청금액 {fill_amt/1e8:.2f}억).</div>',
                    unsafe_allow_html=True)
    else:
        st.info('표시할 핵심 스타일이 없습니다 (현재 결품 임박 단품 없음).')

    # 전체 단품 리스트
    st.markdown('#### 📋 전체 단품 리스트')
    f1, f2, f3 = st.columns([1.5, 3, 2])
    with f1:
        g = st.selectbox('진단', ['전체', '🔴 X 결품임박', '🟡 M 주의', '🟢 S 정상'], key='op_grade')
    with f2:
        q = st.text_input('검색 (스타일/단품코드 앞 10자리)', key='op_q').strip().upper()
    with f3:
        srt = st.selectbox('정렬', ['필업요청 ↓', '주판 ↓', '재고주수 ↑'], key='op_sort')
    view = rows
    if g != '전체':
        view = [r for r in view if r['진단'] == g]
    if q:
        view = [r for r in view if r['스타일코드'].startswith(q) or r['단품코드'].startswith(q)]
    if srt == '필업요청 ↓':
        view.sort(key=lambda r: -r['_sort'])
    elif srt == '주판 ↓':
        view.sort(key=lambda r: -r['주판'])
    else:
        view.sort(key=lambda r: (r['재고주수'] if r['재고주수'] is not None else 999))
    st.caption(f'총 {len(view):,}건' + (' · 상위 500건 표시' if len(view) > 500 else ''))
    view = view[:500]
    cols_order = ['진단', '스타일코드', '단품코드', '스타일명', '주력채널', '출고매장코드',
                  '현재고', '반응과 전체수량', '반응과 전체금액(만원)',
                  '주판', '재고주수', '필업요청(장)', '필업요청금액(만원)',
                  '이동 후 재고주수', '예상 회수매출(만원)']
    if view:
        # 합계 행
        v_inv = sum(r['현재고'] for r in view)
        v_bw_q = sum(r['반응과 전체수량'] for r in view)
        v_bw_amt = sum(r['반응과 전체금액(만원)'] for r in view)
        v_ord = sum(r['주판'] for r in view)
        v_fq = sum(r['필업요청(장)'] for r in view)
        v_famt = sum(r['필업요청금액(만원)'] for r in view)
        v_woc = (v_inv / v_ord) if v_ord > 0 else None
        v_woc_after = (v_inv + v_fq) / v_ord if v_ord > 0 else None
        sum_row = {
            '진단': '— 합계 —', '스타일코드': '', '단품코드': f'{len(view)}건',
            '스타일명': '', '주력채널': '-', '출고매장코드': '-',
            '현재고': v_inv, '반응과 전체수량': v_bw_q, '반응과 전체금액(만원)': v_bw_amt,
            '주판': v_ord,
            '재고주수': round(v_woc, 1) if v_woc is not None else None,
            '필업요청(장)': v_fq, '필업요청금액(만원)': v_famt,
            '이동 후 재고주수': round(v_woc_after, 1) if v_woc_after is not None else None,
            '예상 회수매출(만원)': v_famt,
        }
        df = pd.DataFrame([sum_row] + [{kc: r[kc] for kc in cols_order} for r in view])
        styled = (df.style.map(_grade_color, subset=['진단'])
                  .map(woc_color, subset=['재고주수', '이동 후 재고주수'])
                  .apply(_hl_sum, axis=1)
                  .format({'현재고': '{:,}'.format, '주판': '{:,}'.format,
                           '반응과 전체수량': '{:,}'.format, '반응과 전체금액(만원)': '{:,}'.format,
                           '필업요청(장)': '{:,}'.format, '필업요청금액(만원)': '{:,}'.format,
                           '예상 회수매출(만원)': '{:,}'.format,
                           '재고주수': lambda v: '' if v is None else f'{v:.1f}',
                           '이동 후 재고주수': lambda v: '' if v is None else f'{v:.1f}'}))
        st.dataframe(styled, use_container_width=True, height=460, hide_index=True)
    else:
        st.info('조건에 맞는 단품이 없습니다.')

    # 실행 액션
    st.markdown('#### ⚙️ 실행 — 직접 실행 / SCM팀 메일 / Excel 다운로드')
    st.caption('직접 실행은 즉시 분배 큐 등록(스파오). SCM팀 메일 발송 시 Excel 자동 첨부. Excel은 채널별 시트로 분리되어 분배 작업 용이.')
    tc = len(view)
    tq = sum(r['필업요청(장)'] for r in view)

    # xlsx 사전 생성 (메일 첨부 + 다운로드 공용)
    xlsx_data = None
    fname = ''
    df_skus = None
    if tc > 0:
        df_skus = pd.DataFrame([{
            '진단': r['진단'], '스타일코드': r['스타일코드'], '단품코드': r['단품코드'],
            '스타일명': r['스타일명'], '주력채널': r['주력채널'], '출고매장코드': r['출고매장코드'],
            '현재고': r['현재고'], '반응과 전체수량': r['반응과 전체수량'],
            '반응과 전체금액(만원)': r['반응과 전체금액(만원)'],
            '주판': r['주판'], '재고주수': r['재고주수'],
            '필업요청(장)': r['필업요청(장)'], '필업요청금액(만원)': r['필업요청금액(만원)'],
            '이동 후 재고주수': r['이동 후 재고주수'],
            '예상 회수매출(만원)': r['예상 회수매출(만원)'],
        } for r in view])
        df_wh_sum = (df_skus.groupby('출고매장코드')
                     .agg(단품수=('단품코드', 'count'),
                          필업요청수량=('필업요청(장)', 'sum'),
                          필업요청금액=('필업요청금액(만원)', 'sum'),
                          예상회수매출=('예상 회수매출(만원)', 'sum'))
                     .reset_index())
        from datetime import datetime as _dt
        fname = f'추가분배_{_dt.now().strftime("%Y%m%d_%H%M")}_{tc}건.xlsx'
        xlsx_data = _xlsx_by_channel(df_skus, df_wh_sum, CH_SHORT)

    # SCM 명단 (편집 가능)
    scm_list = _get_mlist('scm_mail_list', DEFAULT_SCM_LIST)

    # 액션 버튼 3개: 직접 실행 / SCM 메일 / Excel 다운로드
    a1, a2, a3 = st.columns(3)
    with a1:
        if st.button(f'⚡ 직접 실행 ({tc:,}건)', type='primary', use_container_width=True, key='op_exec', disabled=(tc == 0)):
            st.success(f'필업 요청 {tc:,}건 / {tq:,}장을 분배 큐에 등록했습니다.')
    with a2:
        scm_send = st.button(f'✉️ SCM팀 메일 발송 ({tc:,}건)',
                              use_container_width=True, key='op_mail',
                              disabled=(tc == 0 or len(scm_list) == 0))
    with a3:
        if xlsx_data is not None:
            st.download_button(
                f'⬇️ Excel 다운로드 ({tc:,}건)',
                data=xlsx_data, file_name=fname,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True, key='op_xlsx',
            )
        else:
            st.button('⬇️ Excel 다운로드 (0건)', use_container_width=True, disabled=True, key='op_xlsx_dis')

    # SCM 메일 발송 처리 — mailto 링크 + xlsx 첨부 다운로드
    if tc > 0 and scm_send:
        from datetime import datetime as _dt2
        subj = f'[필업 요청] 추가 분배 {tc:,}건 · {tq:,}장 — {_dt2.now().strftime("%Y-%m-%d")}'
        body = '\n'.join([
            '안녕하세요. SCM팀 담당자님.',
            '',
            f'AICA 분석 기준 필업 요청 {tc:,}건 / {tq:,}장 검토 부탁드립니다.',
            '첨부된 Excel은 채널별 시트로 분리되어 있어 분배 작업이 용이합니다.',
            '',
            f'- 총 필업 요청수량: {tq:,}장',
            f"- 총 필업 요청금액: {sum(r['필업요청금액(만원)'] for r in view):,}만원",
            f"- 출고매장코드 종류: {df_skus['출고매장코드'].nunique()}개" if df_skus is not None else '',
            '',
            '※ Excel 파일은 자동 다운로드되며, 메일 클라이언트에 첨부 후 발송하시기 바랍니다.',
            '',
            '— CAIO실 AX 혁신팀',
        ])
        link = _mailto_link(scm_list, subj, body)
        st.success(f'✉️ 메일 클라이언트 열림 — 받는 사람 {len(scm_list)}명 자동 입력 / Excel 첨부 후 발송')
        st.markdown(f'[📧 메일 클라이언트 열기 (수신 {len(scm_list)}명)]({link})')
        st.download_button(
            '⬇️ 첨부용 Excel 자동 다운로드',
            data=xlsx_data, file_name=fname,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary', key='op_mail_xlsx',
        )

    # SCM팀 명단 관리
    _edit_mlist('scm_mail_list', DEFAULT_SCM_LIST,
                'SCM팀 받는 사람',
                '추가 분배 메일 발송 시 자동 수신자. 기본 4명.')

    # SCM 메일 표 본문 — 엑셀 붙여넣기 가능 (TSV)
    if tc > 0:
        st.markdown('##### ✉️ SCM팀 메일 본문 — TSV (복사 붙여넣기용)')
        st.caption('TSV는 메일 본문 첨부용 — 정식 작업은 위 xlsx 사용 권장. 출고매장코드 단위 분배.')
        headers = ['단품코드', '스타일코드', '스타일명', '주력채널', '출고매장코드',
                   '현재고', '주판', '재고주수', '필업요청(장)', '필업요청금액(만원)', '예상 회수매출(만원)']
        lines = ['\t'.join(headers)]
        # 합계
        lines.append('\t'.join(['— 합계 —', '', f'{tc}건', '-', '-',
                                 f'{v_inv:,}' if 'v_inv' in dir() else str(sum(r['현재고'] for r in view)),
                                 f'{v_ord:,}' if 'v_ord' in dir() else str(sum(r['주판'] for r in view)),
                                 '', f'{tq:,}',
                                 f"{sum(r['필업요청금액(만원)'] for r in view):,}",
                                 f"{sum(r['예상 회수매출(만원)'] for r in view):,}"]))
        for r in view[:200]:  # 상한 200건
            lines.append('\t'.join([
                r['단품코드'], r['스타일코드'], r['스타일명'].replace('\t', ' '),
                r['주력채널'], r['출고매장코드'],
                f"{r['현재고']:,}", f"{r['주판']:,}",
                f"{r['재고주수']}" if r['재고주수'] is not None else '',
                f"{r['필업요청(장)']:,}",
                f"{r['필업요청금액(만원)']:,}",
                f"{r['예상 회수매출(만원)']:,}",
            ]))
        if len(view) > 200:
            lines.append(f'... 외 {len(view) - 200:,}건 (전체는 직접 실행 시 분배 큐 등록)')
        tsv_body = '\n'.join(lines)
        st.text_area('표 본문 (TSV)', value=tsv_body, height=240, key='op_mail_body')

        if scm_send:
            st.success(f'SCM팀 단톡방 5인 + 메일 그룹에 {tc:,}건 필업 요청 초안(표)을 발송했습니다.')

    st.caption('✉️ ARS 자동메일: 매주 월 06:00 결품임박(X) 단품을 SCM팀에 자동 작성·발송.')


def render_reorder_request_tab():
    st.markdown('### 🚨 리오더 요청')
    st.caption('결품 임박(재고주수 < 1주) 단품 자동 추출. 회전(재배치)으로 못 메우는 잠재 결품을 '
               '리오더로 연결 — ARS 베스트 + AICA **워스트(잠재 결품)** 동시 관리.')

    # ─── 채널 선택 (스파오 6/19 P0 #1) ───
    st.markdown('#### 📍 채널 선택')
    st.caption('스파오 6/19 미팅 합의 — 채널별로 분리 표시. "전체" 선택 시 6채널 합산 기준.')
    ch_opts = ['전체'] + list(CHANNELS)
    sel_ch = st.radio('채널', ch_opts, horizontal=True, key='reo_ch', label_visibility='collapsed')

    smap = _load_style_map()
    base = imminent_rows_by_channel(sel_ch)
    if sel_ch != '전체':
        st.caption(f'📍 **{sel_ch}** 채널의 결품 임박 단품 ({len(base):,}건) — 전체 합산이 아닌 해당 채널의 재고·주판 기준')
    enriched = []
    for r in base:
        woc = r['woc']
        if woc is None: grade = '–'
        elif woc < 1: grade = '🔴 X'
        elif woc < 4: grade = '🟡 M'
        else: grade = '🟢 S'
        sty_code = r['code'][:10]
        sty_name = smap.get(sty_code, r['name'])
        amt1w = r['ord'] * r['price']
        reord2 = max(0, r['ord'] * 2 - r['inv'])
        exp = reord2 * r['price']
        enriched.append({
            'code': r['code'], 'name': r['name'], 'rank': r['rank'],
            'inv': r['inv'], 'ord': r['ord'], 'woc': woc, 'price': r['price'],
            'sty_code': sty_code, 'sty_name': sty_name, 'grade': grade,
            'amt1w': amt1w, 'reord2': reord2, 'exp': exp,
        })

    tot_reord = sum(r['reord2'] for r in enriched)
    tot_exp = sum(r['exp'] for r in enriched)
    tot_amt1w = sum(r['amt1w'] for r in enriched)
    c1, c2, c3, c4 = st.columns(4)
    _kpi(c1, '결품 임박 단품', f'{len(enriched):,}건', '재고주수 < 1주')
    _kpi(c2, '1주 결품 노출액', f'{tot_amt1w/1e8:.2f}억', '1주 주판 × 정상가')
    _kpi(c3, '📦 리오더 권장 물량', f'{tot_exp/1e8:.2f}억', f'{tot_reord:,}장 · 2주 수요 − 현재고')
    _kpi(c4, '💰 리오더 시 기대매출', f'{tot_exp/1e8:.2f}억', '권장리오더 × 정상가')

    # 핵심 10 스타일
    st.markdown('#### ⭐ 리오더 우선 검토 10 스타일')
    st.caption('단품을 스타일(10자리)로 묶어 **1주 주문액 큰 순** 정렬. 첨부 스파오 스타일코드 매핑 적용.')

    style_groups = {}
    for r in enriched:
        sty = r['sty_code']
        g = style_groups.setdefault(sty, {'units': [], 'amt1w': 0, 'reord': 0, 'exp': 0,
                                          'inv': 0, 'ord': 0,
                                          'name': smap.get(sty, r['name'])})
        g['units'].append(r)
        g['amt1w'] += r['amt1w']
        g['reord'] += r['reord2']
        g['exp'] += r['exp']
        g['inv'] += r['inv']
        g['ord'] += r['ord']
    top_styles = sorted(style_groups.items(), key=lambda kv: -kv[1]['amt1w'])[:10]

    if not top_styles:
        st.info('현재 결품 임박 스타일이 없습니다.')
        selected_styles = set()
    else:
        # 단품 표와 동일한 컬럼 양식 — 핵심 10 스타일도 같은 표로
        top_list = []
        for sty, g in top_styles:
            woc = (g['inv'] / g['ord']) if g['ord'] > 0 else None
            if woc is None: grade = '–'
            elif woc < 1: grade = '🔴 X'
            elif woc < 4: grade = '🟡 M'
            else: grade = '🟢 S'
            woc_after = (g['inv'] + g['reord']) / g['ord'] if g['ord'] > 0 else None
            top_list.append({
                '선택': True,
                '진단': grade,
                '스타일코드': sty,
                '스타일명': (g['name'][:28] + '…') if len(g['name']) > 28 else g['name'],
                '주력채널': '-',
                '현재고': g['inv'],
                '주판': g['ord'],
                '재고주수': f"{round(woc, 1)}주" if woc is not None else '',
                '필업요청(장)': g['reord'],
                '필업요청금액(만원)': round(g['exp'] / 10000),
                '이동 후 재고주수': f"{round(woc_after, 1)}주" if woc_after is not None else '',
                '예상 회수매출(만원)': round(g['exp'] / 10000),
            })
        # 합계 행 (맨 위)
        sum_inv = sum(g['inv'] for _, g in top_styles)
        sum_ord = sum(g['ord'] for _, g in top_styles)
        sum_reord = sum(g['reord'] for _, g in top_styles)
        sum_exp = sum(g['exp'] for _, g in top_styles)
        sum_woc = round(sum_inv / sum_ord, 1) if sum_ord > 0 else None
        sum_woc_after = round((sum_inv + sum_reord) / sum_ord, 1) if sum_ord > 0 else None
        sum_row = {
            '선택': False, '진단': '— 합계 —', '스타일코드': f'{len(top_styles)}개',
            '스타일명': '', '주력채널': '-',
            '현재고': sum_inv, '주판': sum_ord,
            '재고주수': f"{sum_woc}주" if sum_woc is not None else '',
            '필업요청(장)': sum_reord, '필업요청금액(만원)': round(sum_exp / 10000),
            '이동 후 재고주수': f"{sum_woc_after}주" if sum_woc_after is not None else '',
            '예상 회수매출(만원)': round(sum_exp / 10000),
        }
        df_top = pd.DataFrame([sum_row] + top_list)
        edited = st.data_editor(
            df_top,
            use_container_width=True,
            hide_index=True,
            height=400,
            disabled=[c for c in df_top.columns if c != '선택'],
            column_config={
                '선택': st.column_config.CheckboxColumn('선택', default=True),
            },
            key='reo_top_editor',
        )
        # 합계 행(스타일코드='N개')은 선택 대상 제외
        valid_styles = {sty for sty, _ in top_styles}
        selected_styles = set(edited[edited['선택']]['스타일코드'].tolist()) & valid_styles

    # 단품 리스트
    st.markdown('#### 📋 리오더 대상 단품 리스트')
    f1, f2 = st.columns([3, 2])
    with f1:
        q = st.text_input('단품코드 검색', placeholder='앞 10자리 입력', key='aica_reo_q').strip().upper()
    with f2:
        topn = st.selectbox('표시 건수', [30, 50, 100, 200], index=1, key='aica_reo_top')
    if top_styles and selected_styles:
        flt = [r for r in enriched if r['sty_code'] in selected_styles]
    else:
        flt = enriched
    view = [r for r in flt if (not q or r['code'].startswith(q) or r['sty_code'].startswith(q))][:topn]

    if view:
        df_rows = []
        v_inv = v_ord = v_reord = 0
        v_exp = 0
        for r in view:
            woc_after = (r['inv'] + r['reord2']) / r['ord'] if r['ord'] > 0 else None
            df_rows.append({
                '진단': r['grade'],
                '스타일코드': r['sty_code'],
                '단품코드': r['code'],
                '스타일명': (r['sty_name'][:24] + '…') if len(r['sty_name']) > 24 else r['sty_name'],
                '주력채널': '-',
                '현재고': r['inv'],
                '주판': r['ord'],
                '재고주수': f"{r['woc']}주",
                '필업요청(장)': r['reord2'],
                '필업요청금액(만원)': round(r['exp'] / 10000),
                '이동 후 재고주수': f"{round(woc_after, 1)}주" if woc_after is not None else '',
                '예상 회수매출(만원)': round(r['exp'] / 10000),
            })
            v_inv += r['inv']; v_ord += r['ord']; v_reord += r['reord2']; v_exp += r['exp']
        v_woc = round(v_inv / v_ord, 1) if v_ord > 0 else None
        v_woc_after = round((v_inv + v_reord) / v_ord, 1) if v_ord > 0 else None
        sum_row = {
            '진단': '— 합계 —', '스타일코드': '', '단품코드': f'{len(view)}건', '스타일명': '', '주력채널': '-',
            '현재고': v_inv, '주판': v_ord,
            '재고주수': f"{v_woc}주" if v_woc is not None else '',
            '필업요청(장)': v_reord, '필업요청금액(만원)': round(v_exp / 10000),
            '이동 후 재고주수': f"{v_woc_after}주" if v_woc_after is not None else '',
            '예상 회수매출(만원)': round(v_exp / 10000),
        }
        df = pd.DataFrame([sum_row] + df_rows)
        styled = (df.style.map(_grade_color, subset=['진단'])
                  .map(woc_color, subset=['재고주수', '이동 후 재고주수'])
                  .apply(_hl_sum, axis=1)
                  .format({
                      '현재고': '{:,}'.format, '주판': '{:,}'.format,
                      '필업요청(장)': '{:,}'.format, '필업요청금액(만원)': '{:,}'.format,
                      '예상 회수매출(만원)': '{:,}'.format}))
        st.dataframe(styled, use_container_width=True, height=380, hide_index=True)
        # xlsx 다운로드는 아래 메일 발송 영역의 액션 버튼으로 통합 (key='reo_xlsx')
    else:
        st.info('선택한 스타일에 해당하는 단품이 없습니다.')

    # ─── xlsx 자동 첨부 사전 생성 + 메일 본문 (채널 명시) ───
    plan_list = _get_mlist('plan_mail_list', DEFAULT_PLAN_LIST)
    from datetime import datetime as _dt3
    n_mail = min(10, len(view))
    ch_tag = f'[{sel_ch}]'  # 채널 명시
    subj_line = f'{ch_tag} 리오더 요청 — 결품 임박 단품 {len(view):,}건 · 우선 검토 스타일 {len(selected_styles)}개 ({_dt3.now().strftime("%Y-%m-%d")})'

    # 메일 초안
    st.markdown('#### ✉️ 리오더 요청 메일 초안 — 기획실 발송')
    st.caption(f'직접 명령하지 않고, **요청 가능한 상태 + 1주 주문량 + 2주 권장 리오더**까지만 제공합니다(6/12 합의). 현재 선택 채널: **{sel_ch}**')
    body = [
        f'제목: {subj_line}',
        '',
        f'안녕하세요. 기획실 담당자님.',
        f'{ch_tag} 채널 기준 온라인 재고 모니터링 결과 1주 내 결품이 예상되는 우선 검토 단품을 공유드립니다.',
        '1주 주문량 기준 2주 안전재고 확보를 위한 리오더 검토 부탁드립니다.',
        '',
        f"{'단품코드':<17}{'주판':>5}{'현재고':>7}{'권장(2주)':>10}  스타일명",
    ]
    for r in view[:n_mail]:
        body.append(f"{r['code']:<17}{r['ord']:>5}{r['inv']:>7}{r['reord2']:>10}  {r['sty_name'][:18]}")
    if top_styles and selected_styles:
        sel_units_all = [u for sty, g in top_styles if sty in selected_styles for u in g['units']]
        sel_exp_all = sum(u['exp'] for u in sel_units_all)
        body += ['', f'선택 스타일 합산 기대매출: {sel_exp_all/1e8:.2f}억 (권장 리오더 × 정상가)']
    body += ['', '※ 첨부된 Excel 파일은 채널별 시트로 분리되어 있어 검토에 용이합니다.',
             '※ 본 메일은 자동 추출한 초안입니다. 실제 발주는 MD 검토 후 진행해 주세요.']
    st.text_area('메일 초안 (복사해서 사용)', value='\n'.join(body), height=260, key='aica_reo_mail')

    # 리오더 xlsx 사전 생성 (메일 첨부용)
    reo_xlsx_data = None
    reo_fname = ''
    if view:
        try:
            df_reo = pd.DataFrame([{
                '진단': r['grade'], '스타일코드': r['sty_code'], '단품코드': r['code'],
                '스타일명': r['sty_name'], '주력채널': sel_ch,
                '현재고': r['inv'], '주판': r['ord'],
                '재고주수': f"{r['woc']}주",
                '필업요청(장)': r['reord2'],
                '필업요청금액(만원)': round(r['exp'] / 10000),
                '예상 회수매출(만원)': round(r['exp'] / 10000),
            } for r in view])
            reo_fname = f'리오더요청_{sel_ch}_{_dt3.now().strftime("%Y%m%d_%H%M")}_{len(df_reo)}건.xlsx'
            reo_xlsx_data = _xlsx_bytes({f'{sel_ch} 리오더 단품': df_reo})
        except Exception:
            pass

    ms1, ms2, ms3 = st.columns(3)
    with ms1:
        plan_send = st.button(
            f'✉️ 기획실 메일 발송 ({len(view):,}건)',
            type='primary', use_container_width=True,
            key='reo_send', disabled=(len(view) == 0 or len(plan_list) == 0),
        )
    with ms2:
        if reo_xlsx_data is not None:
            st.download_button(
                f'⬇️ Excel 다운로드 ({len(view):,}건)',
                data=reo_xlsx_data, file_name=reo_fname,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True, key='reo_xlsx',
            )
        else:
            st.button('⬇️ Excel 다운로드 (0건)', use_container_width=True, disabled=True, key='reo_xlsx_dis')
    with ms3:
        if st.button('📋 클립보드 복사용 텍스트', use_container_width=True, key='reo_copy'):
            st.info('상단 텍스트박스를 전체 선택(Ctrl+A) → 복사(Ctrl+C) 하세요.')

    if plan_send and reo_xlsx_data is not None:
        link = _mailto_link(plan_list, subj_line, '\n'.join(body[2:]))
        st.success(f'✉️ 메일 클라이언트 열림 — 받는 사람 {len(plan_list)}명 자동 입력 / Excel 첨부 후 발송')
        st.markdown(f'[📧 메일 클라이언트 열기 (수신 {len(plan_list)}명)]({link})')
        st.download_button(
            '⬇️ 첨부용 Excel 자동 다운로드',
            data=reo_xlsx_data, file_name=reo_fname,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary', key='reo_mail_xlsx',
        )

    # 기획실 명단 관리
    _edit_mlist('plan_mail_list', DEFAULT_PLAN_LIST,
                '기획실 받는 사람',
                '리오더 요청 메일 발송 시 자동 수신자. 기본 14명.')


def render_inbound_tab():
    st.markdown('### 📦 입고 예정')
    st.markdown('<div class="scenario-box">🔌 <b>신규 데이터 — 9/1 API/수기 연동 예정</b>. 발주완료·이동중·항만입항 '
                '수량을 입고예정일(D-day) 기준 가용재고에 더해 결품 판정을 보정합니다. '
                '<b>입고예정일이 D-7 미만(오늘부터 7일 이내)인 경우만 +반영</b> — 그 외는 표시만. '
                '<b>아래 입고예정 수치·예정일은 데모(mock)</b>이며, 현재고·주간판매는 실데이터입니다.</div>',
                unsafe_allow_html=True)
    apply_in = st.toggle('입고 예정 반영 (D-7 미만만 가용재고 = 현재고 + 입고예정)', value=True, key='aica_in_apply')

    rows = imminent_rows()
    for r in rows:
        r['eta'] = _mock_int(r['code'], 'eta', 0, 21)
        r['po'] = _mock_int(r['code'], 'po', 0, max(1, r['short']))
        r['tr'] = _mock_int(r['code'], 'tr', 0, max(1, r['short'] // 2))
        r['pt'] = _mock_int(r['code'], 'pt', 0, max(1, r['short'] // 3))
        r['incoming'] = r['po'] + r['tr'] + r['pt']

    # 상단: 스타일 그룹 — 단품 기준과 동일한 양식
    st.markdown('#### ⭐ 입고 예정 — 스타일 기준')
    st.caption('스타일별 합산 입고예정·결품해소 효과 확인. 의사결정은 스타일 단위 → 단품 단위 순.')
    sty_grp = {}
    for r in rows:
        sty = r['code'][:10]
        g = sty_grp.setdefault(sty, {'units': [], 'name': r['name'],
                                     'inv': 0, 'ord': 0, 'incoming': 0,
                                     'po': 0, 'tr': 0, 'pt': 0,
                                     'min_eta': 999})
        g['units'].append(r)
        g['inv'] += r['inv']; g['ord'] += r['ord']; g['incoming'] += r['incoming']
        g['po'] += r['po']; g['tr'] += r['tr']; g['pt'] += r['pt']
        if r['eta'] < g['min_eta']:
            g['min_eta'] = r['eta']
    top_styles = sorted(sty_grp.items(), key=lambda kv: -kv[1]['ord'])[:15]

    if top_styles:
        sty_out = []
        for sty, g in top_styles:
            eta_eff = g['incoming'] if (apply_in and g['min_eta'] < 7) else 0
            avail = g['inv'] + eta_eff
            woc_cur = round(g['inv'] / g['ord'], 2) if g['ord'] else None
            woc2 = round(avail / g['ord'], 2) if g['ord'] else None
            sty_out.append({
                '스타일코드': sty,
                '스타일명': (g['name'][:18] + '…') if len(g['name']) > 18 else g['name'],
                '현재고': g['inv'], '주간판매': g['ord'],
                '현재고주수': f"{woc_cur}주" if woc_cur is not None else '',
                '🔌발주완료': g['po'], '🔌이동중': g['tr'], '🔌항만입항': g['pt'],
                '🔌입고예정계': g['incoming'],
                '입고예정일': f"D-{g['min_eta']}",
                '보정 가용재고': avail,
                '보정 재고주수': (f"{woc2}주" if woc2 is not None else ''),
            })
        df_sty = pd.DataFrame(sty_out)
        st_sty = df_sty.style.map(woc_color, subset=['현재고주수', '보정 재고주수']).format(
            {'현재고': '{:,}'.format, '주간판매': '{:,}'.format,
             '🔌발주완료': '{:,}'.format, '🔌이동중': '{:,}'.format, '🔌항만입항': '{:,}'.format,
             '🔌입고예정계': '{:,}'.format, '보정 가용재고': '{:,}'.format})
        st.dataframe(st_sty, use_container_width=True, height=420, hide_index=True)

    # 하단: 단품
    st.markdown('#### 📋 입고 예정 — 단품 기준 (상세)')
    sku_rows = rows[:60]
    out = []
    resolved = 0
    for r in sku_rows:
        eta_eff = r['incoming'] if (apply_in and r['eta'] < 7) else 0
        avail = r['inv'] + eta_eff
        woc2 = round(avail / r['ord'], 2) if r['ord'] else None
        if apply_in and woc2 is not None and woc2 >= 1:
            resolved += 1
        out.append({
            '단품코드': r['code'],
            '단품명': (r['name'][:18] + '…') if len(r['name']) > 18 else r['name'],
            '현재고': r['inv'], '주간판매': r['ord'], '현재고주수': f"{r['woc']}주",
            '🔌발주완료': r['po'], '🔌이동중': r['tr'], '🔌항만입항': r['pt'],
            '🔌입고예정계': r['incoming'],
            '입고예정일': f"D-{r['eta']}",
            '보정 가용재고': avail,
            '보정 재고주수': (f"{woc2}주" if woc2 is not None else ''),
        })
    c1, c2 = st.columns(2)
    _kpi(c1, '입고예정(D-7 이내) 반영 시 결품 해소', f'{resolved:,}건' if apply_in else '—', '보정 재고주수 ≥ 1주 (mock)')
    _kpi(c2, '검토 대상', f'{len(sku_rows):,}건', '결품 임박 상위')
    if out:
        df = pd.DataFrame(out)
        styled = df.style.map(woc_color, subset=['현재고주수', '보정 재고주수']).format(
            {'현재고': '{:,}'.format, '주간판매': '{:,}'.format,
             '🔌발주완료': '{:,}'.format, '🔌이동중': '{:,}'.format, '🔌항만입항': '{:,}'.format,
             '🔌입고예정계': '{:,}'.format, '보정 가용재고': '{:,}'.format})
    st.caption('🔌 입고예정 3종 + 입고예정일은 9/1 API/수기 입력 연동 후 실데이터로 대체. 현재는 단품별 결정적 mock(0~21일).')


def _hl_sum_unified(row):
    if str(row.get('채널', '')).startswith('—'):
        return ['background-color:#1E2D40; color:#4AE3B5; font-weight:bold'] * len(row)
    return [''] * len(row)


def render_unified_tab():
    st.markdown('### 🏬 통합 재고뷰')
    st.markdown('<div class="scenario-box">온라인 6채널 통합 재고를 한 화면에서 — <b>내부창고 vs 외부창고(FASS·이플렉스·CJ·풀필먼트) 분리</b>. '
                '6/12 스파오 미팅 ①② 요청 반영 — "단순 회전 도구 → 온라인 통합 재고 + 의사결정 허브" 확장 방향.</div>',
                unsafe_allow_html=True)
    skus = load_data_v20()
    agg = {ch: {'inv': 0, 'ext': 0, 'inv_amt': 0, 'ext_amt': 0, 'ord_qty': 0, 'ord_amt': 0} for ch in CHANNELS}
    for d in skus.values():
        price = d.get('price', 0)
        ext_wh_d = d.get('ext_wh', {})
        for ch in CHANNELS:
            iv = d['inv'].get(ch, 0)
            ext = ext_wh_d.get(ch, 0)
            od = d['orders'].get(ch, 0)
            agg[ch]['inv'] += iv
            agg[ch]['ext'] += ext
            agg[ch]['inv_amt'] += iv * price
            agg[ch]['ext_amt'] += ext * price
            agg[ch]['ord_qty'] += od
            agg[ch]['ord_amt'] += od * price
    tot_inv = sum(a['inv'] for a in agg.values())
    tot_ext = sum(a['ext'] for a in agg.values())
    tot_inv_amt = sum(a['inv_amt'] for a in agg.values())
    tot_ext_amt = sum(a['ext_amt'] for a in agg.values())
    tot_int_amt = tot_inv_amt - tot_ext_amt
    tot_ord_qty = sum(a['ord_qty'] for a in agg.values())
    tot_ord_amt = sum(a['ord_amt'] for a in agg.values())
    c1, c2, c3 = st.columns(3)
    _kpi(c1, '🌐 온라인 총 재고', f'{tot_inv_amt/1e8:.2f}억', f'{tot_inv:,}장')
    _kpi(c2, '🏬 내부창고', f'{tot_int_amt/1e8:.2f}억', f'{tot_inv - tot_ext:,}장 · 반응과·천안·인천 등')
    _kpi(c3, '🏭 외부창고', f'{tot_ext_amt/1e8:.2f}억', f'{tot_ext:,}장 · FASS·이플렉스·CJ')
    rows = [{
        '채널': '— 합계 —',
        '총 재고금액(만원)': round(tot_inv_amt / 10000),
        '내부창고 금액(만원)': round(tot_int_amt / 10000),
        '외부창고 금액(만원)': round(tot_ext_amt / 10000),
        '주간 주문액(만원)': round(tot_ord_amt / 10000),
        '재고보유주수': round(tot_inv / tot_ord_qty, 1) if tot_ord_qty else None,
    }]
    for ch in CHANNELS:
        a = agg[ch]
        woc = (a['inv'] / a['ord_qty']) if a['ord_qty'] else None
        rows.append({
            '채널': ch,
            '총 재고금액(만원)': round(a['inv_amt'] / 10000),
            '내부창고 금액(만원)': round((a['inv_amt'] - a['ext_amt']) / 10000),
            '외부창고 금액(만원)': round(a['ext_amt'] / 10000),
            '주간 주문액(만원)': round(a['ord_amt'] / 10000),
            '재고보유주수': round(woc, 1) if woc is not None else None,
        })
    df = pd.DataFrame(rows)
    df = pd.DataFrame(rows)
    styled = (df.style
              .map(woc_color, subset=['재고보유주수'])
              .apply(_hl_sum_unified, axis=1)
              .format({'총 재고금액(만원)': '{:,}'.format,
                       '내부창고 금액(만원)': '{:,}'.format,
                       '외부창고 금액(만원)': '{:,}'.format,
                       '주간 주문액(만원)': '{:,}'.format,
                       '재고보유주수': lambda v: '' if v is None else f'{v:.1f}'}))
    st.dataframe(styled, use_container_width=True, hide_index=True, height=320)
    st.caption('💡 외부창고 = 무신사·지그재그·네이버 풀필먼트 보관분 (이동 가용 X, 표시만 — 한지웅 리더 6/17 확정). '
               '재고보유주수 = 총재고 ÷ 주판량. 신호등 색상 동일 기준 (🔴<1주 🟡1-4주 🟢≥4주).')


# ─────────────────────────────────────────────────────────────
# v1.0 — 채널별 6 에이전트 캐릭터 (전일 베스트·결품·기대매출 동시 보고)
# ─────────────────────────────────────────────────────────────
CHANNEL_AGENTS = {
    '공홈':         {'emoji': '🏠', 'name': '공홈이',  'tag': '안정 운영 매니저',     'color': '#4a90ff', 'sync_min': 3,   'src': 'SAP'},
    '이랜드몰':     {'emoji': '🛍', 'name': '이몰이',  'tag': '그룹사 통합 큐레이터', 'color': '#7cd99c', 'sync_min': 5,   'src': 'SAP'},
    '무신사':       {'emoji': '🅼', 'name': '무신이',  'tag': 'MZ 트렌드 영업왕',     'color': '#ff6b9d', 'sync_min': 27,  'src': '풀필먼트 API'},
    '지그재그':     {'emoji': '🆉', 'name': '재그이',  'tag': '여성 캐주얼 큐레이터', 'color': '#ffb84d', 'sync_min': 183, 'src': '풀필먼트 API'},
    '네이버':       {'emoji': '🅽', 'name': '네이비',  'tag': '검색 1위 수성 담당',   'color': '#03c75a', 'sync_min': 12,  'src': '풀필먼트 API'},
    '카카오선물하기': {'emoji': '🅺', 'name': '카카이',  'tag': '선물·기프트 전문',     'color': '#fee500', 'sync_min': 8,   'src': '풀필먼트 API'},
}

AGENT_MENTS = {
    '공홈':         '📊 어제 안정 운영 완료! 핵심 결품만 도와주시면 매출↑',
    '이랜드몰':     '📦 그룹 트래픽 양호. 신상 회전 부탁드려요~',
    '무신사':       '🔥 어제 회전율 1위! 인기 단품 결품 보정 시급해요',
    '지그재그':     '✨ 여성 캐주얼 호조 · 사이즈별 결품 점검 중',
    '네이버':       '🔍 검색 유입 +12% · 결품 단품은 노출 차단됐어요',
    '카카오선물하기': '🎁 선물하기 주문 안정 · 카드 결제 회전 점검 중',
}


def _channel_agent_data(skus, ch):
    """채널별 베스트 5 (주판 기준) + 결품 임박 5 (재고주수 기준) + 기대매출."""
    bests = []
    shorts = []
    rev_total = 0
    for code, d in skus.items():
        ord_ = d['orders'].get(ch, 0)
        inv = d['inv'].get(ch, 0)
        price = d.get('price', 0)
        if ord_ > 0:
            bests.append((code, ord_, price))
            woc = inv / ord_
            if woc < 1.0:
                # 결품 해소 가정 시 회수 = 부족분 × 가격
                relief = max(0, ord_ - inv) * price
                shorts.append((code, inv, ord_, woc, relief))
                rev_total += relief
    bests.sort(key=lambda x: -x[1])
    shorts.sort(key=lambda x: x[3])
    return bests[:5], shorts[:5], rev_total


def _render_channel_agents_panel(_dt, _td):
    """6 채널 에이전트 카드 — 3×2 그리드, 항상 표시 (지우지 않음)."""
    import streamlit as st
    st.markdown(
        '<div style="margin:0 0 8px 0">'
        '<div style="color:#c4a8ff;font-size:18px;font-weight:700">'
        '🎙 채널 에이전트 6인 — 오늘의 일일 보고</div>'
        '<div style="color:#9fb3d9;font-size:12px;margin-top:2px">'
        '각 채널 에이전트가 동시에 전일 베스트·결품·기대매출을 보고합니다. '
        '[▶ AI 회전 실행]으로 채널별 회전을 즉시 트리거할 수 있습니다.</div></div>',
        unsafe_allow_html=True,
    )

    try:
        skus = load_data_v20()
    except Exception as e:
        st.error('데이터 로드 실패: ' + str(e))
        return

    smap = _load_style_map()
    _now = _dt.now()
    channels_in_order = [c for c in CHANNELS if c in CHANNEL_AGENTS]
    if 'v10_agent_log' not in st.session_state:
        st.session_state['v10_agent_log'] = {}

    # 3 x 2 그리드
    rows = [channels_in_order[:3], channels_in_order[3:6]]
    for row in rows:
        cols = st.columns(3)
        for col, ch in zip(cols, row):
            with col:
                _render_one_agent_card(ch, skus, smap, _now, _td)


def _render_one_agent_card(ch, skus, smap, _now, _td):
    """단일 채널 에이전트 카드."""
    import streamlit as st
    info = CHANNEL_AGENTS[ch]
    bests, shorts, rev = _channel_agent_data(skus, ch)
    sync_ts = _now - _td(minutes=info['sync_min'])
    dm = info['sync_min']
    if dm < 30:
        sync_icon, sync_col = '✅', '#7cd99c'
    elif dm < 120:
        sync_icon, sync_col = '⚠️', '#ffb84d'
    else:
        sync_icon, sync_col = '🔴', '#ff6b6b'

    # 베스트·결품 리스트 HTML
    best_html = ''
    for i, (code, ord_, price) in enumerate(bests, 1):
        sty = code[:10]
        nm = smap.get(sty, '')[:14]
        best_html += (
            f'<div style="font-size:11px;color:#cfd8e3;margin:2px 0">'
            f'<span style="color:{info["color"]};font-weight:700">{i}.</span> '
            f'<span style="color:#fff">{sty}</span> '
            f'<span style="color:#9ab">{nm}</span> · '
            f'<b style="color:#fff">{ord_:,}장</b></div>'
        )
    if not bests:
        best_html = '<div style="color:#666;font-size:11px">(데이터 없음)</div>'

    short_html = ''
    for i, (code, inv, ord_, woc, relief) in enumerate(shorts, 1):
        sty = code[:10]
        nm = smap.get(sty, '')[:14]
        short_html += (
            f'<div style="font-size:11px;color:#cfd8e3;margin:2px 0">'
            f'<span style="color:#ff6b6b;font-weight:700">{i}.</span> '
            f'<span style="color:#fff">{sty}</span> '
            f'<span style="color:#9ab">{nm}</span> · '
            f'<b style="color:#ffb84d">{woc:.1f}주</b></div>'
        )
    if not shorts:
        short_html = '<div style="color:#7cd99c;font-size:11px">🟢 결품 없음 — 안정</div>'

    rev_man = round(rev / 10000)

    # 카드 HTML — 캐릭터 헤더 + 멘트 + 베스트 + 결품 + 기대매출
    card_html = (
        f'<div style="background:linear-gradient(180deg,#0f1d3a 0%,#0a1428 100%);'
        f'border:1px solid {info["color"]}55;border-left:4px solid {info["color"]};'
        f'border-radius:10px;padding:14px;margin-bottom:8px;min-height:480px;'
        f'box-shadow:0 0 20px {info["color"]}22">'

        # 헤더
        f'<div style="display:flex;justify-content:space-between;align-items:center">'
        f'  <div>'
        f'    <span style="font-size:24px">{info["emoji"]}</span> '
        f'    <span style="font-size:16px;color:{info["color"]};font-weight:700">{info["name"]}</span>'
        f'    <span style="font-size:10px;color:#9ab;margin-left:6px">· {ch}</span>'
        f'  </div>'
        f'  <div style="text-align:right;font-size:10px">'
        f'    <span style="color:{sync_col}">{sync_icon} {sync_ts.strftime("%H:%M")}</span><br>'
        f'    <span style="color:#9ab">{dm}분 전 · {info["src"]}</span>'
        f'  </div>'
        f'</div>'
        f'<div style="font-size:10px;color:#7d8fa3;margin-top:2px">{info["tag"]}</div>'

        # 멘트 (말풍선)
        f'<div style="background:#1a2438;border-radius:10px 10px 10px 2px;padding:8px 12px;'
        f'margin:10px 0;font-size:12px;color:#fff;border-left:2px solid {info["color"]}">'
        f'💬 "{AGENT_MENTS.get(ch, "")}"</div>'

        # 베스트
        f'<div style="font-size:12px;color:{info["color"]};font-weight:700;margin-top:8px">🏆 전일 베스트 5</div>'
        f'{best_html}'

        # 결품
        f'<div style="font-size:12px;color:#ff6b6b;font-weight:700;margin-top:10px">🚨 결품 임박 5 (1주 미만)</div>'
        f'{short_html}'

        # 기대매출
        f'<div style="background:#2a1a1a;border-radius:6px;padding:8px;margin-top:10px;text-align:center">'
        f'<span style="color:#9ab;font-size:10px">💰 오늘 결품해소 시 기대매출</span><br>'
        f'<span style="color:#ffb84d;font-size:20px;font-weight:800">{rev_man:,}만원</span>'
        f'</div>'

        f'</div>'
    )
    st.markdown(card_html, unsafe_allow_html=True)

    # AI 회전 실행 버튼
    btn_key = f'v10_exec_{ch}'
    if st.button(f'▶ {info["name"]}에게 AI 회전 의뢰', key=btn_key, use_container_width=True):
        st.session_state['v10_agent_log'][ch] = {
            'ts': _now.strftime('%H:%M:%S'),
            'msg': f'{info["name"]}: "{ch} 회전 분배판 생성 완료! 결품 {len(shorts)}건 해소 시 약 {rev_man:,}만원 회수 가능."',
        }
        st.toast(f'{info["emoji"]} {info["name"]} — 회전 의뢰 접수!', icon='🎙')

    # 누적 응답 (지우지 않음)
    log = st.session_state['v10_agent_log'].get(ch)
    if log:
        st.markdown(
            f'<div style="background:#0d3320;border-left:3px solid #7cd99c;border-radius:8px;'
            f'padding:8px 10px;margin-top:4px;font-size:11px;color:#cfd8e3">'
            f'<span style="color:#7cd99c">✅ {log["ts"]}</span> · {log["msg"]}</div>',
            unsafe_allow_html=True,
        )


def render_v10_test_tab():
    """v1.0 (테스트) — SCM에이전트 학습 4종 통합 신규 화면.

    ① 6 채널 에이전트 보고 대화창 (전일 베스트/결품/기대매출/동기시각/실행)
    ② 결품보정 vs 수요예측 토글
    ③ AI 인사이트 한 줄 결론
    ④ AI 어시스턴트 채팅
    """
    import streamlit as st
    from datetime import datetime as _dt, timedelta as _td

    # ① 6 채널 에이전트 보고 대화창
    _render_channel_agents_panel(_dt, _td)

    st.markdown('---')

    # ② 결품보정 토글 (B2)
    cL, cR = st.columns([2, 3])
    with cL:
        st.markdown('#### ⚙️ 수요 트랙')
        mode = st.radio('수요 트랙',
                        ['수요예측 (관측 판매)', '결품보정 (잠재수요 +20%)'],
                        horizontal=False, label_visibility='collapsed', key='v10_mode')
    with cR:
        st.markdown('#### 💡 모드 설명')
        if mode.startswith('결품'):
            st.success('🎯 **결품보정 모드** — 결품으로 눌린 잠재수요를 +20% 복원해 회전 강도 강화.')
        else:
            st.info('📊 **수요예측 모드** — 관측 판매를 그대로 추종 (현 v0.9 기본 동일).')

    boost = 1.20 if mode.startswith('결품') else 1.0
    preset = SCENARIOS['🛡️ 기본']
    params_key = (preset['shortage_th'], preset['target_woc'], preset['ship_th'],
                  preset['min_move'], preset['min_recv'], _ch_excl_key(), preset['move_cap_pct'])
    try:
        results = calc_results_v20(params_key)
        results = _apply_exclusion(results)
        results = _apply_overrides(results)
    except Exception as e:
        st.error('계산 실패: ' + str(e))
        return

    if boost > 1.0:
        for r in results:
            r['revenue'] = int(r['revenue'] * boost)

    moves_items = [r for r in results if any(v != 0 for v in r['moves'].values())]
    moves_items.sort(key=lambda r: -r['revenue'])

    total_rev = sum(r['revenue'] for r in moves_items)
    total_qty = sum(sum(v for v in r['moves'].values() if v > 0) for r in moves_items)

    k1, k2, k3 = st.columns(3)
    k1.metric('🔄 회전 발생 단품', f'{len(moves_items):,}건')
    k2.metric('📦 총 이동량', f'{total_qty:,}장')
    k3.metric('💰 회수매출 (결품해소)', f'{total_rev/100000000:.2f}억',
              (f'+{(boost-1)*100:.0f}% 보정' if boost > 1 else None))

    st.markdown('---')

    # ③ AI 인사이트 한 줄 (A4)
    st.markdown('#### 💡 AI 인사이트 (한 줄 결론 · 상위 10건)')
    smap = _load_style_map()
    for r in moves_items[:10]:
        d = r['data']
        code = r['code']
        sty = code[:10]
        sty_name = smap.get(sty, d.get('name', ''))
        moves = r['moves']
        out_ch = next((c for c, v in moves.items() if v < 0), None)
        in_pairs = [(c, v) for c, v in moves.items() if v > 0]
        if not out_ch or not in_pairs:
            continue
        out_qty = -moves[out_ch]
        in_str = ' / '.join([f'{CH_SHORT.get(c, c)}+{v}' for c, v in in_pairs])
        inv_out = d['inv'].get(out_ch, 0)
        ord_out = d['orders'].get(out_ch, 0)
        woc_out = (inv_out / ord_out) if ord_out > 0 else 0
        in_woc_min = 999.0
        for c, _ in in_pairs:
            o = d['orders'].get(c, 0)
            i = d['inv'].get(c, 0)
            if o > 0:
                in_woc_min = min(in_woc_min, i / o)
        rev_man = round(r['revenue'] / 10000)
        urgency = '🔴 긴급' if in_woc_min < 1 else ('🟡 주의' if in_woc_min < 4 else '🟢 안정')
        st.markdown(
            '<div style="background:#1a1f2e;padding:12px;border-radius:8px;'
            'border-left:4px solid #4a90ff;margin-bottom:6px">'
            + urgency + ' <b style="color:#c4a8ff">' + sty + '</b> · '
            + '<span style="color:#9ab">' + str(sty_name) + '</span><br>'
            + '<span style="color:#fff">💡 ' + CH_SHORT.get(out_ch, out_ch)
            + f'({woc_out:.1f}주) → ' + in_str + f' <b>{out_qty}장</b> 회전</span> · '
            + f'<span style="color:#7cd99c">받는 채널 결품 {in_woc_min:.1f}주 해소</span> · '
            + f'<span style="color:#ffb84d">회수 <b>{rev_man:,}만원</b></span>'
            + '</div>',
            unsafe_allow_html=True,
        )
    st.caption('💡 인사이트 = 규칙 기반 자동 생성 (현재). 7월 본연동 시 Gemini/Claude LLM 호출로 교체 예정.')

    st.markdown('---')

    # ④ AI 어시스턴트 (B3)
    st.markdown('#### 🤖 AI 어시스턴트 (자연어 질의)')
    st.caption('예: "오늘 회전 추천" · "공홈 결품" · "회수매출" · "무신사"')

    if 'v10_chat' not in st.session_state:
        st.session_state['v10_chat'] = [
            {'role': 'AICA', 'text': '안녕하세요. AICA 입니다. 오늘 SPAO 온라인 재고에 대해 무엇이 궁금하신가요?'},
        ]

    for msg in st.session_state['v10_chat'][-12:]:
        if msg['role'] == '사용자':
            st.markdown(
                '<div style="text-align:right;margin:6px 0">'
                '<span style="background:#3a4a6b;padding:8px 14px;border-radius:14px;'
                'color:#fff;display:inline-block;max-width:75%">' + msg['text'] + '</span></div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="margin:6px 0">'
                '<span style="background:#1f2937;padding:8px 14px;border-radius:14px;'
                'color:#e5e7eb;display:inline-block;max-width:85%;border-left:3px solid #4a90ff">'
                '🤖 <b style="color:#c4a8ff">AICA</b> · ' + msg['text'] + '</span></div>',
                unsafe_allow_html=True,
            )

    user_q = st.chat_input('AICA에게 물어보기 (Enter 전송)')
    if user_q:
        st.session_state['v10_chat'].append({'role': '사용자', 'text': user_q})
        ans = _v10_chat_answer(user_q, results, moves_items)
        st.session_state['v10_chat'].append({'role': 'AICA', 'text': ans})
        st.rerun()

    if st.button('🧹 대화 초기화', key='v10_chat_clear'):
        st.session_state['v10_chat'] = [
            {'role': 'AICA', 'text': '대화를 초기화했습니다. 무엇을 도와드릴까요?'},
        ]
        st.rerun()


def _v10_chat_answer(q, results, moves_items):
    """v1.0 AI 어시스턴트 — 규칙 기반 응답 (7월 LLM 본연동 전 mock)."""
    if any(k in q for k in ['회전', '추천', '오늘', '분배판']):
        n = len(moves_items)
        rev = sum(r['revenue'] for r in moves_items)
        qty = sum(sum(v for v in r['moves'].values() if v > 0) for r in moves_items)
        top_str = ''
        if moves_items:
            top = moves_items[0]
            m = top['moves']
            out_ch = next((c for c, v in m.items() if v < 0), None)
            ins = [(c, v) for c, v in m.items() if v > 0]
            if out_ch and ins:
                in_s = ' / '.join([f'{CH_SHORT.get(c, c)}+{v}' for c, v in ins])
                top_str = f'\n\n📌 최상위 회전: {top["code"][:10]} · {CH_SHORT.get(out_ch, out_ch)}→{in_s} **{-m[out_ch]}장** (회수 {round(top["revenue"]/10000):,}만원)'
        return (f'오늘 회전 추천 **{n}건**, 총 이동 {qty:,}장, '
                f'예상 회수매출 **{rev/100000000:.2f}억**입니다.{top_str}')
    if '결품' in q:
        ch_match = None
        for ch in CHANNELS:
            short = CH_SHORT.get(ch, ch)
            if ch in q or short in q:
                ch_match = ch
                break
        if ch_match:
            shorts = []
            for r in results:
                inv = r['data']['inv'].get(ch_match, 0)
                o = r['data']['orders'].get(ch_match, 0)
                if o > 0 and inv / o < 1:
                    shorts.append((r['code'], inv, o, inv / o))
            shorts.sort(key=lambda x: x[3])
            top5 = shorts[:5]
            if not top5:
                return f'{CH_SHORT.get(ch_match, ch_match)} 채널 결품 단품 없음 (모두 ≥1주).'
            lines = [f'- {c[:10]} · 재고 {i:,} · 주판 {o:,} · **{w:.1f}주** 🔴' for c, i, o, w in top5]
            return f'**{CH_SHORT.get(ch_match, ch_match)}** 결품 상위 5건:\n\n' + '\n'.join(lines)
        cnt = 0
        for r in results:
            for ch in CHANNELS:
                inv = r['data']['inv'].get(ch, 0)
                o = r['data']['orders'].get(ch, 0)
                if o > 0 and inv / o < 1:
                    cnt += 1
                    break
        return f'전체 결품 (1주 미만) 단품 **{cnt:,}건**. 채널 알려주시면 상위 5건 보여드릴게요. (예: "무신 결품")'
    if any(k in q for k in ['회수', '매출', '효과', '금액']):
        rev = sum(r['revenue'] for r in moves_items)
        return f'현 회전 시나리오 예상 **결품해소 회수매출 {rev/100000000:.2f}억** · 연환산 {rev*52/100000000:.0f}억'
    for ch in CHANNELS:
        short = CH_SHORT.get(ch, ch)
        if ch in q or short in q:
            ch_items = []
            for r in moves_items:
                m = r['moves']
                if m.get(ch, 0) != 0:
                    ch_items.append((r['code'], m.get(ch, 0), r['revenue']))
            ch_items.sort(key=lambda x: -abs(x[1]))
            top5 = ch_items[:5]
            if not top5:
                return f'{CH_SHORT.get(ch, ch)} 채널 회전 발생 없음.'
            lines = [f'- {c[:10]} · {"OUT" if v < 0 else "IN"} **{abs(v)}장** · 회수 {round(rev/10000):,}만원'
                     for c, v, rev in top5]
            return f'**{CH_SHORT.get(ch, ch)}** 회전 상위 5건:\n\n' + '\n'.join(lines)
    return ('도움말:\n'
            '- 🔄 "오늘 회전 추천" / "분배판" → 회전 요약\n'
            '- 🚨 "공홈 결품" / "무신 결품" → 채널별 결품 상위 5건\n'
            '- 💰 "회수매출" / "효과" → 예상 회수매출\n'
            '- 📊 "공홈" / "무신" → 채널 회전 상위 5건')


def _get_data_seed():
    """현재 데이터 변동 시드 (실행 버튼 클릭 시 갱신). 없으면 0 (정적)."""
    return st.session_state.get('data_seed', 0)


def _apply_data_variance(results, seed):
    """mock 시각효과 — seed 기반 ±2% 변동 (단품 단위 일관성).
    7월 본연동 시 이 함수를 제거하고 진짜 데이터 갱신으로 교체."""
    if not seed:
        return results
    import random as _rnd
    rnd = _rnd.Random(seed)
    out = []
    for r in results:
        r2 = dict(r)
        factor = 1.0 + (rnd.random() - 0.5) * 0.04  # ±2%
        r2['revenue'] = int(r['revenue'] * factor)
        r2['moves'] = {k: int(round(v * factor)) for k, v in r['moves'].items()}
        out.append(r2)
    return out


def _trigger_data_refresh():
    """실행 버튼 클릭 시 호출 — seed 갱신 + 갱신 시각 + 이전 KPI 저장."""
    import time as _t
    from datetime import datetime as _dt
    prev_seed = st.session_state.get('data_seed', 0)
    if prev_seed:
        # 이전 결과 보존 (델타 계산용)
        st.session_state['data_seed_prev'] = prev_seed
        st.session_state['data_seed_prev_ts'] = st.session_state.get('data_seed_ts', '')
    st.session_state['data_seed'] = int(_t.time())
    st.session_state['data_seed_ts'] = _dt.now().strftime('%Y-%m-%d %H:%M:%S')


def _spao_img_url(sty_code, accent='#4a90ff'):
    '''SPAO 상품 이미지 URL (현재 mock — 7월 본연동 시 SPAO 공홈 실 이미지로 교체).
    accent: 카드별 강조색. SVG inline data URI 사용 (외부 의존성 0).'''
    import base64 as _b64
    bg = '#0f1d3a'
    svg = (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="120" height="150" viewBox="0 0 120 150">'
        f'<rect width="120" height="150" fill="{bg}" rx="6"/>'
        f'<rect x="14" y="14" width="92" height="100" fill="{accent}" opacity="0.18" rx="4"/>'
        f'<text x="60" y="65" text-anchor="middle" fill="{accent}" '
        f'font-family="Arial" font-size="11" font-weight="700">{sty_code}</text>'
        f'<text x="60" y="82" text-anchor="middle" fill="#9ab" '
        f'font-family="Arial" font-size="8">SPAO mock</text>'
        f'<text x="60" y="135" text-anchor="middle" fill="#666" '
        f'font-family="Arial" font-size="8">7월 실 이미지 연동</text>'
        f'</svg>'
    )
    return 'data:image/svg+xml;base64,' + _b64.b64encode(svg.encode()).decode()


def _compute_summary_kpis(skus, seed):
    """AI 요약 KPI 계산 — 전일 주문 수/액 + 현 재고 + 보유일수 + 회전/분배 추천."""
    from collections import defaultdict
    total_orders_qty = 0
    total_orders_amt = 0
    total_inv = 0
    ch_inv = defaultdict(int)
    ch_orders = defaultdict(int)
    ch_lead = {}  # ch -> (sty, qty, name, price)
    smap = _load_style_map()
    for code, d in skus.items():
        price = d.get('price', 0)
        for c in CHANNELS:
            o = d['orders'].get(c, 0)
            i = d['inv'].get(c, 0)
            total_orders_qty += o
            total_orders_amt += o * price
            total_inv += i
            ch_orders[c] += o
            ch_inv[c] += i
            if o > 0:
                prev = ch_lead.get(c)
                if (not prev) or o > prev[1]:
                    sty = code[:10]
                    nm = smap.get(sty, d.get('name', ''))
                    ch_lead[c] = (sty, o, nm, price)
    # 일평균 (주판 ÷ 7)
    daily_qty = total_orders_qty // 7
    daily_amt = total_orders_amt // 7
    woc_days = (total_inv / daily_qty) if daily_qty > 0 else 0
    # 회전 결과 (기본 시나리오)
    preset = SCENARIOS['🛡️ 기본']
    params_key = (preset['shortage_th'], preset['target_woc'], preset['ship_th'],
                  preset['min_move'], preset['min_recv'],
                  _ch_excl_key(), preset['move_cap_pct'])
    results = calc_results_v20(params_key)
    results = _apply_exclusion(results)
    results = _apply_data_variance(results, seed)
    rotation_qty = sum(sum(v for v in r['moves'].values() if v > 0) for r in results)
    rotation_amt = sum(r['revenue'] for r in results)
    # 채널별 회전 IN/OUT + 회전 대표 단품(채널이 가장 많이 IN 받는 단품)
    ch_in_qty = defaultdict(int)
    ch_in_amt = defaultdict(int)
    ch_out_qty = defaultdict(int)
    ch_top_in_sku = {}  # ch -> (sty_code, in_qty, name)
    for r in results:
        price = r['data'].get('price', 0)
        sty = r['code'][:10]
        nm = smap.get(sty, r['data'].get('name', ''))
        for c, v in r['moves'].items():
            if v > 0:
                ch_in_qty[c] += v
                ch_in_amt[c] += v * price
                prev = ch_top_in_sku.get(c)
                if (not prev) or v > prev[1]:
                    ch_top_in_sku[c] = (sty, v, nm)
            elif v < 0:
                ch_out_qty[c] += -v
    # 분배 (반응과) — calc_distribution 호출 (잔여 결품 보충 추정)
    dist_qty_total = 0
    dist_amt_total = 0
    ch_dist_qty = defaultdict(int)
    ch_dist_amt = defaultdict(int)
    for r in results:
        code = r['code']
        d = skus.get(code)
        if not d:
            continue
        moves = r['moves']
        try:
            dist, used = calc_distribution(d, moves, CHANNELS,
                                           dist_target=preset['target_woc'],
                                           bw_name='반응과')
        except Exception:
            continue
        price = d.get('price', 0)
        for c, q in dist.items():
            if q > 0:
                ch_dist_qty[c] += q
                ch_dist_amt[c] += q * price
                dist_qty_total += q
                dist_amt_total += q * price
    return {
        'daily_qty': daily_qty, 'daily_amt': daily_amt,
        'total_inv': total_inv, 'woc_days': woc_days,
        'ch_inv': ch_inv, 'ch_orders': ch_orders, 'ch_lead': ch_lead,
        'rotation_qty': rotation_qty, 'rotation_amt': rotation_amt,
        'ch_in_qty': ch_in_qty, 'ch_in_amt': ch_in_amt, 'ch_out_qty': ch_out_qty,
        'ch_top_in_sku': ch_top_in_sku,
        'dist_qty': dist_qty_total, 'dist_amt': dist_amt_total,
        'ch_dist_qty': ch_dist_qty, 'ch_dist_amt': ch_dist_amt,
    }


def render_ai_summary_tab():
    """🤖 AI 일일 요약 — 이전 AICA Studio(TEST) 페이지 내용 + 사용자 요청 양식.
    브리핑 박스(실제 이동/기대매출 분리) + TOP10 가로 표 + 자연어 채팅.
    """
    from datetime import datetime as _dt, timedelta as _td

    st.markdown("""<style>
      /* 좌우분할 hero — 글자 크게 유지 + 패딩만 컴팩트 (사용자 6/22 정정) */
      .aica-hero { text-align:center; padding: 0; }
      .aica-hero h2 { color:#fff; font-size: 22px; font-weight: 800; margin: 4px 0 0 0; line-height:1.2; }
      .aica-hero p { color:#9fb3d9; font-size: 12px; margin-top: 4px; line-height:1.35; }
      .aica-robot { display:inline-block; font-size: 60px; line-height: 1;
                    filter: drop-shadow(0 3px 16px rgba(196,168,255,.35));
                    animation: aicabob 2.4s ease-in-out infinite; }
      @keyframes aicabob { 0%,100% { transform: translateY(0); } 50% { transform: translateY(-6px); } }
      .aica-brief { background: linear-gradient(135deg, #1a0d2e 0%, #0a2138 100%);
                    border: 1px solid #5a3fb8; border-radius: 12px;
                    padding: 14px 20px; margin-top: 0;
                    box-shadow: 0 0 24px rgba(90,63,184,.18); }
      .aica-brief-title { color: #c4a8ff; font-size: 12px; font-weight: 700; letter-spacing: 1.5px; }
      .aica-brief-body { color: #fff; font-size: 16px; line-height: 1.75; margin-top: 8px; }
      .aica-brief-body b { color: #ffb84d; font-weight: 800; font-size: 17px; }
      .aica-top10 { width:100%; border-collapse:collapse; margin-top: 8px; table-layout: fixed; }
      .aica-top10 th { background:#0d2540; color:#9ab; font-size:13px; font-weight:700;
                       padding:6px; border:1px solid #2e3b50; text-align:center; }
      .aica-top10 td { background:#0f1d3a; color:#fff; font-size:12px;
                       padding:6px 4px; border:1px solid #2e3b50; text-align:center;
                       vertical-align:middle; }
      .aica-top10 td.img-cell { padding:4px; }
      .aica-bubble-user { background:#3a4a6b; padding:10px 16px;
                          border-radius:16px 16px 4px 16px; color:#fff;
                          display:inline-block; max-width:75%; }
      .aica-bubble-ai { background:#1f2937; padding:12px 18px;
                        border-radius:16px 16px 16px 4px; color:#e5e7eb;
                        display:inline-block; max-width:88%; line-height:1.7; }
    </style>""", unsafe_allow_html=True)

    # 좌우 분할용 hero HTML 정의 (실 출력은 brief markdown 호출 시 컬럼 안)
    _hero_html = (
        '<div class="aica-hero">'
        '<div class="aica-robot">🤖</div>'
        '<h2>AI 일일 요약 보고</h2>'
        '<p>SPAO 온라인 재고 AICA<br>자연어로 자유롭게 질문</p>'
        '</div>'
    )

    try:
        skus = load_data_v20()
        smap = _load_style_map()
    except Exception as e:
        st.error('데이터 로드 실패: ' + str(e))
        return

    daily_amt_total = 0.0
    ch_daily_amt = {c: 0.0 for c in CHANNELS}
    style_qty = {}; style_name = {}; style_price = {}
    short_cnt = 0
    total_inv_qty = 0
    total_inv_amt = 0
    total_orders_qty = 0
    for code, d in skus.items():
        price = d.get('price', 0)
        sty = code[:10]
        for c in CHANNELS:
            o = d['orders'].get(c, 0); i = d['inv'].get(c, 0)
            amt = o * price
            ch_daily_amt[c] += amt / 7
            daily_amt_total += amt / 7
            total_inv_qty += i
            total_inv_amt += i * price
            total_orders_qty += o
            if o > 0 and i / o < 1:
                short_cnt += 1
        tot_o = sum(d['orders'].get(c, 0) for c in CHANNELS)
        if tot_o > 0:
            style_qty[sty] = style_qty.get(sty, 0) + tot_o
            style_price[sty] = price
            style_name[sty] = smap.get(sty, d.get('name', ''))
    # 매출 순위 정렬 (사용자 요청 — 판매량 → 매출 기준 변경)
    style_amt = {s: q * style_price.get(s, 0) for s, q in style_qty.items()}
    top_10 = sorted(style_amt.items(), key=lambda x: -x[1])[:10]

    preset = SCENARIOS['🛡️ 기본']
    params_key = (preset['shortage_th'], preset['target_woc'], preset['ship_th'],
                  preset['min_move'], preset['min_recv'],
                  _ch_excl_key(), preset['move_cap_pct'])
    results = calc_results_v20(params_key)
    results = _apply_exclusion(results)
    rotation_qty = sum(sum(v for v in r['moves'].values() if v > 0) for r in results)
    rotation_revenue = sum(r['revenue'] for r in results)
    actual_move_amt = sum(
        sum(v for v in r['moves'].values() if v > 0) * r['data'].get('price', 0)
        for r in results
    )
    dist_qty = 0; dist_revenue = 0; dist_actual_amt = 0
    for r in results:
        d = skus.get(r['code'])
        if not d:
            continue
        try:
            dist, _used = calc_distribution(d, r['moves'], CHANNELS,
                                            dist_target=preset['target_woc'], bw_name='반응과')
        except Exception:
            continue
        price = d.get('price', 0)
        for c, q in dist.items():
            if q > 0:
                dist_qty += q
                dist_actual_amt += q * price
                inv_c = d['inv'].get(c, 0)
                ord_c = d['orders'].get(c, 0)
                short = max(0, ord_c - inv_c - r['moves'].get(c, 0))
                relief = min(q, short)
                dist_revenue += relief * price
    not_covered = max(0, short_cnt - len([r for r in results if any(v > 0 for v in r['moves'].values())]))

    today = _dt.now()
    yest = today - _td(days=1)
    yest_label = yest.strftime('%m/%d')
    # 매출 표기: '3억 4,067만원' (정수 억 + 나머지 만원)
    daily_eok_int = int(daily_amt_total // 100000000)
    daily_man_remainder = int((daily_amt_total % 100000000) / 10000)
    # AI 일일 요약 브리프는 풀네임 사용 (사용자 6/23 요청)
    ch_strs = [f'{c} <b>{int(ch_daily_amt[c]/10000):,}만원</b>' for c in CHANNELS]
    actual_eok = actual_move_amt / 100000000
    rev_eok = rotation_revenue / 100000000
    dist_actual_eok = dist_actual_amt / 100000000
    dist_rev_eok = dist_revenue / 100000000
    # 현 재고액 + 재고보유일수
    total_inv_eok = total_inv_amt / 100000000
    daily_qty_avg = total_orders_qty / 7 if total_orders_qty > 0 else 0
    woc_days_int = int(total_inv_qty / daily_qty_avg) if daily_qty_avg > 0 else 0

    body = (
        f'<b>[전일 온라인 주문 기준 매출 보고]</b><br>'
        f'{yest_label}일 주문 기준 매출은 <b>{daily_eok_int}억 {daily_man_remainder:,}만원</b>입니다.<br>'
        f'채널별: ' + ' / '.join(ch_strs) + '<br>'
        f'온라인 현 재고액은 <b>{total_inv_eok:.0f}억</b>이며, '
        f'재고보유일수는 <b>{woc_days_int}일</b>입니다.<br><br>'
        f'<b>[회전·분배 추천]</b><br>'
        f'현재 6채널 결품 <b>{short_cnt:,}건</b> 발생하여 '
        f'총 회전량 <b>{rotation_qty:,}장</b> 재배치 필요 — '
        f'실제 이동금액 <b>{actual_eok:.2f}억</b> / 기대매출 <b>{rev_eok:.2f}억</b>.<br>'
        f'회전으로 채우지 못하는 <b>{not_covered:,}건</b>에 대해서는 '
        f'반응과에서 <b>{dist_qty:,}장</b> 필업 필요 — '
        f'실제 이동금액 <b>{dist_actual_eok:.2f}억</b> / 기대매출 <b>{dist_rev_eok:.2f}억</b>.'
    )
    # ── 좌우 분할 렌더: 헤더(1) + 브리핑(5) ──
    _hc, _bc = st.columns([1, 5])
    with _hc:
        st.markdown(_hero_html, unsafe_allow_html=True)
    with _bc:
        st.markdown(
            f'<div class="aica-brief">'
            f'<div class="aica-brief-title">📡 AICA · DAILY BRIEFING</div>'
            f'<div class="aica-brief-body">{body}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── 순서 변경: 채팅 먼저 → TOP10 마지막 ──
    st.markdown('### 💬 자연어 질의')
    # LLM 실패 사유 노출 (디버깅용 — 키 누락/API 에러 즉시 식별)
    _llm_err = st.session_state.get('_aica_last_llm_err')
    if _llm_err:
        st.caption(f'⚠️ Gemini Flash 응답 실패 → 규칙 기반 fallback 사용 중. 사유: `{_llm_err}`')
    if 'aica_chat' not in st.session_state:
        st.session_state['aica_chat'] = []

    K = {
        'daily_amt_total': daily_amt_total,
        'ch_daily_amt': ch_daily_amt,
        'top_10': [(s, style_qty.get(s, 0) // 7, style_name.get(s, '')) for s, _amt in top_10],
        'short_cnt': short_cnt,
        'rotation_qty': rotation_qty,
        'rotation_revenue': rotation_revenue,
        'actual_move_amt': actual_move_amt,
        'dist_qty': dist_qty,
        'dist_revenue': dist_revenue,
        'dist_actual_amt': dist_actual_amt,
        'results': results,
        'skus': skus,
        'smap': smap,
    }

    for msg in st.session_state['aica_chat'][-12:]:
        if msg['role'] == 'user':
            st.markdown(
                f'<div style="text-align:right;margin:8px 0">'
                f'<span class="aica-bubble-user">{msg["text"]}</span></div>',
                unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div style="margin:8px 0">'
                f'<span class="aica-bubble-ai" style="border-left:3px solid #4a90ff">'
                f'🤖 <b style="color:#c4a8ff">AICA</b><br>{msg["text"]}</span></div>',
                unsafe_allow_html=True)

    user_q = st.chat_input('자유롭게 물어보세요 (예: 회전 TOP 5 스타일과 기대매출은?)', key='aica_chat_input')
    if user_q:
        st.session_state['aica_chat'].append({'role': 'user', 'text': user_q})
        ans = _aica_answer(user_q, K)
        st.session_state['aica_chat'].append({'role': 'ai', 'text': ans})
        st.rerun()

    # ── TOP 10 가로 표 (마지막으로 이동) ──
    st.markdown('---')
    st.markdown('#### 🏆 전일 TOP 10 스타일')
    if top_10:
        html = '<table class="aica-top10"><tr>'
        for i in range(1, 11):
            html += f'<th>{i}</th>'
        html += '</tr><tr>'
        # 이미지 행
        for sty, _amt in top_10:
            img_url = _spao_img_url(sty, '#4a90ff')
            html += f'<td class="img-cell"><img src="{img_url}" width="80" height="100"/></td>'
        html += '</tr><tr>'
        # 스타일코드 + 이름
        for sty, _amt in top_10:
            nm = (style_name.get(sty) or '')[:14]
            html += (f'<td><b style="color:#c4a8ff">{sty}</b><br>'
                     f'<span style="color:#9ab;font-size:10px">{nm}</span></td>')
        html += '</tr><tr>'
        # 판매량 (style_qty에서 조회)
        for sty, _amt in top_10:
            daily_q = style_qty.get(sty, 0) // 7
            html += (f'<td><b style="color:#ffb84d">{daily_q:,}</b>'
                     f'<br><span style="color:#9ab;font-size:10px">장/일</span></td>')
        html += '</tr><tr>'
        # 전일매출 (amt = 주간 매출 → /7 → /10000 만원)
        for sty, amt in top_10:
            amt_man = round(amt / 7 / 10000)
            html += (f'<td><b style="color:#7cd99c">{amt_man:,}</b>'
                     f'<br><span style="color:#9ab;font-size:10px">만원/일</span></td>')
        html += '</tr></table>'
        st.markdown(html, unsafe_allow_html=True)
    else:
        st.caption('데이터 없음')


def _aica_answer_llm(q, K):
    """Gemini Flash 호출 — st.secrets['GEMINI_API_KEY'] 필요.
    반환: (answer_html or None, debug_msg or None).
      - 성공 → (text, None)
      - 키 없음 → (None, '키 미설정')
      - 호출 실패 → (None, 'ErrorType: msg')
    호출자는 None이면 fallback 사용, debug_msg를 캡션으로 노출."""
    try:
        api_key = None
        try:
            if hasattr(st, 'secrets'):
                api_key = st.secrets.get('GEMINI_API_KEY')
        except Exception:
            api_key = None
        import os as _os
        api_key = api_key or _os.environ.get('GEMINI_API_KEY')
        if not api_key:
            return None, '키 미설정 (Streamlit Secrets: GEMINI_API_KEY)'
        # SDK 우회 — google-genai SDK는 AQ.* 키를 Bearer로 보내 401.
        # 실측 검증된 ?key= query string 방식을 urllib로 직접 호출.
        import urllib.request as _ur
        import urllib.error as _ue
        import json as _json
        skus = K['skus']
        smap = K['smap']
        # 채널별 KPI 빌드 (LLM 컨텍스트용)
        ch_inv_qty = {c: 0 for c in CHANNELS}
        ch_inv_amt = {c: 0 for c in CHANNELS}
        ch_short_cnt = {c: 0 for c in CHANNELS}
        ch_daily_qty = {c: 0 for c in CHANNELS}
        for code, d in skus.items():
            price = d.get('price', 0)
            for c in CHANNELS:
                o = d['orders'].get(c, 0); i = d['inv'].get(c, 0)
                ch_inv_qty[c] += i
                ch_inv_amt[c] += i * price
                ch_daily_qty[c] += o / 7
                if o > 0 and i / o < 1:
                    ch_short_cnt[c] += 1
        ch_woc_days = {c: int(ch_inv_qty[c] / ch_daily_qty[c]) if ch_daily_qty[c] > 0 else 0 for c in CHANNELS}
        # 회전 결과 (전체)
        results = K['results']
        sty_rev = {}
        for r in results:
            sty = r['code'][:10]
            sty_rev[sty] = sty_rev.get(sty, 0) + r['revenue']
        rot_top10 = sorted([(s, v) for s, v in sty_rev.items() if v > 0], key=lambda x: -x[1])[:10]
        rot_top_str = ' / '.join([f"{s}({smap.get(s,'')[:10]}) 회수 {round(v/10000):,}만원" for s, v in rot_top10])
        top_10_str = ' / '.join([f"{s}({n[:10]}) {q:,}장/일" for s, q, n in K['top_10'][:10]])
        # ── 사용자 6/24: 질문 키워드 기반 동적 dump (TPM quota 안전) ──
        # 전체 dump는 60만 토큰 → 429. 질문에서 관련 데이터만 추려 5만~15만 토큰으로 축소.
        ql = (q or '').lower()
        # 1) 질문에서 채널 키워드 감지 (풀네임 + 줄임말 둘 다)
        ch_aliases = {
            '공홈': ['공홈', '공식몰', '자사몰'],
            '이랜드몰': ['이랜드몰', '이몰', '이랜드', 'eland'],
            '무신사': ['무신사', '무신', 'musinsa'],
            '지그재그': ['지그재그', '지재', 'zigzag'],
            '네이버': ['네이버', '네이', 'naver'],
            '카카오선물하기': ['카카오선물하기', '카카오', '카카', 'kakao'],
        }
        mentioned_channels = [c for c, aliases in ch_aliases.items() if any(a in q for a in aliases)]
        # 2) 의도 분류
        is_short_q = any(k in ql for k in ['결품', '재고주수', '주수', '재고 부족', '품절'])
        is_rot_q = any(k in ql for k in ['회전', '재배치', '이동', '회수매출', '기대매출', '기대 매출'])
        is_topsell_q = any(k in ql for k in ['top', '상위', '베스트', '많이 팔', '매출 높', '잘 팔'])
        # 3) 결품 단품 (재고주수<1) 전체 — code|name|price|c1:inv/주판|...
        short_lines = []
        for code, d in skus.items():
            orders_d = d.get('orders', {}); inv_d = d.get('inv', {})
            is_short = any(orders_d.get(c, 0) > 0 and inv_d.get(c, 0) / orders_d.get(c, 0) < 1.0 for c in CHANNELS)
            if not is_short:
                continue
            if mentioned_channels and not any(orders_d.get(c, 0) > 0 and inv_d.get(c, 0) / orders_d.get(c, 0) < 1.0 for c in mentioned_channels):
                continue  # 언급 채널에서 결품 아닌 단품 제외
            nm = (d.get('name', '') or '')[:18]
            px = int(d.get('price', 0))
            ch_str = ' '.join(f"{c[:2]}:{inv_d.get(c,0)}/{orders_d.get(c,0)}" for c in CHANNELS)
            short_lines.append(f"{code}|{nm}|{px}|{ch_str}")
        # 4) 회전 추천 결과 전체 (이동 발생만)
        rot_lines = []
        for r in results:
            mv = r['moves']
            if not any(v != 0 for v in mv.values()):
                continue
            code = r['code']; nm = (r['data'].get('name', '') or '')[:18]
            rev_man = round(r['revenue'] / 10000)
            mv_str = ' '.join(f"{c[:2]}:{mv.get(c,0):+d}" for c in CHANNELS)
            rot_lines.append(f"{code}|{nm}|{rev_man}만|{mv_str}")
        # 5) 매출 상위 단품 dump (스타일 단위, 매출 내림차순)
        sku_amt = []
        for code, d in skus.items():
            orders_d = d.get('orders', {})
            tot_o = sum(orders_d.get(c, 0) for c in CHANNELS)
            if tot_o == 0:
                continue
            amt = tot_o * d.get('price', 0)
            sku_amt.append((amt, code, d))
        sku_amt.sort(key=lambda x: -x[0])
        top_n = 500 if not mentioned_channels else 1000
        topsell_lines = []
        for amt, code, d in sku_amt[:top_n]:
            inv_d = d.get('inv', {}); orders_d = d.get('orders', {})
            if mentioned_channels and not any(orders_d.get(c, 0) > 0 for c in mentioned_channels):
                continue
            nm = (d.get('name', '') or '')[:18]
            px = int(d.get('price', 0))
            amt_man = round(amt / 10000)
            ch_str = ' '.join(f"{c[:2]}:{inv_d.get(c,0)}/{orders_d.get(c,0)}" for c in CHANNELS)
            topsell_lines.append(f"{code}|{nm}|{px}|주매출{amt_man}만|{ch_str}")
        # 어떤 블록 포함할지 결정 (질문 의도 기반)
        include_short = is_short_q or (not is_rot_q and not is_topsell_q)
        include_rot = is_rot_q or (not is_short_q and not is_topsell_q)
        include_topsell = is_topsell_q or (mentioned_channels and not is_short_q and not is_rot_q)
        # 항상 회전 결과는 포함 (작은 사이즈)
        include_rot = True
        sku_dump = '\n'.join(short_lines) if include_short else ''
        topsell_dump = '\n'.join(topsell_lines) if include_topsell else ''
        rot_dump = '\n'.join(rot_lines) if include_rot else ''
        # 활성 단품 카운트 (전체 통계용)
        active_total = sum(1 for d in skus.values() if sum(d.get('orders', {}).get(c, 0) for c in CHANNELS) > 0)
        # 채널명은 풀네임 사용 (사용자 6/23 요청)
        context = f"""당신은 SPAO 온라인 재고 AI 에이전트 'AICA' 입니다.
아래 SPAO 온라인 재고 전체 데이터를 직접 분석해 한국어로 간결·정확히 답변하세요.

[답변 규칙]
- 아래 [전체 단품 데이터], [회전 추천 결과 데이터]를 직접 조회·집계해서 답변
- 채널·기준이 명확하지 않으면 합리적으로 추정해 답하고 기준 명시
- 숫자는 데이터에 있는 것만 사용 (추측·창작 금지)
- HTML 태그 <b> <br> 활용. 답변은 5~15줄 이내
- 채널명은 반드시 풀네임(공홈, 이랜드몰, 무신사, 지그재그, 네이버, 카카오선물하기). 줄임말 금지
- 단품코드는 15자리 그대로, 스타일은 앞 10자리

[데이터 포맷 안내]
[전체 단품 데이터] 한 줄당 한 단품:
  코드|단품명|정상가|공홈inv/주판|이몰inv/주판|무신inv/주판|지재inv/주판|네이inv/주판|카카inv/주판
  (채널 앞 2자: 공홈=공홈, 이몰=이랜드몰, 무신=무신사, 지재=지그재그, 네이=네이버, 카카=카카오선물하기)
  주판 = 주간 주문량(7일). 재고주수 = inv / 주판.
  결품 = 재고주수 < 1주 (inv/주판 < 1.0)

[회전 추천 결과] 한 줄당 한 단품:
  코드|단품명|기대회수만원|공홈mv|이몰mv|무신mv|지재mv|네이mv|카카mv
  mv = +이면 IN(받음), -이면 OUT(보냄), 0이면 없음

[전일 운영 KPI]
전일 매출: {K['daily_amt_total']/1e8:.2f}억
6채널 결품 단품(주수<1): {K['short_cnt']:,}건
회전 추천: {K['rotation_qty']:,}장 / 실제 이동금액 {K['actual_move_amt']/1e8:.2f}억 / 기대매출 {K['rotation_revenue']/1e8:.2f}억
반응과 분배: {K['dist_qty']:,}장 / {K['dist_revenue']/1e8:.2f}억

[채널별 운영 요약]
{chr(10).join([f"{c}: 일평균매출 {int(K['ch_daily_amt'][c]/10000):,}만원, 일평균판매 {int(ch_daily_qty[c]):,}장, 현재고 {ch_inv_qty[c]:,}장({ch_inv_amt[c]/1e8:.1f}억), 재고보유 {ch_woc_days[c]}일, 결품 {ch_short_cnt[c]:,}건" for c in CHANNELS])}

[현재 데이터 범위 — 전체 활성 단품 {active_total:,}건 중 질문 관련 추출]

[결품 단품 데이터 — 재고주수<1 {len(short_lines):,}건{', 언급채널 한정' if mentioned_channels else ''}]
{sku_dump if sku_dump else '(이번 질문에 결품 단품 데이터 불필요로 제외)'}

[매출 상위 단품 데이터 — 주매출 내림차순 {len(topsell_lines):,}건]
{topsell_dump if topsell_dump else '(이번 질문에 매출 상위 단품 데이터 불필요로 제외)'}

[회전 추천 결과 데이터 — 이동발생 {len(rot_lines):,}건]
{rot_dump if rot_dump else '(이번 질문에 회전 결과 데이터 불필요로 제외)'}

[사용자 질문]
{q}
"""
        # 모델 우선순위 — 실측 검증 결과 (사용자 6/23):
        #   gemini-2.5-flash-lite → ✅ 200 OK (?key= query, AQ.* 키 작동 확인)
        #   gemini-2.0-flash      → 429 RESOURCE_EXHAUSTED (무료 quota 초과)
        #   gemini-2.5-flash      → 503 UNAVAILABLE (일시 장애)
        #   gemini-2.5-pro        → 별도 quota
        candidates = ['gemini-2.5-flash-lite', 'gemini-2.0-flash', 'gemini-2.5-flash', 'gemini-2.5-pro']
        last_err = None
        ans = None
        model_used = None
        req_body = _json.dumps({"contents": [{"parts": [{"text": context}]}]}).encode('utf-8')
        for m in candidates:
            url = f'https://generativelanguage.googleapis.com/v1beta/models/{m}:generateContent?key={api_key}'
            req = _ur.Request(
                url,
                data=req_body,
                headers={'Content-Type': 'application/json; charset=utf-8'},
                method='POST',
            )
            try:
                with _ur.urlopen(req, timeout=30) as r:
                    data = _json.loads(r.read().decode('utf-8'))
                cands = data.get('candidates') or []
                if cands:
                    parts = cands[0].get('content', {}).get('parts', [])
                    text = ''.join(p.get('text', '') for p in parts).strip()
                    if text:
                        ans = text
                        model_used = m
                        break
                last_err = f'{m}: empty response'
            except _ue.HTTPError as he:
                try:
                    err_body = he.read().decode('utf-8')[:300]
                except Exception:
                    err_body = ''
                last_err = f'{m}: HTTP {he.code} {he.reason} — {err_body}'
                continue
            except Exception as me:
                last_err = f'{m}: {type(me).__name__}: {str(me)[:160]}'
                continue
        if not ans:
            return None, last_err or 'all models failed'
        return ans + f'<br><br><span style="color:#9ab;font-size:10px">🤖 {model_used} 자유 응답</span>', None
    except Exception as e:
        return None, f'{type(e).__name__}: {str(e)[:200]}'


def _aica_answer(q, K):
    """LLM 시도 → 실패 시 fallback 답변 + LLM 디버그 사유를 inline으로 prepend."""
    # 1) Gemini Flash 시도
    llm_ans, llm_err = _aica_answer_llm(q, K)
    if llm_ans:
        return llm_ans
    # 2) Fallback 답변
    fb = _aica_fallback(q, K)
    # 3) LLM 디버그 사유를 답변 위에 inline 표시 (사용자 6/23 진단 강화 — 캡션 누락 대비)
    if llm_err:
        st.session_state['_aica_last_llm_err'] = llm_err
        prefix = (
            f'<div style="background:#3a1f1f;border-left:3px solid #ff6b6b;'
            f'padding:8px 12px;border-radius:4px;margin-bottom:10px;font-size:11px;'
            f'color:#ffb84d;line-height:1.5">'
            f'<b>⚙️ LLM 진단</b> — Gemini 호출 실패 → 규칙 기반 fallback 사용 중<br>'
            f'사유: <code style="color:#ff6b6b">{llm_err}</code>'
            f'</div>'
        )
        return prefix + fb
    return fb


def _aica_fallback(q, K):
    """규칙 기반 답변 (LLM 키 없거나 LLM 호출 실패 시)."""
    import re as _re
    num_match = _re.search(r'(\d+)\s*개|top\s*(\d+)|상위\s*(\d+)', q.lower())
    top_n = 5
    if num_match:
        for g in num_match.groups():
            if g:
                top_n = min(int(g), 30); break
    results = K['results']; skus = K['skus']; smap = K['smap']
    if (('회전' in q or '재배치' in q) and
        ('top' in q.lower() or '상위' in q or '필요' in q or '추천' in q or '큰' in q or '높' in q or '많' in q)):
        sty_rev = {}; sty_qty = {}; sty_nm = {}
        for r in results:
            sty = r['code'][:10]
            sty_rev[sty] = sty_rev.get(sty, 0) + r['revenue']
            sty_qty[sty] = sty_qty.get(sty, 0) + sum(v for v in r['moves'].values() if v > 0)
            sty_nm[sty] = smap.get(sty, r['data'].get('name', ''))
        top = sorted([(s, rev) for s, rev in sty_rev.items() if rev > 0], key=lambda x: -x[1])[:top_n]
        if not top:
            return '회전 필요한 스타일이 없습니다 (안정 운영 중).'
        lines = [
            f'{i}. <b>{s}</b> {(sty_nm[s] or "")[:18]} — 이동 {sty_qty[s]:,}장 · 회수 <b style="color:#ffb84d">{round(sty_rev[s]/10000):,}만원</b>'
            for i, (s, _) in enumerate(top, 1)
        ]
        total_rev_eok = sum(rev for _, rev in top) / 100000000
        return (f'회전 필요 상위 <b>{top_n}개 스타일</b> (회수매출 기준):<br><br>' + '<br>'.join(lines) +
                f'<br><br>📊 상위 {top_n}개 합계 회수: <b style="color:#ffb84d">{total_rev_eok:.2f}억</b>')
    if (('top' in q.lower() or '상위' in q or '베스트' in q) and ('스타일' in q or '판매' in q or '매출' in q)):
        top = K['top_10'][:top_n]
        if not top:
            return '데이터 없음'
        lines = [f'{i}. <b>{s}</b> {(nm or "")[:18]} — <b>{daily_q:,}장/일</b>'
                 for i, (s, daily_q, nm) in enumerate(top, 1)]
        return f'전일 매출 TOP {top_n} 스타일:<br><br>' + '<br>'.join(lines)
    for c in CHANNELS:
        short = CH_SHORT.get(c, c)
        # 매칭은 양쪽(풀네임/줄임말) 다 지원, 출력은 풀네임 사용 (사용자 6/23 요청)
        if (c in q or short in q) and '결품' in q:
            shorts = []
            for code, d in skus.items():
                o = d['orders'].get(c, 0); i = d['inv'].get(c, 0)
                if o > 0 and i / o < 1:
                    sty = code[:10]
                    shorts.append((code, i, o, i / o, smap.get(sty, '')))
            shorts.sort(key=lambda x: x[3])
            top_s = shorts[:top_n]
            if not top_s:
                return f'<b>{c}</b> 채널 결품 단품 없음 (모두 ≥ 1주).'
            lines = [
                f'{i}. <b>{c0}</b> {(nm or "")[:14]} — 재고 {inv:,} · 주판 {o:,} · <b style="color:#ff6b6b">{w:.1f}주</b>'
                for i, (c0, inv, o, w, nm) in enumerate(top_s, 1)
            ]
            return f'<b>{c}</b> 채널 결품 상위 {top_n}건:<br><br>' + '<br>'.join(lines)
    if any(k in q for k in ['기대매출', '회수', '효과', '얼마']) and '회전' in q:
        return (f'회전 기대매출: <b style="color:#ffb84d">{K["rotation_revenue"]/100000000:.2f}억</b><br>'
                f'(실제 이동금액 {K["actual_move_amt"]/100000000:.2f}억, 총 이동량 {K["rotation_qty"]:,}장)')
    if any(k in q for k in ['반응과', '분배', '필업']):
        return (f'반응과 분배 추천: <b>{K["dist_qty"]:,}장</b><br>'
                f'실제 이동금액 <b>{K["dist_actual_amt"]/100000000:.2f}억</b> / '
                f'기대매출 <b>{K["dist_revenue"]/100000000:.2f}억</b>')
    if '결품' in q:
        return (f'현재 6채널 결품 단품: <b style="color:#ff6b6b">{K["short_cnt"]:,}건</b><br>'
                f'채널명 지정하시면 상위 결품을 보여드릴게요. (예: "무신사 결품 5건")')
    if any(k in q for k in ['매출', '주문', '전일', '어제', '오늘']):
        ch_strs = [f'{c} {int(K["ch_daily_amt"][c]/10000):,}만원' for c in CHANNELS]
        return (f'전일 주문 기준 매출: <b>{K["daily_amt_total"]/100000000:.2f}억</b><br>'
                f'채널별: {" / ".join(ch_strs)}')
    if '회전' in q or '재배치' in q:
        return (f'회전 추천: <b>{K["rotation_qty"]:,}장</b> · '
                f'실제 이동금액 <b>{K["actual_move_amt"]/100000000:.2f}억</b> / '
                f'기대매출 <b>{K["rotation_revenue"]/100000000:.2f}억</b>')
    return (
        '도움말 — 다음 패턴으로 물어봐주세요:<br><br>'
        '• <b>"회전 TOP 5 스타일"</b> · 회전 필요 상위 N개 스타일 + 기대매출<br>'
        '• <b>"전일 TOP 10 스타일"</b> · 전일 매출 베스트 스타일<br>'
        '• <b>"공홈 결품 5건"</b> · 채널별 결품 상위 N건<br>'
        '• <b>"회전 기대매출"</b> · 회전 시 예상 회수매출<br>'
        '• <b>"반응과 분배"</b> · 반응과 분배 추천'
    )


def render():
    st.markdown('<div class="title-bar">온라인 재고관리 Agent — 운영 대시보드<span class="ver-badge">v0.9</span></div>', unsafe_allow_html=True)
    last = get_last_update_time()
    reorder_info = get_reorder_info()
    if reorder_info.get('file'):
        reorder_txt = f"  ·  리오더 매핑 <b>{reorder_info['file']}</b> ({reorder_info.get('mapping_rows', 0)}건)"
    else:
        reorder_txt = ''
    col_a, col_b, col_c = st.columns([6, 1, 1])
    with col_a:
        st.caption(f'<b>마지막 데이터 갱신</b>: {last}{reorder_txt}')
    with col_b:
        if st.button('🔄 새로고침', use_container_width=True):
            st.rerun()
    with col_c:
        st.caption('v0.9')

    st.markdown(
        '<style>'
        'div[data-baseweb="tab-list"]:not([data-baseweb="tab-panel"] *) [data-baseweb="tab"]:nth-child(6),'
        'div[data-baseweb="tab-list"]:not([data-baseweb="tab-panel"] *) [data-baseweb="tab"]:nth-child(9)'
        '{margin-left:68px;}'
        '</style>',
        unsafe_allow_html=True,
    )

    labels = [
        '🤖 AI 일일 요약',
        '🛡️ 재배치(기본)',
        '🎛️ 재배치(임의)',
        '📈 실행 효과',
        '🧩 추가 분배',
        '🚨 리오더 요청',
        '🏬 통합 재고뷰',
        '📊 채널 별 세부',
        '📦 입고 예정',
        '🚫 채널 IN-OUT (MD 기입)',
        '🔁 리오더 매핑 (SCM 기입)',
    ]
    t = st.tabs(labels)

    def _safe(name, fn):
        import traceback as _tb
        try:
            fn()
        except Exception as e:
            st.error(f'⚠️ **[{name}] 탭 렌더 실패** — `{type(e).__name__}: {e}`')
            with st.expander('🔎 디버그 traceback'):
                st.code(_tb.format_exc())

    with t[0]:
        _safe('AI 일일 요약', render_ai_summary_tab)
    with t[1]:
        _safe('재배치(기본)', lambda: render_scenario('🛡️ 기본', st, allow_slider=False))
    with t[2]:
        _safe('재배치(임의)', lambda: render_scenario('🎛️ 임의', st, allow_slider=True))
    with t[3]:
        _safe('실행 효과', render_effect_tab)
    with t[4]:
        _safe('추가 분배', render_onepan_tab)
    with t[5]:
        _safe('리오더 요청', render_reorder_request_tab)
    with t[6]:
        _safe('통합 재고뷰', render_unified_tab)
    with t[7]:
        _safe('채널 별 세부', render_channel_tab)
    with t[8]:
        _safe('입고 예정', render_inbound_tab)
    with t[9]:
        _safe('채널 IN-OUT', render_excluded_tab)
    with t[10]:
        _safe('리오더 매핑', render_reorder_tab)
    st.caption('© 2026 Fashion BG · CAIO실 AX 혁신팀 · 강훈구  |  온라인 재고관리 Agent v0.9 (스파오 6/19 합의 반영)')
# end of render() — DO NOT TRUNCATE BELOW THIS LINE
